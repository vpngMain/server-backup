"""Robustní parsování .xls a .xlsx souborů s tolerantním mapováním hlaviček.

Formát Order Confirmation (confirmation.xls / .xlsx):
Importuje se pouze těchto 16 sloupců v tomto pořadí jako v souboru:
  Description, Description 2, Pot-Size, Qty., Ordered Qty., Per Unit,
  Qty. Per Shelf, Shelf per CC, Unit per CC, Sales Price, Amount,
  EAN Code, VBN Code, Plant Passport No., Customer Line Info, Image Reference.
Řádek souhrnu na konci (prázdný Description nebo jen součty) se při importu přeskočí.
Čísla: desetinná tečka/čárka, mezera jako oddělovač tisíců (např. 3 136,32).
"""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from types import SimpleNamespace
from typing import Optional, Any

import xlrd
import openpyxl

logger = logging.getLogger(__name__)

# Pro kompatibilitu _cell_to_value (xlrd ctypes)
XL_CELL_EMPTY = 0
XL_CELL_TEXT = 1
XL_CELL_NUMBER = 2

# Sloupce z Order Confirmation (confirmation.xls) – přesně tyto hlavičky se rozpoznají
ORDER_CONFIRMATION_HEADERS = [
    "Description",
    "Description 2",
    "Pot-Size",
    "Qty.",
    "Ordered Qty.",
    "Per Unit",
    "Qty. Per Shelf",
    "Shelf per CC",
    "Unit per CC",
    "Sales Price",
    "Amount",
    "EAN Code",
    "VBN Code",
    "Plant Passport No.",
    "Customer Line Info",
    "Image Reference",
]

# Mapování: interní klíč -> možné názvy sloupců (různé zdroje mění názvy – bereme všechny varianty)
COLUMN_ALIASES: dict[str, list[str]] = {
    "description": [
        "description", "popis", "název", "name", "desc", "product", "produkt",
        "description (1)", "desc 1", "popis produktu", "product name",
    ],
    "description2": [
        "description 2", "description2", "description (2)", "popis 2", "popis2",
        "doplňkový popis", "desc 2", "subtitle", "doplněk",
    ],
    "pot_size": [
        "pot-size", "pot size", "pot_size", "potsize", "velikost květináče",
        "pot", "velikost", "size", "container", "objem",
    ],
    "qty": ["qty.", "qty", "množství", "quantity", "ks", "qty"],
    "ordered_qty": ["ordered qty.", "ordered qty", "orderedqty", "objednané množství", "ordered", "objednáno"],
    "per_unit": ["per unit", "per_unit", "perunit", "za jednotku", "jednotka", "unit"],
    "qty_per_shelf": ["qty. per shelf", "qty per shelf", "qty_per_shelf", "per shelf", "na polici"],
    "shelf_per_cc": ["shelf per cc", "shelf_per_cc", "shelf per cc", "polic na cc"],
    "unit_per_cc": ["unit per cc", "unit_per_cc", "unit per cc", "units per cc", "jednotek na cc"],
    "sales_price": [
        "sales price", "sales_price", "salesprice", "prodejní cena", "cena",
        "price", "sell price", "prodej", "cena prodej",
    ],
    "amount": ["amount", "částka", "celkem", "suma", "total", "celková částka"],
    "ean_code": ["ean code", "ean_code", "ean", "EAN Code", "ean code", "barcode"],
    "vbn_code": ["vbn code", "vbn_code", "vbn", "VBN Code", "vbn code"],
    "plant_passport_no": [
        "plant passport no.", "plant passport no", "plant passport no",
        "plant_passport_no", "Plant Passport No.", "passport", "pasport", "číslo pasu",
    ],
    "customer_line_info": ["customer line info", "customer_line_info", "Customer Line Info", "info zákazník", "poznámka řádek"],
    "image_reference": ["image reference", "image_reference", "Image Reference", "obrázek", "image", "foto"],
    "purchase_price_imported": [
        "cena + doprava", "cena doprava", "cena+doprava", "nakup s dopravou",
        "purchase", "purchase price", "nákup", "cena nákup", "cena a doprava",
    ],
    "margin_7_imported": [
        "7% marže + doprava", "7% marže doprava", "7% marže+doprava",
        "marže", "7% marže", "marže 7", "7% margin",
    ],
    "vip_eur_imported": ["vip eur", "VIP Eur", "vip_eur", "vip eur", "cena vip eur"],
    "vip_czk_imported": [
        "vip czk", "VIP CZK", "vip_czk", "vip czk", "vip", "cena czk vip",
        "cena vip czk", "vip czk",
    ],
    "trade_price_imported": [
        "cena obchod", "Cena obchod", "trade price", "obchod", "obchodní cena",
        "d1", "cena d1", "trade", "wholesale",
    ],
    "d4_price_imported": ["d4", "D4", "d4_price", "d4 price", "cena d4"],
}

# Sloupce ukládané jako číslo (Decimal)
NUMERIC_KEYS = frozenset({
    "qty", "ordered_qty", "qty_per_shelf", "shelf_per_cc", "unit_per_cc",
    "sales_price", "amount", "purchase_price_imported", "margin_7_imported",
    "vip_eur_imported", "vip_czk_imported", "trade_price_imported", "d4_price_imported",
})


@dataclass
class RowParseError:
    """Chyba parsování jednoho řádku."""
    row_index: int  # 0-based v rámci dat (první datový řádek = 0)
    message: str


@dataclass
class ParsedXlsResult:
    """Výsledek parsování jednoho .xls souboru."""
    rows: list[dict] = field(default_factory=list)
    row_errors: list[RowParseError] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    header_row_index: Optional[int] = None
    column_map: dict[str, int] = field(default_factory=dict)  # interní klíč -> index sloupce
    detected_headers: dict[str, str] = field(default_factory=dict)  # interní klíč -> původní text hlavičky v souboru


def _normalize_header(cell_value: Optional[str]) -> str:
    """Normalizuje hlavičku pro porovnání: BOM, trim, lowercase, jedna mezera, odstranění závorek a teček."""
    if cell_value is None:
        return ""
    s = str(cell_value)
    s = s.replace("\ufeff", "").strip()
    s = re.sub(r"[\x00-\x1f\x7f]", "", s)
    s = s.strip().lower()
    # Odstranit obsah v závorkách např. "Description (1)" -> "description"
    s = re.sub(r"\s*\([^)]*\)\s*", " ", s)
    # Tečky a pomlčky na mezeru (aby "Qty." = "qty" a "Pot-Size" = "pot size")
    s = s.replace(".", " ").replace("-", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _header_matches(header_norm: str, alias_norm: str) -> bool:
    """True pokud hlavička odpovídá aliasu (exact nebo začátek/slovo shoda)."""
    if not header_norm or not alias_norm:
        return False
    if header_norm == alias_norm:
        return True
    # Tolerantní: "description 1" po normalizaci; alias je obsažen v hlavičce nebo naopak
    if len(alias_norm) >= 3 and alias_norm in header_norm:
        return True
    if len(header_norm) >= 3 and header_norm in alias_norm:
        return True
    # Slovo shoda: "prodejni cena" vs "sales price" – společná slova; zde stačí obsah
    h_words = set(header_norm.split())
    a_words = set(alias_norm.split())
    if h_words & a_words and len(h_words & a_words) >= 1:
        return True
    return False


def _find_column_index(row: list, key: str) -> Optional[int]:
    """Najde index sloupce podle aliasů. Exact match má přednost, pak tolerantní."""
    aliases = COLUMN_ALIASES.get(key, [key])
    if isinstance(aliases, str):
        aliases = [aliases]
    alias_norms = [_normalize_header(a) for a in aliases]
    # 1. Exact match
    for i, cell in enumerate(row):
        norm = _normalize_header(cell)
        if norm and norm in alias_norms:
            return i
    # 2. Tolerantní (obsahuje)
    for i, cell in enumerate(row):
        norm = _normalize_header(cell)
        if not norm:
            continue
        for an in alias_norms:
            if _header_matches(norm, an):
                return i
    return None


def _find_header_row(sheet: Any) -> Optional[int]:
    """Najde index řádku s hlavičkou (obsahuje sloupec description/popis). Scan prvních 20 řádků."""
    max_scan = min(20, sheet.nrows)
    for row_idx in range(max_scan):
        row_cells = [sheet.cell_value(row_idx, c) for c in range(sheet.ncols)]
        if _find_column_index(row_cells, "description") is not None:
            return row_idx
    return None


def _cell_to_value(cell: Any, numeric: bool):
    """Převede buňku na hodnotu (str nebo float). Podporuje xlrd buňku i openpyxl wrapper (ctype 0/1/2)."""
    ctype = getattr(cell, "ctype", XL_CELL_EMPTY)
    if ctype == XL_CELL_EMPTY:
        return None
    if numeric:
        if ctype == XL_CELL_NUMBER:
            return float(cell.value)
        if ctype == XL_CELL_TEXT and cell.value:
            try:
                return float(str(cell.value).replace(",", ".").replace(" ", ""))
            except ValueError:
                return None
        return None
    # text
    if ctype == XL_CELL_TEXT:
        return (cell.value or "").strip() or None
    if ctype == XL_CELL_NUMBER:
        v = cell.value
        return str(int(v)) if isinstance(v, (int, float)) and v == int(v) else str(v)
    return str(cell.value).strip() if cell.value else None


def _load_sheet_xlsx(file_path: Path) -> Any:
    """Načte .xlsx a vrátí objekt s nrows, ncols, cell_value(row,col), cell(row,col) s .value a .ctype."""
    wb = openpyxl.load_workbook(file_path, read_only=False, data_only=True)
    ws = wb.active
    nrows = ws.max_row or 0
    ncols = ws.max_column or 0

    def cell_value(r: int, c: int):
        cell = ws.cell(row=r + 1, column=c + 1)
        v = cell.value
        if v is None or (isinstance(v, str) and not v.strip()):
            return None
        return v

    def cell(r: int, c: int):
        cell = ws.cell(row=r + 1, column=c + 1)
        v = cell.value
        if v is None or (isinstance(v, str) and not v.strip()):
            return SimpleNamespace(value=None, ctype=XL_CELL_EMPTY)
        if isinstance(v, (int, float)):
            return SimpleNamespace(value=v, ctype=XL_CELL_NUMBER)
        return SimpleNamespace(value=str(v).strip(), ctype=XL_CELL_TEXT)

    return SimpleNamespace(nrows=nrows, ncols=ncols, cell_value=cell_value, cell=cell)


def parse_xls_file(file_path: Path) -> ParsedXlsResult:
    """
    Načte .xls nebo .xlsx soubor a vrátí strukturovaný výsledek:
    - rows: seznam řádků jako dict (klíče = interní názvy sloupců)
    - row_errors: chyby po řádcích (řádek se nepodařilo zpracovat)
    - warnings: varování (chybějící sloupce, atd.)
    """
    result = ParsedXlsResult()
    suffix = file_path.suffix.lower()

    try:
        if suffix == ".xlsx":
            sheet = _load_sheet_xlsx(file_path)
        else:
            with xlrd.open_workbook(str(file_path), formatting_info=False) as book:
                sheet = book.sheet_by_index(0)
    except Exception as e:
        result.warnings.append(f"Nelze otevřít soubor: {e}")
        logger.warning("Nelze otevřít soubor %s: %s", file_path, e)
        return result

    if sheet.nrows == 0:
        result.warnings.append("Soubor je prázdný.")
        return result

    header_row_index = _find_header_row(sheet)
    if header_row_index is None:
        result.warnings.append("Nenalezen řádek s hlavičkou (očekáván sloupec Description nebo Popis).")
        return result

    result.header_row_index = header_row_index
    header_row = [sheet.cell_value(header_row_index, c) for c in range(sheet.ncols)]

    # Sloupce dopočítané z dopravy a kurzu – nevarovat
    COMPUTED_OPTIONAL = frozenset({"purchase_price_imported", "margin_7_imported", "vip_eur_imported", "vip_czk_imported", "trade_price_imported", "d4_price_imported"})
    header_cells = [str(c).strip() if c is not None else "" for c in header_row]
    for key in COLUMN_ALIASES:
        idx = _find_column_index(header_row, key)
        if idx is not None:
            result.column_map[key] = idx
            result.detected_headers[key] = header_cells[idx] if idx < len(header_cells) else key
        elif key not in ("description", "pot_size") and key not in COMPUTED_OPTIONAL:
            result.warnings.append(f"Sloupec '{key}' nebyl nalezen (volitelný).")

    if "description" not in result.column_map:
        result.warnings.append("Povinný sloupec 'Description' nebyl nalezen.")
        return result

    # Parsovat datové řádky
    for row_idx in range(header_row_index + 1, sheet.nrows):
        data_row_index = row_idx - (header_row_index + 1)  # 0-based pro report
        try:
            row_dict = {}
            for key, col_idx in result.column_map.items():
                if col_idx >= sheet.ncols:
                    row_dict[key] = None
                    continue
                cell = sheet.cell(row_idx, col_idx)
                numeric = key in NUMERIC_KEYS
                row_dict[key] = _cell_to_value(cell, numeric)
            result.rows.append(row_dict)
        except Exception as e:
            msg = str(e)
            result.row_errors.append(RowParseError(row_index=data_row_index, message=msg))
            logger.debug("Řádek %s v %s: %s", row_idx + 1, file_path.name, msg)

    return result
