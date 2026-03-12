import os
import json
import threading
import requests
from functools import wraps
from datetime import datetime

from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    session,
    flash,
    send_from_directory,
    jsonify,
)
import click
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func as sql_func
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import joinedload, selectinload
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.exceptions import RequestEntityTooLarge

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "change-in-production")
if os.environ.get("TESTING"):
    app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get("TEST_DATABASE_URI", "sqlite:///:memory:")
    app.config["TESTING"] = True
else:
    # V produkci lze nastavit SQLALCHEMY_DATABASE_URI nebo DATABASE_URI (např. sqlite:////srv/http/objednavac/data/warehouse.db)
    _project_root = os.path.dirname(os.path.abspath(__file__))
    _data_dir = os.path.join(_project_root, "data")
    _default_db_path = os.path.join(_data_dir, "warehouse.db")
    os.makedirs(_data_dir, exist_ok=True)
    app.config["SQLALCHEMY_DATABASE_URI"] = (
        os.environ.get("SQLALCHEMY_DATABASE_URI")
        or os.environ.get("DATABASE_URI")
        or (("sqlite:///" + os.path.abspath(_default_db_path)).replace("\\", "/"))
    )
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["UPLOAD_FOLDER"] = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploads")
app.config["MAX_CONTENT_LENGTH"] = None  # Bez limitu – velké XLS/XLSX importy; při 413 od proxy zvýšit client_max_body_size v NGINX
app.config["APP_NAME"] = os.environ.get("APP_NAME", "Objednávač")

os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)

db = SQLAlchemy(app)


@app.errorhandler(RequestEntityTooLarge)
def handle_413(e):
    """413 může vracet reverse proxy (NGINX) – zvýšit client_max_body_size v konfiguraci proxy."""
    return (
        render_template(
            "error_413.html",
            message="Soubor je příliš velký (413). Aplikace nemá limit; pokud používáte NGINX nebo jiný proxy, zvýšte client_max_body_size v jeho konfiguraci.",
        ),
        413,
    )


def _ensure_warehouse_note_column():
    """Přidá sloupec warehouse_note do order_items, pokud chybí (migrace pro existující DB)."""
    if getattr(_ensure_warehouse_note_column, "_done", False):
        return
    _ensure_warehouse_note_column._done = True
    try:
        from sqlalchemy import text
        db.session.execute(text("ALTER TABLE order_items ADD COLUMN warehouse_note TEXT"))
        db.session.commit()
    except Exception:
        db.session.rollback()


def _ensure_user_warehouses_migration():
    """Vytvoří tabulku user_warehouses a sloupec users.warehouse_id, pokud chybí."""
    if getattr(_ensure_user_warehouses_migration, "_done", False):
        return
    _ensure_user_warehouses_migration._done = True
    from sqlalchemy import text
    try:
        db.session.execute(text(
            "CREATE TABLE IF NOT EXISTS user_warehouses (user_id INTEGER NOT NULL, warehouse_id INTEGER NOT NULL, "
            "PRIMARY KEY (user_id, warehouse_id), FOREIGN KEY (user_id) REFERENCES users (id), "
            "FOREIGN KEY (warehouse_id) REFERENCES warehouses (id))"
        ))
        db.session.commit()
    except Exception:
        db.session.rollback()
    try:
        r = db.session.execute(text("PRAGMA table_info(users)"))
        cols = [row[1] for row in r.fetchall()]
        if "warehouse_id" not in cols:
            # Bez REFERENCES kvůli kompatibilitě se staršími SQLite
            db.session.execute(text("ALTER TABLE users ADD COLUMN warehouse_id INTEGER"))
            db.session.commit()
    except Exception as e:
        db.session.rollback()
        app.logger.warning("Migrace users.warehouse_id selhala: %s", e)


def _ensure_product_eans_table():
    """Vytvoří tabulku product_eans pro více EAN na produkt, pokud chybí."""
    if getattr(_ensure_product_eans_table, "_done", False):
        return
    _ensure_product_eans_table._done = True
    from sqlalchemy import text
    try:
        db.session.execute(text(
            "CREATE TABLE IF NOT EXISTS product_eans (id INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT, "
            "product_id INTEGER NOT NULL REFERENCES products (id), ean VARCHAR(50) NOT NULL UNIQUE)"
        ))
        db.session.commit()
    except Exception:
        db.session.rollback()


def _ensure_order_check_status_column():
    """Přidá sloupec check_status do orders (výsledek kontroly; hlavní stav zůstává pending/partially_shipped/shipped)."""
    if getattr(_ensure_order_check_status_column, "_done", False):
        return
    _ensure_order_check_status_column._done = True
    from sqlalchemy import text
    try:
        r = db.session.execute(text("PRAGMA table_info(orders)"))
        cols = [row[1] for row in r.fetchall()]
        if "check_status" not in cols:
            db.session.execute(text("ALTER TABLE orders ADD COLUMN check_status VARCHAR(30)"))
            db.session.commit()
    except Exception:
        db.session.rollback()


def _ensure_order_supplier_id_column():
    """Přidá sloupec supplier_id do orders (FK na suppliers)."""
    if getattr(_ensure_order_supplier_id_column, "_done", False):
        return
    _ensure_order_supplier_id_column._done = True
    from sqlalchemy import text
    try:
        r = db.session.execute(text("PRAGMA table_info(orders)"))
        cols = [row[1] for row in r.fetchall()]
        if "supplier_id" not in cols:
            db.session.execute(text("ALTER TABLE orders ADD COLUMN supplier_id INTEGER REFERENCES suppliers(id)"))
            db.session.commit()
    except Exception:
        db.session.rollback()


def _ensure_goods_receipt_tables():
    """Vytvoří tabulky goods_receipts a goods_receipt_items pro příjem zboží od dodavatele."""
    if getattr(_ensure_goods_receipt_tables, "_done", False):
        return
    _ensure_goods_receipt_tables._done = True
    from sqlalchemy import text
    try:
        db.session.execute(text("""
            CREATE TABLE IF NOT EXISTS goods_receipts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                warehouse_id INTEGER NOT NULL REFERENCES warehouses(id),
                received_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                created_by_id INTEGER REFERENCES users(id),
                note TEXT
            )
        """))
        db.session.execute(text("""
            CREATE TABLE IF NOT EXISTS goods_receipt_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                goods_receipt_id INTEGER NOT NULL REFERENCES goods_receipts(id),
                product_id INTEGER NOT NULL REFERENCES products(id),
                quantity REAL NOT NULL
            )
        """))
        db.session.commit()
    except Exception:
        db.session.rollback()


def _ensure_goods_receipt_supplier_id_column():
    """Přidá sloupec supplier_id do goods_receipts."""
    if getattr(_ensure_goods_receipt_supplier_id_column, "_done", False):
        return
    _ensure_goods_receipt_supplier_id_column._done = True
    from sqlalchemy import text
    try:
        r = db.session.execute(text("PRAGMA table_info(goods_receipts)"))
        cols = [row[1] for row in r.fetchall()]
        if "supplier_id" not in cols:
            db.session.execute(text("ALTER TABLE goods_receipts ADD COLUMN supplier_id INTEGER REFERENCES suppliers(id)"))
            db.session.commit()
    except Exception:
        db.session.rollback()


def _ensure_warehouse_code_column():
    """Přidá sloupec code do warehouses."""
    if getattr(_ensure_warehouse_code_column, "_done", False):
        return
    _ensure_warehouse_code_column._done = True
    from sqlalchemy import text
    try:
        r = db.session.execute(text("PRAGMA table_info(warehouses)"))
        cols = [row[1] for row in r.fetchall()]
        if "code" not in cols:
            db.session.execute(text("ALTER TABLE warehouses ADD COLUMN code VARCHAR(50)"))
            db.session.commit()
    except Exception:
        db.session.rollback()


def _ensure_order_warehouse_id_for_created_by_warehouse():
    """Doplní warehouse_id u objednávek, které vytvořil skladový uživatel a které mají warehouse_id NULL.
    Pouze pokud má uživatel přiřazen právě jeden sklad (jednoznačné přiřazení)."""
    if getattr(_ensure_order_warehouse_id_for_created_by_warehouse, "_done", False):
        return
    _ensure_order_warehouse_id_for_created_by_warehouse._done = True
    try:
        orders = Order.query.filter(
            Order.warehouse_id.is_(None),
            Order.created_by_warehouse_id.isnot(None),
        ).all()
        for order in orders:
            user = User.query.get(order.created_by_warehouse_id)
            if not user:
                continue
            ids = user.get_warehouse_ids()
            if len(ids) == 1:
                order.warehouse_id = ids[0]
        db.session.commit()
    except Exception:
        db.session.rollback()


def _ensure_orders_drop_invoice_columns():
    """Odstraní sloupce invoice_ok a invoice_note z orders (SQLite 3.35+). U starších SQLite se sloupce ponechají v DB, model je již nepoužívá."""
    if getattr(_ensure_orders_drop_invoice_columns, "_done", False):
        return
    _ensure_orders_drop_invoice_columns._done = True
    try:
        r = db.session.execute(text("PRAGMA table_info(orders)"))
        cols = [row[1] for row in r.fetchall()]
        for col in ("invoice_ok", "invoice_note"):
            if col in cols:
                db.session.execute(text(f"ALTER TABLE orders DROP COLUMN {col}"))
                db.session.commit()
    except Exception:
        db.session.rollback()


def _ensure_suppliers_table():
    """Vytvoří tabulku dodavatelů, pokud neexistuje."""
    try:
        db.session.execute(text("""
            CREATE TABLE IF NOT EXISTS suppliers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                supplier_id VARCHAR(100) NOT NULL UNIQUE,
                supplier_name VARCHAR(255) NOT NULL,
                warehouse_id INTEGER REFERENCES warehouses(id)
            )
        """))
        db.session.commit()
    except Exception:
        db.session.rollback()


def _ensure_suppliers_street_municipality():
    """Přidá sloupce ulice a obec do suppliers."""
    if getattr(_ensure_suppliers_street_municipality, "_done", False):
        return
    try:
        cols = [row[0] for row in db.session.execute(text("PRAGMA table_info(suppliers)")).fetchall()]
        if "street" not in cols:
            db.session.execute(text("ALTER TABLE suppliers ADD COLUMN street VARCHAR(255)"))
        if "municipality" not in cols:
            db.session.execute(text("ALTER TABLE suppliers ADD COLUMN municipality VARCHAR(255)"))
        db.session.commit()
        _ensure_suppliers_street_municipality._done = True
    except Exception:
        db.session.rollback()


def _run_migrations():
    """Spustí migrace (sloupce/tabulky). Voláno z before_request a při startu aplikace."""
    _ensure_warehouse_note_column()
    _ensure_user_warehouses_migration()
    _ensure_product_eans_table()
    _ensure_order_check_status_column()
    _ensure_goods_receipt_tables()
    _ensure_suppliers_table()
    _ensure_suppliers_street_municipality()
    _ensure_goods_receipt_supplier_id_column()
    _ensure_order_supplier_id_column()
    _ensure_warehouse_code_column()
    _ensure_order_warehouse_id_for_created_by_warehouse()
    _ensure_orders_drop_invoice_columns()


@app.before_request
def _run_migrations_request():
    _run_migrations()


# Při načtení modulu nejdřív vytvořit tabulky (create_all), pak migrace (důležité pro gunicorn a pro testy s prázdnou DB)
with app.app_context():
    db.create_all()
    _run_migrations()


def find_product_by_ean(ean):
    """Najde produkt podle EAN – hledá v Product.ean i v ProductEan.ean."""
    if not ean or not str(ean).strip():
        return None
    ean = str(ean).strip()
    p = Product.query.filter(Product.ean == ean).first()
    if p:
        return p
    pe = ProductEan.query.filter(ProductEan.ean == ean).first()
    return pe.product if pe else None


def find_product_by_ean_or_code(ean, kod_zbozi=None):
    """Najde produkt podle kódu zboží nebo EAN. Priorita: 1. kod_zbozi, 2. EAN."""
    if kod_zbozi is not None and str(kod_zbozi).strip():
        code = str(kod_zbozi).strip()
        p = Product.query.filter(Product.kod_zbozi == code).first()
        if p:
            return p
    if ean and str(ean).strip():
        return find_product_by_ean(ean)
    return None


ROLES = ("admin", "branch", "warehouse")
# Workflow stavů objednávky (ERP inspirace: pending → partially_shipped → shipped → verified/error)
ORDER_STATUSES = ("pending", "partially_shipped", "shipped", "verified", "error")
STATUS_CZ = {
    "pending": "Čeká",
    "processed": "Zpracováno",  # jen pro zobrazení starých záznamů
    "shipped": "Odesláno",
    "partially_shipped": "Částečně odesláno",
    "verified": "Zkontrolováno",
    "error": "Chyba kontroly",
}


user_branches = db.Table(
    "user_branches",
    db.Column("user_id", db.Integer, db.ForeignKey("users.id"), primary_key=True),
    db.Column("branch_id", db.Integer, db.ForeignKey("branches.id"), primary_key=True),
)

user_warehouses = db.Table(
    "user_warehouses",
    db.Column("user_id", db.Integer, db.ForeignKey("users.id"), primary_key=True),
    db.Column("warehouse_id", db.Integer, db.ForeignKey("warehouses.id"), primary_key=True),
)


class User(db.Model):
    __tablename__ = "users"
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(20), nullable=False)
    branch_id = db.Column(db.Integer, db.ForeignKey("branches.id"), nullable=True)  # výchozí / zpětná kompatibilita
    branch = db.relationship("Branch", backref="users", foreign_keys=[branch_id])
    branches = db.relationship("Branch", secondary=user_branches, backref=db.backref("branch_users", lazy="dynamic"), lazy="dynamic")
    warehouse_id = db.Column(db.Integer, db.ForeignKey("warehouses.id"), nullable=True)  # výchozí sklad
    warehouse = db.relationship("Warehouse", backref="users_default", foreign_keys=[warehouse_id])
    warehouses = db.relationship("Warehouse", secondary=user_warehouses, backref=db.backref("warehouse_users", lazy="dynamic"), lazy="dynamic")

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    def get_branch_ids(self):
        """Seznam ID poboček, k nimž je uživatel přiřazen (včetně branch_id pro zpětnou kompatibilitu)."""
        ids = []
        if self.branch_id:
            ids.append(self.branch_id)
        for b in self.branches:
            if b.id not in ids:
                ids.append(b.id)
        return ids

    def has_any_branch(self):
        return bool(self.branch_id or self.branches.count() > 0)

    def branch_names_display(self):
        """Pro zobrazení v adminu: názvy poboček oddělené čárkou."""
        ids = self.get_branch_ids()
        if not ids:
            return "—"
        names = [Branch.query.get(bid).name for bid in ids if Branch.query.get(bid)]
        return ", ".join(names) if names else "—"

    def get_warehouse_ids(self):
        """Seznam ID skladů, k nimž je uživatel přiřazen."""
        ids = []
        if self.warehouse_id:
            ids.append(self.warehouse_id)
        for w in self.warehouses:
            if w.id not in ids:
                ids.append(w.id)
        return ids

    def has_any_warehouse(self):
        return bool(self.warehouse_id or self.warehouses.count() > 0)

    def warehouse_names_display(self):
        """Pro zobrazení v adminu: názvy skladů oddělené čárkou."""
        ids = self.get_warehouse_ids()
        if not ids:
            return "—"
        names = [Warehouse.query.get(wid).name for wid in ids if Warehouse.query.get(wid)]
        return ", ".join(names) if names else "—"


class Branch(db.Model):
    __tablename__ = "branches"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(255), nullable=False)
    code = db.Column(db.String(50), nullable=True)
    orders = db.relationship("Order", backref="branch", lazy="dynamic")


class Warehouse(db.Model):
    __tablename__ = "warehouses"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(255), nullable=False)
    code = db.Column(db.String(50), nullable=True, unique=True)


class Product(db.Model):
    """Produkty dle schématu: naz_skup, kod_zbozi, nazev, ean, mj, nc, pc."""
    __tablename__ = "products"
    id = db.Column(db.Integer, primary_key=True)
    naz_skup = db.Column(db.String(100), nullable=True)      # group name
    kod_zbozi = db.Column(db.String(100), nullable=True, index=True)  # product code
    nazev = db.Column(db.String(255), nullable=False)        # product name
    ean = db.Column(db.String(50), nullable=True, index=True)  # barcode (hlavní EAN)
    mj = db.Column(db.String(20), nullable=True)            # unit
    nc = db.Column(db.Float, nullable=True)                  # purchase price
    pc = db.Column("pc_float", db.Float, nullable=True)      # selling price
    is_internal = db.Column(db.Boolean, default=False, nullable=False)
    extra_eans = db.relationship("ProductEan", backref="product", lazy="dynamic", cascade="all, delete-orphan")


class ProductEan(db.Model):
    """Další EAN kódy přiřazené k produktu (jeden produkt může mít více EAN)."""
    __tablename__ = "product_eans"
    id = db.Column(db.Integer, primary_key=True)
    product_id = db.Column(db.Integer, db.ForeignKey("products.id"), nullable=False)
    ean = db.Column(db.String(50), nullable=False, unique=True, index=True)


class InternalProduct(db.Model):
    """Samostatná tabulka interních produktů (kancelář) – oddělená od běžných produktů."""
    __tablename__ = "internal_products"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(255), nullable=False)
    sku = db.Column(db.String(100), nullable=True)
    unit = db.Column(db.String(50), nullable=True)
    group_name = db.Column(db.String(255), nullable=True)
    pc = db.Column(db.String(100), nullable=True)


class Order(db.Model):
    __tablename__ = "orders"
    id = db.Column(db.Integer, primary_key=True)
    branch_id = db.Column(db.Integer, db.ForeignKey("branches.id"), nullable=False)
    warehouse_id = db.Column(db.Integer, db.ForeignKey("warehouses.id"), nullable=True)
    status = db.Column(db.String(30), default="pending")
    check_status = db.Column(db.String(30), nullable=True)  # výsledek kontroly: verified/error; hlavní stav je status (pending/partially_shipped/shipped)
    created_at = db.Column(db.DateTime, default=datetime.now)
    created_by_id = db.Column(db.Integer, db.ForeignKey("users.id"), nullable=True)
    created_by = db.relationship("User", foreign_keys=[created_by_id])
    items = db.relationship("OrderItem", backref="order", cascade="all, delete-orphan")
    order_type = db.Column(db.String(30), default="normal", nullable=False)
    created_by_warehouse_id = db.Column(db.Integer, db.ForeignKey("users.id"), nullable=True)
    created_by_warehouse = db.relationship("User", foreign_keys=[created_by_warehouse_id])
    warehouse = db.relationship("Warehouse", backref="orders")
    supplier_id = db.Column(db.Integer, db.ForeignKey("suppliers.id"), nullable=True)
    supplier = db.relationship("Supplier", backref="orders")


class OrderItem(db.Model):
    __tablename__ = "order_items"
    id = db.Column(db.Integer, primary_key=True)
    order_id = db.Column(db.Integer, db.ForeignKey("orders.id"), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey("products.id"), nullable=False)   # běžné objednávky; u interních se použije placeholder
    internal_product_id = db.Column(db.Integer, db.ForeignKey("internal_products.id"), nullable=True)  # interní objednávky
    ordered_quantity = db.Column(db.Float, nullable=False)
    shipped_quantity = db.Column(db.Float, nullable=True)
    branch_note = db.Column(db.String(255), nullable=True)
    warehouse_note = db.Column(db.Text, nullable=True)  # poznámka skladu k položce
    custom_product_name = db.Column(db.String(255), nullable=True)  # název produktu, když není v DB
    unavailable = db.Column(db.Boolean, default=False, nullable=False)
    product = db.relationship("Product", backref="order_items")
    internal_product = db.relationship("InternalProduct", backref="order_items")

    def display_name(self):
        if self.internal_product:
            return self.internal_product.name
        return self.custom_product_name or (self.product.nazev if self.product else "—")


class OrderItemCheck(db.Model):
    """Výsledek kontroly položky objednávky (čtečka čárových kódů)."""
    __tablename__ = "order_item_checks"
    id = db.Column(db.Integer, primary_key=True)
    order_item_id = db.Column(db.Integer, db.ForeignKey("order_items.id"), nullable=False)
    scanned_quantity = db.Column(db.Float, nullable=False)
    expected_quantity = db.Column(db.Float, nullable=False)
    result = db.Column(db.String(20), nullable=False)  # 'correct' | 'incorrect'
    created_at = db.Column(db.DateTime, default=datetime.now, nullable=False)
    order_item = db.relationship("OrderItem", backref=db.backref("checks", lazy="dynamic"))


class GoodsReceipt(db.Model):
    """Příjem zboží od dodavatele (sklad)."""
    __tablename__ = "goods_receipts"
    id = db.Column(db.Integer, primary_key=True)
    warehouse_id = db.Column(db.Integer, db.ForeignKey("warehouses.id"), nullable=False)
    supplier_id = db.Column(db.Integer, db.ForeignKey("suppliers.id"), nullable=True)
    received_at = db.Column(db.DateTime, default=datetime.now, nullable=False)
    created_by_id = db.Column(db.Integer, db.ForeignKey("users.id"), nullable=True)
    created_by = db.relationship("User", foreign_keys=[created_by_id])
    note = db.Column(db.Text, nullable=True)
    warehouse = db.relationship("Warehouse", backref="goods_receipts")
    supplier = db.relationship("Supplier", backref="goods_receipts")
    items = db.relationship("GoodsReceiptItem", backref="goods_receipt", cascade="all, delete-orphan")


class GoodsReceiptItem(db.Model):
    """Položka příjmu zboží – produkt a přijaté množství."""
    __tablename__ = "goods_receipt_items"
    id = db.Column(db.Integer, primary_key=True)
    goods_receipt_id = db.Column(db.Integer, db.ForeignKey("goods_receipts.id"), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey("products.id"), nullable=False)
    quantity = db.Column(db.Float, nullable=False)
    product = db.relationship("Product", backref="goods_receipt_items")


class Supplier(db.Model):
    """Dodavatel (kód, název, ulice, obec). Volitelně vázaný na sklad."""
    __tablename__ = "suppliers"
    id = db.Column(db.Integer, primary_key=True)
    supplier_id = db.Column(db.String(100), nullable=False, unique=True)
    supplier_name = db.Column(db.String(255), nullable=False)
    street = db.Column(db.String(255), nullable=True)
    municipality = db.Column(db.String(255), nullable=True)
    warehouse_id = db.Column(db.Integer, db.ForeignKey("warehouses.id"), nullable=True)
    warehouse = db.relationship("Warehouse", backref="suppliers")


class BranchCart(db.Model):
    """Trvalý košík pro uživatele a pobočku – synchronizace mezi zařízeními. Interní košík zvlášť."""
    __tablename__ = "branch_carts"
    user_id = db.Column(db.Integer, db.ForeignKey("users.id"), primary_key=True)
    branch_id = db.Column(db.Integer, db.ForeignKey("branches.id"), primary_key=True)
    cart_json = db.Column(db.Text, nullable=False, default="{}")
    cart_internal_json = db.Column(db.Text, nullable=False, default="{}")
    notes_json = db.Column(db.Text, nullable=False, default="{}")
    custom_json = db.Column(db.Text, nullable=False, default="[]")


class AuditLog(db.Model):
    """Log změn pro audit (kdo, kdy, co)."""
    __tablename__ = "audit_log"
    id = db.Column(db.Integer, primary_key=True)
    created_at = db.Column(db.DateTime, default=datetime.now, nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey("users.id"), nullable=True)
    username = db.Column(db.String(80), nullable=False)
    action = db.Column(db.String(80), nullable=False)
    entity_type = db.Column(db.String(50), nullable=True)
    entity_id = db.Column(db.Integer, nullable=True)
    details = db.Column(db.Text, nullable=True)
    user = db.relationship("User", backref="audit_logs")


def notify_worker(order_id, branch_name):
    # --- NASTAVENÍ ---
    TOKEN = "8308981650:AAE9ruZxe9lZuMZFM24eEng1UuxwFKem4w4"
    CHAT_ID = "-1003891440639"
    # -----------------

    # Sestavení krátké zprávy
    cas = datetime.now().strftime('%H:%M:%S')
    message = (
        f"🔔 *Nová objednávka!*\n"
        f"🏢 Pobočka: *{branch_name}*\n"
        f"🕒 Čas: {cas}\n"
        f"🆔 ID: #{order_id}"
    )

    url = f"https://api.telegram.org/bot{TOKEN}/sendMessage"
    try:
        requests.post(url, json={
            "chat_id": CHAT_ID,
            "text": message,
            "parse_mode": "Markdown"
        }, timeout=10)
    except Exception as e:
        print(f"Telegram error: {e}")


def login_required(f):
    @wraps(f)
    def inner(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return inner


def role_required(*allowed):
    def decorator(f):
        @wraps(f)
        def inner(*args, **kwargs):
            if "user_id" not in session:
                return redirect(url_for("login"))
            user = User.query.get(session["user_id"])
            effective = user.role if user else None
            if user and user.role == "admin" and session.get("acting_as_role") in allowed:
                effective = session.get("acting_as_role")
            if not user or effective not in allowed:
                flash("Přístup odepřen.", "error")
                if user and user.role == "admin":
                    return redirect(url_for("admin_dashboard"))
                if user and user.role == "warehouse":
                    return redirect(url_for("warehouse_dashboard"))
                return redirect(url_for("login"))
            return f(*args, **kwargs)
        return inner
    return decorator


def get_current_user():
    if "user_id" not in session:
        return None
    return User.query.get(session["user_id"])


def get_current_branch_id():
    """Vrátí ID pobočky, pro kterou uživatel právě objednává (session nebo výchozí). Admin v režimu pobočky používá acting_as_branch_id."""
    user = get_current_user()
    if user and user.role == "admin" and session.get("acting_as_role") == "branch":
        bid = session.get("acting_as_branch_id")
        if bid is not None and Branch.query.get(bid):
            return int(bid)
        return None
    if not user or user.role != "branch":
        return None
    current = session.get("current_branch_id")
    if current is not None:
        try:
            current = int(current)
        except (TypeError, ValueError):
            current = None
    if current and user.get_branch_ids() and current in user.get_branch_ids():
        return current
    ids = user.get_branch_ids()
    if ids:
        session["current_branch_id"] = ids[0]
        return ids[0]
    return user.branch_id


def _default_warehouse_teplice_id():
    """Vrátí ID skladu s názvem nebo kódem 'Teplice' (výchozí sklad)."""
    w = Warehouse.query.filter(
        db.or_(
            db.func.lower(Warehouse.name) == "teplice",
            (Warehouse.code.isnot(None)) & (db.func.lower(Warehouse.code) == "teplice"),
        )
    ).first()
    return w.id if w else None


def get_current_warehouse_id():
    """Vrátí ID skladu, pod kterým uživatel právě působí. Admin v režimu sklad používá acting_as_warehouse_id.
    Výchozí sklad pro roli warehouse je Teplice (pokud je v seznamu)."""
    user = get_current_user()
    if user and user.role == "admin" and session.get("acting_as_role") == "warehouse":
        wid = session.get("acting_as_warehouse_id")
        if wid is not None and Warehouse.query.get(wid):
            return int(wid)
        return None
    if not user or user.role != "warehouse":
        return None
    current = session.get("current_warehouse_id")
    if current is not None:
        try:
            current = int(current)
        except (TypeError, ValueError):
            current = None
    if current and user.get_warehouse_ids() and current in user.get_warehouse_ids():
        return current
    ids = user.get_warehouse_ids()
    if ids:
        teplice_id = _default_warehouse_teplice_id()
        first_id = teplice_id if (teplice_id and teplice_id in ids) else ids[0]
        session["current_warehouse_id"] = first_id
        return first_id
    return user.warehouse_id


def status_cz(key):
    return STATUS_CZ.get(key, key)


def audit_log(action, entity_type=None, entity_id=None, details=None):
    """Zapíše záznam do audit logu. U admina v režimu pobočka/sklad přidá poznámku."""
    user = get_current_user()
    username = user.username if user else "anonymous"
    uid = user.id if user else None
    acting = session.get("acting_as_role")
    if user and user.role == "admin" and acting:
        suffix = f" [admin jako {acting}]"
        details = (details or "") + suffix
    try:
        log = AuditLog(
            user_id=uid,
            username=username,
            action=action,
            entity_type=entity_type,
            entity_id=entity_id,
            details=details,
        )
        db.session.add(log)
        db.session.commit()
    except Exception:
        db.session.rollback()


@app.template_filter("format_datetime")
def format_datetime_filter(dt):
    if dt is None:
        return "—"
    return dt.strftime("%d.%m.%Y %H:%M")


@app.template_filter("format_qty")
def format_qty_filter(value):
    """Zobrazí množství jako celé číslo, pokud je to celé (3.0 → 3)."""
    if value is None:
        return "—"
    try:
        f = float(value)
        if f == int(f):
            return str(int(f))
        return str(f)
    except (TypeError, ValueError):
        return str(value)


# Možnosti mg pro filtr (pořadí v UI)
PRODUCT_MG_OPTIONS = ["0mg", "3mg", "6mg", "10mg", "11mg", "12mg", "18mg", "20mg"]
# Pořadí při detekci v názvu (od nejdelšího, aby "20mg" neodpovídalo jako "0mg")
PRODUCT_MG_DETECT_ORDER = ["20mg", "18mg", "12mg", "11mg", "10mg", "6mg", "3mg", "0mg"]


def _product_category(name):
    """Z názvu produktu extrahuje mg (0mg, 3mg, 6mg, 10mg, 20mg) pro filtrování a označení."""
    if not name or not isinstance(name, str):
        return None
    n = name.lower()
    for mg in PRODUCT_MG_DETECT_ORDER:
        if mg in n:
            return mg
    return None


@app.template_filter("product_category")
def product_category_filter(name):
    """Template filter: vrátí '0mg', '3mg', '6mg', '10mg', '20mg' nebo prázdný řetězec."""
    cat = _product_category(name)
    return cat if cat else ""


def _get_branches_for_current_user():
    """Pobočky, k nimž je přihlášený uživatel (branch) přiřazen – pro výběr pobočky. Admin v režimu pobočky vidí jen tu jednu."""
    user = get_current_user()
    if user and user.role == "admin" and session.get("acting_as_role") == "branch":
        bid = session.get("acting_as_branch_id")
        if bid and Branch.query.get(bid):
            return [Branch.query.get(bid)]
        return []
    if not user or user.role != "branch":
        return []
    ids = user.get_branch_ids()
    if not ids:
        return []
    return Branch.query.filter(Branch.id.in_(ids)).order_by(Branch.name).all()


def _get_warehouses_for_current_user():
    """Sklady, k nimž je přihlášený uživatel (warehouse) přiřazen. Admin v režimu sklad vidí jen ten jeden."""
    user = get_current_user()
    if user and user.role == "admin" and session.get("acting_as_role") == "warehouse":
        wid = session.get("acting_as_warehouse_id")
        if wid and Warehouse.query.get(wid):
            return [Warehouse.query.get(wid)]
        return []
    if not user or user.role != "warehouse":
        return []
    ids = user.get_warehouse_ids()
    if not ids:
        return []
    return Warehouse.query.filter(Warehouse.id.in_(ids)).order_by(Warehouse.name).all()


def _get_branch_cart_count():
    """Počet položek v košíku pobočky (součet množství) pro zobrazení v navbaru/sidebaru."""
    if session.get("order_type") == "internal":
        cart = session.get("cart_internal_by_branch", {}).get(str(get_current_branch_id() or ""), {})
    else:
        cart = session.get("cart_by_branch", {}).get(str(get_current_branch_id() or ""), {})
    custom = session.get("cart_custom_by_branch", {}).get(str(get_current_branch_id() or ""), [])
    total = sum(q for q in cart.values() if q and q > 0)
    total += sum(c.get("quantity", 0) for c in custom if c.get("quantity", 0) > 0)
    return total


def _get_branch_cart_count_internal():
    """Počet položek v interním košíku pobočky."""
    cart = session.get("cart_internal_by_branch", {}).get(str(get_current_branch_id() or ""), {})
    return sum(q for q in cart.values() if q and q > 0)


@app.context_processor
def inject_user():
    current = get_current_user()
    acting_as_role = session.get("acting_as_role") if current and current.role == "admin" else None
    acting_as_branch = None
    if acting_as_role == "branch" and session.get("acting_as_branch_id"):
        acting_as_branch = Branch.query.get(session["acting_as_branch_id"])
    acting_as_warehouse = None
    if acting_as_role == "warehouse" and session.get("acting_as_warehouse_id"):
        acting_as_warehouse = Warehouse.query.get(session["acting_as_warehouse_id"])
    branch_list = _get_branches_for_current_user() if (current and (current.role == "branch" or acting_as_role == "branch")) else []
    warehouse_list = _get_warehouses_for_current_user() if (current and (current.role == "warehouse" or acting_as_role == "warehouse")) else []
    current_branch_id = get_current_branch_id() if current else None
    current_branch = Branch.query.get(current_branch_id) if current_branch_id else None
    current_warehouse_id = get_current_warehouse_id() if current else None
    current_warehouse = Warehouse.query.get(current_warehouse_id) if current_warehouse_id else None
    cart_count = 0
    cart_count_internal = 0
    if current and (current.role == "branch" or acting_as_role == "branch") and current_branch_id:
        _load_branch_cart_from_db()
        cart_count = _get_branch_cart_count()
        cart_count_internal = _get_branch_cart_count_internal()
    router_url = os.environ.get("ROUTER_URL", "").strip()
    if not router_url and request:
        try:
            from urllib.parse import urlparse
            p = urlparse(request.url_root)
            router_url = f"{p.scheme}://{p.hostname}" if p.port in (80, 8000, None) else f"{p.scheme}://{p.hostname}:8000"
        except Exception:
            router_url = "http://localhost:8000"
    else:
        router_url = router_url.rstrip("/") if router_url else "http://localhost:8000"

    all_warehouses = []
    branch_preferred_warehouse_id = None
    if current and (current.role == "branch" or acting_as_role == "branch"):
        all_warehouses = Warehouse.query.order_by(Warehouse.name).all()
        try:
            branch_preferred_warehouse_id = session.get("branch_preferred_warehouse_id")
            if branch_preferred_warehouse_id is not None:
                branch_preferred_warehouse_id = int(branch_preferred_warehouse_id)
        except (TypeError, ValueError):
            branch_preferred_warehouse_id = None

    return {
        "current_user": current,
        "status_cz": status_cz,
        "current_branch": current_branch,
        "user_branches": branch_list,
        "current_warehouse": current_warehouse,
        "user_warehouses": warehouse_list,
        "acting_as_role": acting_as_role,
        "acting_as_branch": acting_as_branch,
        "acting_as_warehouse": acting_as_warehouse,
        "cart_count": cart_count,
        "cart_count_internal": cart_count_internal,
        "router_url": router_url,
        "all_warehouses": all_warehouses,
        "branch_preferred_warehouse_id": branch_preferred_warehouse_id,
        "default_warehouse_teplice_id": _default_warehouse_teplice_id(),
    }


def _sync_branches_from_auth():
    """Synchronizuje pobočky s auth-system: přidá chybějící, smaže ty co v auth-system už nejsou (bez objednávek)."""
    auth_url = (os.environ.get("AUTH_API_URL") or "http://localhost:8080").rstrip("/")
    try:
        r = requests.get(auth_url + "/api/branches", timeout=10)
        if r.status_code != 200:
            return
        data = r.json() if r.headers.get("content-type", "").startswith("application/json") else []
        if not isinstance(data, list):
            return
        auth_names = {(item.get("name") or "").strip() for item in data if (item.get("name") or "").strip()}
        for item in data:
            name = (item.get("name") or "").strip()
            if not name or Branch.query.filter_by(name=name).first():
                continue
            db.session.add(Branch(name=name))
        for b in Branch.query.all():
            if b.name in auth_names:
                continue
            if Order.query.filter_by(branch_id=b.id).count() > 0:
                continue
            User.query.filter_by(branch_id=b.id).update({"branch_id": None})
            for u in User.query.filter(User.branches.any(Branch.id == b.id)).all():
                u.branches.remove(b)
            BranchCart.query.filter_by(branch_id=b.id).delete()
            db.session.delete(b)
        db.session.commit()
    except Exception:
        db.session.rollback()


def _sync_warehouses_from_auth():
    """Synchronizuje sklady s auth-system: přidá chybějící, smaže ty co v auth-system už nejsou (bez objednávek a příjmů)."""
    auth_url = (os.environ.get("AUTH_API_URL") or "http://localhost:8080").rstrip("/")
    try:
        r = requests.get(auth_url + "/api/warehouses", timeout=10)
        if r.status_code != 200:
            return
        data = r.json() if r.headers.get("content-type", "").startswith("application/json") else []
        if not isinstance(data, list):
            return
        auth_names = {(item.get("name") or "").strip() for item in data if (item.get("name") or "").strip()}
        for item in data:
            name = (item.get("name") or "").strip()
            if not name or Warehouse.query.filter_by(name=name).first():
                continue
            db.session.add(Warehouse(name=name, code=(item.get("code") or "").strip() or None))
        for w in Warehouse.query.all():
            if w.name in auth_names:
                continue
            if Order.query.filter_by(warehouse_id=w.id).count() > 0 or GoodsReceipt.query.filter_by(warehouse_id=w.id).count() > 0:
                continue
            User.query.filter_by(warehouse_id=w.id).update({"warehouse_id": None})
            for u in User.query.filter(User.warehouses.any(Warehouse.id == w.id)).all():
                u.warehouses.remove(w)
            Supplier.query.filter_by(warehouse_id=w.id).update({"warehouse_id": None})
            db.session.delete(w)
        db.session.commit()
    except Exception:
        db.session.rollback()


def _auth_api_login(username, pin):
    """Ověří přihlášení přes centrální auth API (application=objednavac). Vrátí (data, None) nebo (None, error)."""
    auth_url = (os.environ.get("AUTH_API_URL") or "http://localhost:8080").rstrip("/")
    try:
        r = requests.post(
            auth_url + "/api/login",
            json={"username": username, "pin": pin, "application": "objednavac"},
            timeout=10,
        )
        data = r.json() if r.headers.get("content-type", "").startswith("application/json") else {}
        if r.status_code == 200 and data.get("ok"):
            return data, None
        return None, data.get("error", "Neplatné přihlašovací údaje.")
    except requests.RequestException as e:
        return None, str(e)


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        pin = request.form.get("pin", "").strip()
        password = request.form.get("password", "")
        # Centrální API: username + PIN
        if username and pin:
            data, err = _auth_api_login(username, pin)
            if err:
                flash(err if "spojit" not in err else f"Nepodařilo se spojit s centrálním přihlášením: {err}", "error")
                return render_template("login.html")
            _sync_branches_from_auth()
            _sync_warehouses_from_auth()
            role = data.get("role", "user")
            warehouse_name = (data.get("warehouse") or "").strip()
            branch_list = []
            if data.get("branches") and isinstance(data["branches"], list):
                for item in data["branches"]:
                    name = (item.get("name") if isinstance(item, dict) else str(item)).strip()
                    if not name:
                        continue
                    b = Branch.query.filter_by(name=name).first()
                    if not b:
                        b = Branch(name=name)
                        db.session.add(b)
                        db.session.flush()
                    branch_list.append(b)
            if not branch_list and (data.get("branch") or "").strip():
                branch_name = (data.get("branch") or "").strip()
                branch = Branch.query.filter_by(name=branch_name).first()
                if not branch:
                    branch = Branch(name=branch_name)
                    db.session.add(branch)
                    db.session.flush()
                branch_list = [branch]
            branch = branch_list[0] if branch_list else None
            warehouse = None
            if warehouse_name:
                warehouse = Warehouse.query.filter_by(name=warehouse_name).first()
                if not warehouse:
                    warehouse = Warehouse(name=warehouse_name)
                    db.session.add(warehouse)
                    db.session.flush()
            user = User.query.filter_by(username=username).first()
            if not user:
                user = User(
                    username=username,
                    password_hash=generate_password_hash(os.urandom(16).hex()),
                    role=role,
                    branch_id=branch.id if branch and role == "branch" else None,
                    warehouse_id=warehouse.id if warehouse and role == "warehouse" else None,
                )
                db.session.add(user)
                db.session.commit()
                if branch_list and role == "branch":
                    for b in branch_list:
                        if b not in user.branches:
                            user.branches.append(b)
                if warehouse and role == "warehouse":
                    user.warehouses.append(warehouse)
                db.session.commit()
            else:
                user.role = role
                user.branch_id = branch.id if branch and role == "branch" else None
                user.warehouse_id = warehouse.id if warehouse and role == "warehouse" else None
                for b in list(user.branches):
                    user.branches.remove(b)
                if branch_list and role == "branch":
                    for b in branch_list:
                        user.branches.append(b)
                for w in list(user.warehouses):
                    user.warehouses.remove(w)
                if warehouse and role == "warehouse":
                    user.warehouses.append(warehouse)
                db.session.commit()
            session["user_id"] = user.id
            session.pop("current_branch_id", None)
            session.pop("current_warehouse_id", None)
            session.pop("acting_as_role", None)
            session.pop("acting_as_branch_id", None)
            session.pop("acting_as_warehouse_id", None)
            if user.role == "branch":
                ids = user.get_branch_ids()
                if ids:
                    session["current_branch_id"] = ids[0]
            if user.role == "warehouse":
                ids = user.get_warehouse_ids()
                if ids:
                    teplice_id = _default_warehouse_teplice_id()
                    session["current_warehouse_id"] = teplice_id if (teplice_id and teplice_id in ids) else ids[0]
            if user.role == "admin":
                return redirect(url_for("admin_dashboard"))
            if user.role == "warehouse":
                return redirect(url_for("warehouse_dashboard"))
            if user.role == "branch":
                return redirect(url_for("branch_dashboard"))
            return redirect(url_for("index"))
        # Zpětná kompatibilita: heslo proti lokální DB
        if username and password:
            user = User.query.filter_by(username=username).first()
            if user and user.check_password(password):
                session["user_id"] = user.id
                session.pop("current_branch_id", None)
                session.pop("current_warehouse_id", None)
                if user.role == "branch":
                    ids = user.get_branch_ids()
                    if ids:
                        session["current_branch_id"] = ids[0]
                if user.role == "warehouse":
                    ids = user.get_warehouse_ids()
                    if ids:
                        teplice_id = _default_warehouse_teplice_id()
                        session["current_warehouse_id"] = teplice_id if (teplice_id and teplice_id in ids) else ids[0]
                if user.role == "admin":
                    return redirect(url_for("admin_dashboard"))
                if user.role == "warehouse":
                    return redirect(url_for("warehouse_dashboard"))
                if user.role == "branch":
                    return redirect(url_for("branch_dashboard"))
                return redirect(url_for("index"))
        flash("Zadejte uživatelské jméno a PIN (nebo jméno a heslo).", "error")
    return render_template("login.html")


@app.route("/auth/sso")
def auth_sso():
    """SSO ze Směrosu: ověření tokenu u auth-system a vytvoření lokální session."""
    token = (request.args.get("token") or "").strip()
    if not token:
        flash("Chybí SSO token.", "error")
        return redirect(url_for("login"))
    auth_url = (os.environ.get("AUTH_API_URL") or "http://localhost:8080").rstrip("/")
    try:
        r = requests.get(auth_url + "/api/sso/verify", params={"token": token}, timeout=10)
        data = r.json() if r.headers.get("content-type", "").startswith("application/json") else {}
        if r.status_code != 200 or not data.get("ok"):
            flash(data.get("error", "Neplatný nebo vypršený SSO token."), "error")
            return redirect(url_for("login"))
    except requests.RequestException as e:
        flash(f"Nepodařilo se ověřit token: {e}", "error")
        return redirect(url_for("login"))
    username = (data.get("username") or "").strip()
    if not username:
        flash("Neplatná odpověď přihlášení.", "error")
        return redirect(url_for("login"))
    _sync_branches_from_auth()
    _sync_warehouses_from_auth()
    role = data.get("role", "user")
    warehouse_name = (data.get("warehouse") or "").strip()
    branch_list = []
    if data.get("branches") and isinstance(data["branches"], list):
        for item in data["branches"]:
            name = (item.get("name") if isinstance(item, dict) else str(item)).strip()
            if not name:
                continue
            b = Branch.query.filter_by(name=name).first()
            if not b:
                b = Branch(name=name)
                db.session.add(b)
                db.session.flush()
            branch_list.append(b)
    if not branch_list and (data.get("branch") or "").strip():
        branch_name = (data.get("branch") or "").strip()
        branch = Branch.query.filter_by(name=branch_name).first()
        if not branch:
            branch = Branch(name=branch_name)
            db.session.add(branch)
            db.session.flush()
        branch_list = [branch]
    branch = branch_list[0] if branch_list else None
    warehouse = None
    if warehouse_name:
        warehouse = Warehouse.query.filter_by(name=warehouse_name).first()
        if not warehouse:
            warehouse = Warehouse(name=warehouse_name)
            db.session.add(warehouse)
            db.session.flush()
    user = User.query.filter_by(username=username).first()
    if not user:
        user = User(
            username=username,
            password_hash=generate_password_hash(os.urandom(16).hex()),
            role=role,
            branch_id=branch.id if branch and role == "branch" else None,
            warehouse_id=warehouse.id if warehouse and role == "warehouse" else None,
        )
        db.session.add(user)
        db.session.commit()
        if branch_list and role == "branch":
            for b in branch_list:
                if b not in user.branches:
                    user.branches.append(b)
        if warehouse and role == "warehouse":
            user.warehouses.append(warehouse)
        db.session.commit()
    else:
        user.role = role
        user.branch_id = branch.id if branch and role == "branch" else None
        user.warehouse_id = warehouse.id if warehouse and role == "warehouse" else None
        for b in list(user.branches):
            user.branches.remove(b)
        if branch_list and role == "branch":
            for b in branch_list:
                user.branches.append(b)
        for w in list(user.warehouses):
            user.warehouses.remove(w)
        if warehouse and role == "warehouse":
            user.warehouses.append(warehouse)
        db.session.commit()
    session["user_id"] = user.id
    session.pop("current_branch_id", None)
    session.pop("current_warehouse_id", None)
    session.pop("acting_as_role", None)
    session.pop("acting_as_branch_id", None)
    session.pop("acting_as_warehouse_id", None)
    if user.role == "branch":
        ids = user.get_branch_ids()
        if ids:
            session["current_branch_id"] = ids[0]
    if user.role == "warehouse":
        ids = user.get_warehouse_ids()
        if ids:
            teplice_id = _default_warehouse_teplice_id()
            session["current_warehouse_id"] = teplice_id if (teplice_id and teplice_id in ids) else ids[0]
    if user.role == "admin":
        return redirect(url_for("admin_dashboard"))
    if user.role == "warehouse":
        return redirect(url_for("warehouse_dashboard"))
    if user.role == "branch":
        return redirect(url_for("branch_dashboard"))
    return redirect(url_for("index"))


@app.route("/branch/switch", methods=["POST"])
@login_required
@role_required("branch")
def branch_switch():
    branch_id = request.form.get("branch_id", type=int)
    user = get_current_user()
    if branch_id and user.get_branch_ids() and branch_id in user.get_branch_ids():
        session["current_branch_id"] = branch_id
        flash("Pobočka změněna.")
    return redirect(request.referrer or url_for("branch_dashboard"))


@app.route("/warehouse/switch", methods=["POST"])
@login_required
@role_required("warehouse")
def warehouse_switch():
    warehouse_id = request.form.get("warehouse_id", type=int)
    user = get_current_user()
    if warehouse_id and user.get_warehouse_ids() and warehouse_id in user.get_warehouse_ids():
        session["current_warehouse_id"] = warehouse_id
        flash("Sklad změněn.")
    return redirect(request.referrer or url_for("warehouse_dashboard"))


@app.route("/branch/set-warehouse", methods=["POST"])
@login_required
@role_required("branch")
def branch_set_warehouse():
    """Pobočka zvolí sklad, ze kterého chce objednávat (výchozí je Teplice)."""
    warehouse_id = request.form.get("warehouse_id", type=int)
    if warehouse_id and Warehouse.query.get(warehouse_id):
        session["branch_preferred_warehouse_id"] = warehouse_id
        flash("Sklad pro objednávky změněn.")
    else:
        session.pop("branch_preferred_warehouse_id", None)
        flash("Pro objednávky bude použit výchozí sklad (Teplice).")
    return redirect(request.referrer or url_for("index"))


def _router_url():
    url = os.environ.get("ROUTER_URL", "").strip()
    if url:
        return url.rstrip("/")
    try:
        from urllib.parse import urlparse
        p = urlparse(request.url_root)
        return f"{p.scheme}://{p.hostname}" if p.port in (80, 8000, None) else f"{p.scheme}://{p.hostname}:8000"
    except Exception:
        return "http://localhost:8000"


def _logout_chain_next_url():
    """Další krok řetězu po Objednávači: DPD logout s chain=1."""
    try:
        from urllib.parse import urlparse
        p = urlparse(_router_url())
        base = f"{p.scheme}://{p.hostname}"
        return f"{base}:8083/logout?chain=1"
    except Exception:
        return "http://localhost:8083/logout?chain=1"


@app.route("/logout")
def logout():
    session.pop("user_id", None)
    session.pop("current_branch_id", None)
    session.pop("current_warehouse_id", None)
    session.pop("acting_as_role", None)
    session.pop("acting_as_branch_id", None)
    session.pop("acting_as_warehouse_id", None)
    next_url = request.args.get("next", "").strip()
    if next_url and next_url.startswith("http"):
        return redirect(next_url)
    if request.args.get("chain") == "1":
        return redirect(_logout_chain_next_url())
    return redirect(_router_url() + "/logout")


def _load_branch_cart_from_db():
    """Načte košík z DB do session – vždy (běžný i interní). Košík je sdílený na úroveň pobočky (všichni uživatelé pobočky vidí stejný košík)."""
    user = get_current_user()
    bid = get_current_branch_id()
    if not user or not bid:
        return
    row = BranchCart.query.filter_by(branch_id=bid).order_by(BranchCart.user_id).first()
    cart_by_branch = session.get("cart_by_branch", {})
    cart_internal_by_branch = session.get("cart_internal_by_branch", {})
    notes_by_branch = session.get("cart_branch_notes", {})
    custom_by_branch = session.get("cart_custom_by_branch", {})
    if row:
        try:
            cart_by_branch[str(bid)] = json.loads(row.cart_json) if row.cart_json else {}
            try:
                cart_internal_by_branch[str(bid)] = json.loads(row.cart_internal_json) if getattr(row, "cart_internal_json", None) else {}
            except (json.JSONDecodeError, TypeError, AttributeError):
                cart_internal_by_branch[str(bid)] = {}
            notes_by_branch[str(bid)] = json.loads(row.notes_json) if row.notes_json else {}
            custom_by_branch[str(bid)] = json.loads(row.custom_json) if row.custom_json else []
        except (json.JSONDecodeError, TypeError):
            cart_by_branch[str(bid)] = {}
            cart_internal_by_branch[str(bid)] = {}
            notes_by_branch[str(bid)] = {}
            custom_by_branch[str(bid)] = []
    else:
        cart_by_branch[str(bid)] = {}
        cart_internal_by_branch[str(bid)] = {}
        notes_by_branch[str(bid)] = {}
        custom_by_branch[str(bid)] = []
    session["cart_by_branch"] = cart_by_branch
    session["cart_internal_by_branch"] = cart_internal_by_branch
    session["cart_branch_notes"] = notes_by_branch
    session["cart_custom_by_branch"] = custom_by_branch


def _save_branch_cart_to_db():
    """Uloží aktuální košíky (běžný i interní) aktuální pobočky do DB."""
    user = get_current_user()
    bid = get_current_branch_id()
    if not user or not bid:
        return
    cart_by_branch = session.get("cart_by_branch", {})
    cart_internal_by_branch = session.get("cart_internal_by_branch", {})
    notes_by_branch = session.get("cart_branch_notes", {})
    custom_by_branch = session.get("cart_custom_by_branch", {})
    cart = cart_by_branch.get(str(bid), {})
    cart_internal = cart_internal_by_branch.get(str(bid), {})
    notes = notes_by_branch.get(str(bid), {})
    custom = custom_by_branch.get(str(bid), [])
    row = BranchCart.query.filter_by(branch_id=bid).order_by(BranchCart.user_id).first()
    if not row:
        row = BranchCart(user_id=user.id, branch_id=bid)
        db.session.add(row)
    row.cart_json = json.dumps(cart)
    if hasattr(row, "cart_internal_json"):
        row.cart_internal_json = json.dumps(cart_internal)
    row.notes_json = json.dumps(notes)
    row.custom_json = json.dumps(custom)
    db.session.commit()


def _get_branch_cart():
    """Košík podle order_type: běžný nebo interní. Načte z DB, pokud ještě nebyl načten."""
    bid = get_current_branch_id()
    if not bid:
        return {}
    _load_branch_cart_from_db()
    if session.get("order_type") == "internal":
        return session.get("cart_internal_by_branch", {}).get(str(bid), {})
    return session.get("cart_by_branch", {}).get(str(bid), {})


def _set_branch_cart(cart):
    user = get_current_user()
    bid = get_current_branch_id()
    if not user or not bid:
        return
    if session.get("order_type") == "internal":
        cart_internal_by_branch = session.get("cart_internal_by_branch", {})
        cart_internal_by_branch[str(bid)] = cart
        session["cart_internal_by_branch"] = cart_internal_by_branch
    else:
        cart_by_branch = session.get("cart_by_branch", {})
        cart_by_branch[str(bid)] = cart
        session["cart_by_branch"] = cart_by_branch
    _save_branch_cart_to_db()


def _get_branch_cart_notes():
    user = get_current_user()
    bid = get_current_branch_id()
    if not user or not bid:
        return {}
    notes = session.get("cart_branch_notes", {})
    return notes.get(str(bid), {})


def _set_branch_cart_notes(notes):
    user = get_current_user()
    bid = get_current_branch_id()
    if not user or not bid:
        return
    by_branch = session.get("cart_branch_notes", {})
    by_branch[str(bid)] = notes
    session["cart_branch_notes"] = by_branch
    _save_branch_cart_to_db()


def _get_branch_cart_custom():
    user = get_current_user()
    bid = get_current_branch_id()
    if not user or not bid:
        return []
    by_branch = session.get("cart_custom_by_branch", {})
    return by_branch.get(str(bid), [])


def _set_branch_cart_custom(items):
    user = get_current_user()
    bid = get_current_branch_id()
    if not user or not bid:
        return
    by_branch = session.get("cart_custom_by_branch", {})
    by_branch[str(bid)] = items
    session["cart_custom_by_branch"] = by_branch
    _save_branch_cart_to_db()


@app.route("/branch")
@login_required
@role_required("branch")
def branch_dashboard():
    from sqlalchemy import func
    bid = get_current_branch_id()
    if not bid:
        return redirect(url_for("index"))
    branch = Branch.query.get(bid)
    if not branch:
        return redirect(url_for("index"))
    orders_q = Order.query.filter_by(branch_id=bid)
    total_orders = orders_q.count()
    by_status = (
        db.session.query(Order.status, func.count(Order.id))
        .filter(Order.branch_id == bid)
        .group_by(Order.status)
        .all()
    )
    status_counts = {s: c for s, c in by_status}
    recent_orders = Order.query.filter_by(branch_id=bid).order_by(Order.created_at.desc()).limit(4).all()
    last_order = Order.query.filter_by(branch_id=bid, order_type="normal").order_by(Order.created_at.desc()).first()
    last_order_unavailable = []
    if last_order:
        last_order_unavailable = [i for i in last_order.items if i.unavailable]
    return render_template(
        "branch_dashboard.html",
        user=get_current_user(),
        branch=branch,
        total_orders=total_orders,
        status_counts=status_counts,
        recent_orders=recent_orders,
        status_cz=status_cz,
        last_order=last_order,
        last_order_unavailable=last_order_unavailable,
    )


def _product_brand(name):
    """První slovo názvu = značka."""
    if not name or not name.strip():
        return ""
    return (name.strip().split() or [""])[0]


@app.route("/")
@login_required
@role_required("branch")
def index():
    """Běžné produkty (ne interní) – samostatná stránka. Vyhledávání je live (bez refresh), filtruje v rámci načtených produktů."""
    from collections import defaultdict
    from sqlalchemy import or_
    session["order_type"] = "normal"
    user = get_current_user()
    q = request.args.get("q", "").strip()
    group_filter = request.args.get("group", "").strip() or None
    brand_filter = request.args.get("brand", "").strip() or None
    category_filter = request.args.get("category", "").strip() or None  # 10mg | 20mg
    page = request.args.get("page", 1, type=int)
    per_page = 300
    if page < 1:
        page = 1
    base = Product.query.filter(
        Product.nazev != "[Vlastní – produkt mimo katalog]",
        db.or_(Product.is_internal == False, Product.is_internal.is_(None)),
    )
    if group_filter:
        base = base.filter(Product.naz_skup == group_filter)
    if category_filter and category_filter in PRODUCT_MG_OPTIONS:
        base = base.filter(Product.nazev.ilike(f"%{category_filter}%"))
    if brand_filter:
        base = base.filter(db.or_(Product.nazev.ilike(brand_filter + " %"), Product.nazev == brand_filter))
    if q:
        q_like = f"%{q}%"
        ean_in_extra = db.session.query(ProductEan.id).filter(
            ProductEan.product_id == Product.id, ProductEan.ean.ilike(q_like)
        ).exists()
        base = base.filter(
            or_(
                Product.nazev.ilike(q_like),
                Product.kod_zbozi.ilike(q_like),
                (Product.naz_skup.isnot(None)) & (Product.naz_skup.ilike(q_like)),
                (Product.ean.isnot(None)) & (Product.ean.ilike(q_like)),
                ean_in_extra,
            )
        )
    base = base.order_by(Product.naz_skup, Product.nazev)
    total = base.count()
    total_pages = max(1, (total + per_page - 1) // per_page) if total else 1
    if page > total_pages:
        page = total_pages
    products = base.offset((page - 1) * per_page).limit(per_page).all()
    # Dropdown možnosti z celého filtrovaného výsledku (ne jen z aktuální stránky)
    groups_query = base.with_entities(Product.naz_skup).distinct().all()
    groups = sorted({r[0] for r in groups_query if r and r[0]})
    nazev_for_brands = base.with_entities(Product.nazev).distinct().limit(15000).all()
    brands = sorted({_product_brand(r[0]) for r in nazev_for_brands if r and r[0] and _product_brand(r[0])})
    categories = PRODUCT_MG_OPTIONS.copy()
    by_brand = defaultdict(list)
    for p in products:
        by_brand[_product_brand(p.nazev) or "—"].append(p)
    products_by_brand = sorted(by_brand.items(), key=lambda x: x[0])
    cart = _get_branch_cart()
    cart_notes = _get_branch_cart_notes()
    waiting_by_product = {}
    bid = get_current_branch_id()
    if bid:
        from sqlalchemy import func
        rows = (
            db.session.query(
                OrderItem.product_id,
                func.sum(OrderItem.ordered_quantity - func.coalesce(OrderItem.shipped_quantity, 0)),
            )
            .join(Order)
            .filter(
                Order.branch_id == bid,
                Order.order_type == "normal",
                Order.status.in_(["pending", "partially_shipped"]),
                OrderItem.unavailable == False,
                OrderItem.product_id.isnot(None),
            )
            .group_by(OrderItem.product_id)
            .all()
        )
        for pid, qty in rows:
            if pid and qty is not None and qty > 0:
                waiting_by_product[pid] = float(qty)
    ctx = dict(
        products=products,
        products_by_brand=products_by_brand,
        cart=cart,
        cart_notes=cart_notes,
        waiting_by_product=waiting_by_product,
        user=user,
        search_q=q,
        groups=groups,
        group_filter=group_filter,
        brands=brands,
        brand_filter=brand_filter,
        categories=categories,
        category_filter=category_filter,
        page=page,
        total_pages=total_pages,
        total=total,
    )
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return render_template("_products_grid_partial.html", **ctx)
    return render_template("products.html", **ctx)


@app.route("/internal")
@login_required
@role_required("branch")
def internal_products():
    """Interní produkty ze samostatné tabulky internal_products – oddělená stránka a databáze."""
    from collections import defaultdict
    from sqlalchemy import or_
    session["order_type"] = "internal"
    user = get_current_user()
    q = request.args.get("q", "").strip()
    group_filter = request.args.get("group", "").strip() or None
    brand_filter = request.args.get("brand", "").strip() or None
    base = InternalProduct.query
    if group_filter:
        base = base.filter(InternalProduct.group_name == group_filter)
    if q:
        q_like = f"%{q}%"
        products = base.filter(
            or_(InternalProduct.name.ilike(q_like), db.and_(InternalProduct.sku.isnot(None), InternalProduct.sku.ilike(q_like)))
        ).order_by(InternalProduct.name).all()
    else:
        products = base.order_by(InternalProduct.group_name, InternalProduct.name).all()
    brands = sorted({_product_brand(p.name) for p in products if _product_brand(p.name)})
    if brand_filter:
        products = [p for p in products if _product_brand(p.name) == brand_filter]
    groups = sorted({p.group_name for p in products if p.group_name})
    by_brand = defaultdict(list)
    for p in products:
        by_brand[_product_brand(p.name) or "—"].append(p)
    products_by_brand = sorted(by_brand.items(), key=lambda x: x[0])
    cart = _get_branch_cart()
    cart_notes = _get_branch_cart_notes()
    return render_template(
        "internal_products.html",
        products=products,
        products_by_brand=products_by_brand,
        cart=cart,
        cart_notes=cart_notes,
        user=user,
        search_q=q,
        groups=groups,
        group_filter=group_filter,
        brands=brands,
        brand_filter=brand_filter,
    )


def _get_or_create_custom_placeholder_product():
    p = Product.query.filter_by(nazev="[Vlastní – produkt mimo katalog]").first()
    if not p:
        p = Product(nazev="[Vlastní – produkt mimo katalog]", kod_zbozi=None, mj="ks")
        db.session.add(p)
        db.session.flush()
    return p


def _redirect_back_to_products(product_id=None):
    """Přesměruje zpět na stránku produktů s původním hledáním (bez scrollu na řádek)."""
    from urllib.parse import urlparse, urlunparse
    referrer = request.referrer
    if referrer and referrer.startswith(request.host_url):
        parsed = urlparse(referrer)
        if parsed.path.rstrip("/") in ("", "/") and "/cart" not in referrer:
            next_url = urlunparse((parsed.scheme, parsed.netloc, parsed.path, parsed.params, parsed.query, ""))
            return redirect(next_url)
    return redirect(url_for("index"))


@app.route("/cart/add-custom", methods=["POST"])
@login_required
@role_required("branch")
def cart_add_custom():
    name = request.form.get("custom_product_name", "").strip()
    quantity = request.form.get("quantity", type=float, default=1)
    if not name or quantity <= 0:
        flash("Zadejte název produktu a množství.", "error")
        return redirect(url_for("index"))
    custom = _get_branch_cart_custom()
    custom.append({"name": name, "quantity": quantity})
    _set_branch_cart_custom(custom)
    flash(f"Přidáno: {name} ({quantity} ks)")
    return _redirect_back_to_products()


@app.route("/cart/add", methods=["POST"])
@login_required
@role_required("branch")
def cart_add():
    product_id = request.form.get("product_id", type=int)
    quantity = request.form.get("quantity", type=float, default=1)
    wants_json = request.headers.get("X-Requested-With") == "XMLHttpRequest"
    if not product_id or quantity <= 0:
        if wants_json:
            return jsonify({"ok": False, "error": "Neplatné množství."}), 400
        flash("Neplatné množství.", "error")
        return redirect(url_for("index"))
    product = Product.query.get(product_id)
    if not product:
        if wants_json:
            return jsonify({"ok": False, "error": "Produkt nenalezen."}), 400
        flash("Produkt nenalezen.", "error")
        return redirect(url_for("index"))
    cart = _get_branch_cart()
    key = str(product_id)
    cart[key] = cart.get(key, 0) + quantity
    _set_branch_cart(cart)
    branch_note = request.form.get("branch_note", "").strip() or None
    notes = _get_branch_cart_notes()
    if branch_note:
        notes[key] = branch_note
    else:
        notes.pop(key, None)
    _set_branch_cart_notes(notes)
    if wants_json:
        total = sum((v for v in _get_branch_cart().values() if v and v > 0))
        return jsonify({"ok": True, "cart_qty": cart[key], "cart_total": total, "message": f"Přidáno: {product.nazev}"})
    return _redirect_back_to_products(product_id=product_id)


@app.route("/cart/set-quantity", methods=["POST"])
@login_required
@role_required("branch")
def cart_set_quantity():
    """Nastaví množství produktu v košíku (pro tlačítka +/- na kartách). Vrací JSON."""
    product_id = request.form.get("product_id", type=int)
    quantity = request.form.get("quantity", type=float)
    wants_json = request.headers.get("X-Requested-With") == "XMLHttpRequest"
    if not product_id:
        if wants_json:
            return jsonify({"ok": False, "error": "Chybí produkt."}), 400
        return redirect(url_for("index"))
    if quantity is None or quantity < 0:
        quantity = 0
    cart = _get_branch_cart()
    key = str(product_id)
    if quantity <= 0:
        cart.pop(key, None)
        notes = _get_branch_cart_notes()
        notes.pop(key, None)
        _set_branch_cart_notes(notes)
        qty = 0
    else:
        cart[key] = quantity
        qty = quantity
    _set_branch_cart(cart)
    if wants_json:
        total = sum((v for v in _get_branch_cart().values() if v and v > 0))
        return jsonify({"ok": True, "cart_qty": qty, "cart_total": total})
    return redirect(request.referrer or url_for("index"))


@app.route("/cart/add-internal", methods=["POST"])
@login_required
@role_required("branch")
def cart_add_internal():
    """Přidá interní produkt (z tabulky internal_products) do interního košíku."""
    session["order_type"] = "internal"
    internal_product_id = request.form.get("internal_product_id", type=int) or request.form.get("product_id", type=int)
    quantity = request.form.get("quantity", type=float, default=1)
    wants_json = request.headers.get("X-Requested-With") == "XMLHttpRequest"
    if not internal_product_id or quantity <= 0:
        if wants_json:
            return jsonify({"ok": False, "error": "Neplatné množství."}), 400
        flash("Neplatné množství.", "error")
        return redirect(url_for("internal_products"))
    ip = InternalProduct.query.get(internal_product_id)
    if not ip:
        if wants_json:
            return jsonify({"ok": False, "error": "Interní produkt nenalezen."}), 400
        flash("Interní produkt nenalezen.", "error")
        return redirect(url_for("internal_products"))
    cart = _get_branch_cart()
    key = str(internal_product_id)
    cart[key] = cart.get(key, 0) + quantity
    _set_branch_cart(cart)
    if wants_json:
        return jsonify({"ok": True, "cart_qty": cart[key], "message": f"Přidáno: {ip.name}"})
    flash(f"Přidáno: {ip.name} ({quantity} ks)")
    return redirect(request.referrer or url_for("internal_products"))


@app.route("/cart")
@login_required
@role_required("branch")
def cart_view():
    if request.args.get("type") == "internal":
        session["order_type"] = "internal"
    elif request.args.get("type") == "normal":
        session["order_type"] = "normal"
    user = get_current_user()
    cart = _get_branch_cart()
    cart_notes = _get_branch_cart_notes()
    order_type = session.get("order_type", "normal")
    items = []
    if order_type == "internal":
        for pid_str, qty in cart.items():
            if qty <= 0:
                continue
            pid = int(pid_str)
            p = InternalProduct.query.get(pid)
            if p:
                items.append({
                    "product": p,
                    "quantity": qty,
                    "branch_note": cart_notes.get(pid_str, ""),
                    "is_custom": False,
                })
    else:
        for pid_str, qty in cart.items():
            if qty <= 0:
                continue
            pid = int(pid_str)
            p = Product.query.get(pid)
            if p:
                items.append({
                    "product": p,
                    "quantity": qty,
                    "branch_note": cart_notes.get(pid_str, ""),
                    "is_custom": False,
                })
        custom = _get_branch_cart_custom()
        for idx, c in enumerate(custom):
            if c.get("quantity", 0) <= 0:
                continue
            items.append({
                "product": None,
                "quantity": c["quantity"],
                "branch_note": "",
                "is_custom": True,
                "custom_name": c.get("name", ""),
                "custom_index": idx,
            })
    return render_template("cart.html", items=items, user=user, order_type=order_type)


@app.route("/cart/update", methods=["POST"])
@login_required
@role_required("branch")
def cart_update():
    cart = _get_branch_cart()
    notes = _get_branch_cart_notes()
    for key in list(cart.keys()):
        new_qty = request.form.get(f"qty_{key}", type=float)
        if new_qty is not None:
            if new_qty <= 0:
                cart.pop(key, None)
                notes.pop(key, None)
            else:
                cart[key] = new_qty
        note_val = request.form.get(f"note_{key}", "").strip() or None
        if key in cart:
            if note_val:
                notes[key] = note_val
            else:
                notes.pop(key, None)
    _set_branch_cart(cart)
    _set_branch_cart_notes(notes)
    if session.get("order_type") != "internal":
        custom = []
        idx = 0
        while idx < 500:
            name_val = request.form.get(f"custom_{idx}_name")
            if name_val is None:
                break
            name = name_val.strip()
            qty = request.form.get(f"custom_{idx}_qty", type=float)
            if name and qty is not None and qty > 0:
                custom.append({"name": name, "quantity": qty})
            idx += 1
        _set_branch_cart_custom(custom)
    return redirect(url_for("cart_view"))


@app.route("/cart/remove-custom/<int:index>", methods=["POST"])
@login_required
@role_required("branch")
def cart_remove_custom(index):
    custom = _get_branch_cart_custom()
    if 0 <= index < len(custom):
        custom.pop(index)
        _set_branch_cart_custom(custom)
    return redirect(url_for("cart_view"))


@app.route("/cart/remove/<int:product_id>", methods=["POST"])
@login_required
@role_required("branch")
def cart_remove(product_id):
    cart = _get_branch_cart()
    key = str(product_id)
    cart.pop(key, None)
    notes = _get_branch_cart_notes()
    notes.pop(key, None)
    _set_branch_cart(cart)
    _set_branch_cart_notes(notes)
    return redirect(url_for("cart_view"))


@app.route("/order/submit", methods=["POST"])
@login_required
@role_required("branch")
def order_submit():
    user = get_current_user()
    bid = get_current_branch_id()
    if not user or not bid:
        flash("Nejprve zvolte pobočku (nebo musíte být přiřazeni k pobočce).", "error")
        return redirect(url_for("cart_view"))
    cart = _get_branch_cart()
    if not cart:
        flash("Košík je prázdný.", "error")
        return redirect(url_for("cart_view"))
    order_type = session.get("order_type", "normal")
    order = Order(branch_id=bid, status="pending", created_by_id=user.id, order_type=order_type)
    preferred_wh = None
    try:
        preferred_wh = session.get("branch_preferred_warehouse_id")
        if preferred_wh is not None:
            preferred_wh = int(preferred_wh)
    except (TypeError, ValueError):
        preferred_wh = None
    preferred_wh = preferred_wh or _default_warehouse_teplice_id()
    if preferred_wh:
        order.warehouse_id = preferred_wh
    db.session.add(order)
    db.session.flush()
    notes = _get_branch_cart_notes()
    placeholder = _get_or_create_custom_placeholder_product()
    if order_type == "internal":
        for pid_str, qty in cart.items():
            if qty <= 0:
                continue
            pid = int(pid_str)
            ip = InternalProduct.query.get(pid)
            if ip:
                branch_note = notes.get(pid_str) if notes else None
                item = OrderItem(
                    order_id=order.id,
                    product_id=placeholder.id,
                    internal_product_id=ip.id,
                    ordered_quantity=qty,
                    branch_note=branch_note,
                )
                db.session.add(item)
    else:
        for pid_str, qty in cart.items():
            if qty <= 0:
                continue
            pid = int(pid_str)
            p = Product.query.get(pid)
            if p:
                branch_note = notes.get(pid_str) if notes else None
                item = OrderItem(
                    order_id=order.id,
                    product_id=p.id,
                    ordered_quantity=qty,
                    branch_note=branch_note,
                )
                db.session.add(item)
        for c in _get_branch_cart_custom():
            if (c.get("quantity") or 0) <= 0:
                continue
            item = OrderItem(
                order_id=order.id,
                product_id=placeholder.id,
                ordered_quantity=c["quantity"],
                custom_product_name=c.get("name", "").strip() or None,
            )
            db.session.add(item)
    db.session.commit()
    audit_log("order_created", "order", order.id, f"Objednávka #{order.id} pro pobočku {bid}")

    # --- SPUŠTĚNÍ NOTIFIKACE NA POZADÍ ---
    branch = Branch.query.get(bid)
    branch_name = branch.name if branch else f"ID {bid}"

    threading.Thread(
        target=notify_worker,
        args=(order.id, branch_name)
    ).start()

    if order_type == "internal":
        _set_branch_cart({})  # vymaže interní košík (session order_type je stále internal)
    else:
        _set_branch_cart({})
        _set_branch_cart_notes({})
        _set_branch_cart_custom([])
    flash("Objednávka odeslána.")
    if order_type == "internal":
        return redirect(url_for("branch_orders", type="internal"))
    return redirect(url_for("branch_orders"))


@app.route("/branch/move-unavailable-to-cart", methods=["POST"])
@login_required
@role_required("branch")
def branch_move_unavailable_to_cart():
    """Převede položky z poslední objednávky (které sklad neměl) do košíku. Uživatel objednávku vytvoří ručně z košíku."""
    bid = get_current_branch_id()
    if not bid:
        flash("Nejprve zvolte pobočku.", "error")
        return redirect(url_for("branch_dashboard"))
    last_order = Order.query.filter_by(branch_id=bid, order_type="normal").order_by(Order.created_at.desc()).first()
    if not last_order:
        flash("Nemáte žádnou objednávku.", "error")
        return redirect(url_for("branch_dashboard"))
    unavailable_items = [i for i in last_order.items if i.unavailable]
    if not unavailable_items:
        flash("V poslední objednávce není žádné nedodané zboží k převodu.", "error")
        return redirect(url_for("branch_dashboard"))
    _load_branch_cart_from_db()
    cart = _get_branch_cart()
    notes = _get_branch_cart_notes()
    custom = _get_branch_cart_custom()
    cart_internal = session.get("cart_internal_by_branch", {}).get(str(bid), {})
    for i in unavailable_items:
        if i.internal_product_id:
            key = str(i.internal_product_id)
            cart_internal[key] = cart_internal.get(key, 0) + (i.ordered_quantity or 0)
        elif i.product_id:
            key = str(i.product_id)
            cart[key] = cart.get(key, 0) + (i.ordered_quantity or 0)
            if i.branch_note:
                notes[key] = i.branch_note
        elif i.custom_product_name:
            custom.append({"name": i.custom_product_name, "quantity": i.ordered_quantity or 0})
    _set_branch_cart(cart)
    _set_branch_cart_notes(notes)
    _set_branch_cart_custom(custom)
    session.setdefault("cart_internal_by_branch", {})[str(bid)] = cart_internal
    _save_branch_cart_to_db()
    flash(f"Položky z nedodaného zboží ({len(unavailable_items)}) vloženy do košíku. Dokončete objednávku v Košíku.")
    return redirect(url_for("cart_view"))


@app.route("/orders")
@login_required
@role_required("branch")
def branch_orders():
    user = get_current_user()
    bid = get_current_branch_id()
    order_type_filter = request.args.get("type", "normal").strip() or "normal"
    if not bid:
        orders = []
    else:
        orders = Order.query.filter_by(branch_id=bid, order_type=order_type_filter).order_by(Order.created_at.desc()).all()
    return render_template("branch_orders.html", orders=orders, user=user, order_type_filter=order_type_filter)


@app.route("/orders/<int:order_id>")
@login_required
@role_required("branch")
def branch_order_detail(order_id):
    user = get_current_user()
    order = Order.query.get_or_404(order_id)
    if order.branch_id != get_current_branch_id():
        flash("Objednávka nepatří vaší pobočce.", "error")
        return redirect(url_for("branch_orders"))
    total_qty = sum((i.ordered_quantity or 0) for i in order.items)
    shipped_qty = sum((i.shipped_quantity or 0) for i in order.items)
    order_audit_log = (
        AuditLog.query.filter_by(entity_type="order", entity_id=order.id)
        .order_by(AuditLog.created_at.desc())
        .limit(30)
        .all()
    )
    order_total_selling = _order_total_selling(order)
    return render_template(
        "branch_order_detail.html",
        order=order,
        user=user,
        status_cz=status_cz,
        total_qty=total_qty,
        shipped_qty=shipped_qty,
        order_audit_log=order_audit_log,
        order_total_selling=order_total_selling,
    )


@app.route("/admin")
@login_required
@role_required("admin")
def admin_dashboard():
    from sqlalchemy import func
    total_orders = Order.query.count()
    by_status = (
        db.session.query(Order.status, func.count(Order.id))
        .group_by(Order.status)
        .all()
    )
    status_counts = {s: c for s, c in by_status}
    top_products = (
        db.session.query(Product.id, Product.nazev, Product.kod_zbozi, Product.mj, func.sum(OrderItem.ordered_quantity).label("total"))
        .join(OrderItem, OrderItem.product_id == Product.id)
        .group_by(Product.id, Product.nazev, Product.kod_zbozi, Product.mj)
        .order_by(func.sum(OrderItem.ordered_quantity).desc())
        .limit(12)
        .all()
    )
    recent_orders = Order.query.order_by(Order.created_at.desc()).limit(4).all()
    return render_template(
        "admin_dashboard.html",
        user=get_current_user(),
        total_orders=total_orders,
        status_counts=status_counts,
        top_products=top_products,
        recent_orders=recent_orders,
        status_cz=status_cz,
    )


@app.route("/admin/act-as", methods=["GET", "POST"])
@login_required
@role_required("admin")
def admin_act_as():
    """Přepnutí admina do režimu pobočky nebo skladu (vše s audit logem)."""
    if request.method == "POST":
        try:
            action = request.form.get("action")
            if action == "branch":
                branch_id = request.form.get("branch_id", type=int)
                branch = Branch.query.get(branch_id) if branch_id else None
                if branch:
                    session["acting_as_role"] = "branch"
                    session["acting_as_branch_id"] = branch.id
                    audit_log("admin_act_as", None, None, f"Admin působí jako pobočka: {branch.name} (#{branch.id})")
                    flash(f"Režim: působíte jako pobočka {branch.name}. Všechny akce se zapisují do audit logu.")
                    return redirect(url_for("branch_dashboard"))
                flash("Zvolte pobočku.", "error")
            elif action == "warehouse":
                warehouse_id = request.form.get("warehouse_id", type=int)
                warehouse = Warehouse.query.get(warehouse_id) if warehouse_id else None
                if warehouse:
                    session["acting_as_role"] = "warehouse"
                    session["acting_as_branch_id"] = None
                    session["acting_as_warehouse_id"] = warehouse.id
                    audit_log("admin_act_as", None, None, f"Admin působí jako sklad: {warehouse.name} (#{warehouse.id})")
                    flash(f"Režim: působíte jako sklad {warehouse.name}. Všechny akce se zapisují do audit logu.")
                    return redirect(url_for("warehouse_dashboard"))
                flash("Zvolte sklad.", "error")
            elif action == "end":
                session.pop("acting_as_role", None)
                session.pop("acting_as_branch_id", None)
                session.pop("acting_as_warehouse_id", None)
                audit_log("admin_act_as", None, None, "Admin ukončil režim pobočka/sklad")
                flash("Režim ukončen.")
                return redirect(url_for("admin_dashboard"))
        except Exception as e:
            db.session.rollback()
            flash(f"Přepnutí režimu se nezdařilo: {e}", "error")
            return redirect(url_for("admin_act_as"))
    branches = Branch.query.order_by(Branch.name).all()
    warehouses = Warehouse.query.order_by(Warehouse.name).all()
    return render_template("admin_act_as.html", branches=branches, warehouses=warehouses, user=get_current_user())


@app.route("/admin/back")
@login_required
@role_required("admin")
def admin_back():
    """Vrátí admina z režimu „působí jako“ zpět na admin dashboard."""
    session.pop("acting_as_role", None)
    session.pop("acting_as_branch_id", None)
    session.pop("acting_as_warehouse_id", None)
    audit_log("admin_act_as", None, None, "Admin se vrátil do admin režimu")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/management")
@login_required
@role_required("admin")
def admin_management():
    """Mezistránka Správa: odkazy na Uživatelé, Pobočky, Sklady, Dodavatelé."""
    return render_template("admin/management.html", user=get_current_user())


@app.route("/admin/orders")
@login_required
@role_required("admin")
def admin_orders():
    status_filter = request.args.get("status", "").strip()
    branch_filter = request.args.get("branch_id", type=int)
    warehouse_filter = request.args.get("warehouse_id", type=int)
    q = Order.query
    if status_filter:
        if status_filter == "shipped":
            # Zobraz i objednávky fakticky odeslané (všechny položky odeslány), i když status zůstal partially_shipped
            has_incomplete = db.session.query(OrderItem.id).filter(
                OrderItem.order_id == Order.id,
                db.func.coalesce(OrderItem.shipped_quantity, 0) < OrderItem.ordered_quantity,
            ).exists()
            q = q.filter(
                db.or_(
                    Order.status.in_(["shipped", "verified", "error"]),
                    (Order.status == "partially_shipped") & ~has_incomplete,
                )
            )
        else:
            q = q.filter(Order.status == status_filter)
    if branch_filter:
        q = q.filter(Order.branch_id == branch_filter)
    if warehouse_filter:
        q = q.filter(Order.warehouse_id == warehouse_filter)
    orders = q.order_by(Order.created_at.desc()).all()
    branches = Branch.query.order_by(Branch.name).all()
    warehouses = Warehouse.query.order_by(Warehouse.name).all()
    return render_template(
        "admin_orders.html",
        orders=orders,
        user=get_current_user(),
        status_filter=status_filter,
        branch_filter=branch_filter,
        warehouse_filter=warehouse_filter,
        branches=branches,
        warehouses=warehouses,
    )


@app.route("/admin/orders/<int:order_id>")
@login_required
@role_required("admin")
def admin_order_detail(order_id):
    order = Order.query.get_or_404(order_id)
    total_qty = sum((i.ordered_quantity or 0) for i in order.items)
    shipped_qty = sum((i.shipped_quantity or 0) for i in order.items)
    order_audit_log = (
        AuditLog.query.filter_by(entity_type="order", entity_id=order.id)
        .order_by(AuditLog.created_at.desc())
        .limit(30)
        .all()
    )
    order_total_selling = _order_total_selling(order)
    warehouses = Warehouse.query.order_by(Warehouse.name).all()
    order_check_status = getattr(order, "check_status", None) or ""
    check_ok_items, check_error_items = _order_check_results(order) if order_check_status in ("verified", "error") else ([], [])
    item_scanned_qty = _order_item_scanned_qtys(order)
    return render_template(
        "admin_order_detail.html",
        order=order,
        user=get_current_user(),
        status_cz=status_cz,
        total_qty=total_qty,
        shipped_qty=shipped_qty,
        order_audit_log=order_audit_log,
        order_total_selling=order_total_selling,
        warehouses=warehouses,
        check_ok_items=check_ok_items,
        check_error_items=check_error_items,
        order_check_status=order_check_status,
        item_scanned_qty=item_scanned_qty,
    )


@app.route("/admin/orders/<int:order_id>/set-warehouse", methods=["POST"])
@login_required
@role_required("admin")
def admin_order_set_warehouse(order_id):
    """Nastavení skladu odeslání (shipping warehouse) u objednávky."""
    order = Order.query.get_or_404(order_id)
    wh_id = request.form.get("warehouse_id", type=int)
    if wh_id is not None:
        w = Warehouse.query.get(wh_id)
        if w:
            order.warehouse_id = w.id
            db.session.commit()
            audit_log("order_updated", "order", order.id, f"Admin nastavil sklad odeslání: {w.name}")
            flash("Sklad odeslání byl upraven.")
        else:
            flash("Zvolený sklad neexistuje.", "error")
    else:
        order.warehouse_id = None
        db.session.commit()
        flash("Sklad odeslání byl odebrán.")
    return redirect(url_for("admin_order_detail", order_id=order_id))


@app.route("/admin/orders/<int:order_id>/delete", methods=["POST"])
@login_required
@role_required("admin")
def admin_order_delete(order_id):
    order = Order.query.get_or_404(order_id)
    if order.status != "pending":
        flash("Lze mazat pouze objednávky se stavem „Čeká“.", "error")
        return redirect(url_for("admin_order_detail", order_id=order_id))
    oid = order.id
    for item in order.items:
        db.session.delete(item)
    db.session.delete(order)
    db.session.commit()
    audit_log("order_deleted", "order", oid, f"Admin smazal objednávku #{oid}")
    flash("Objednávka byla smazána.")
    return redirect(url_for("admin_orders"))


@app.route("/admin/monthly-overview")
@login_required
@role_required("admin")
def admin_monthly_overview():
    """Měsíční přehled nákupních nákladů (nc) podle skladu a pobočky. Zahrnuje všechny objednávky v daném měsíci (včetně částečně odeslaných); do sumy nc jdou jen položky s odeslaným množstvím (shipped_quantity > 0)."""
    from collections import defaultdict
    year = request.args.get("year", type=int)
    month = request.args.get("month", type=int)
    if not year or not month:
        now = datetime.now()
        year = year or now.year
        month = month or now.month
    month_start = datetime(year, month, 1)
    if month == 12:
        month_end = datetime(year + 1, 1, 1)
    else:
        month_end = datetime(year, month + 1, 1)
    # Všechny objednávky v měsíci (bez filtru na status) – částečně odeslané se zahrnou
    orders = Order.query.filter(Order.created_at >= month_start, Order.created_at < month_end).all()
    # (warehouse_id, branch_id) -> total nc a total pc
    totals_nc = defaultdict(float)
    totals_pc = defaultdict(float)
    warehouse_names = {}
    branch_names = {}
    for order in orders:
        wh_id = order.warehouse_id
        br_id = order.branch_id
        if wh_id and wh_id not in warehouse_names:
            w = Warehouse.query.get(wh_id)
            warehouse_names[wh_id] = w.name if w else f"ID {wh_id}"
        if br_id not in branch_names:
            b = Branch.query.get(br_id)
            branch_names[br_id] = b.name if b else f"ID {br_id}"
        for item in order.items:
            if (item.shipped_quantity or 0) <= 0 or not item.product:
                continue
            key = (wh_id, br_id)
            if item.product.nc is not None:
                totals_nc[key] += (item.product.nc or 0) * (item.shipped_quantity or 0)
            if item.product.pc is not None:
                totals_pc[key] += (item.product.pc or 0) * (item.shipped_quantity or 0)
    keys_sorted = sorted(set(totals_nc.keys()) | set(totals_pc.keys()), key=lambda x: (str(x[0] or ""), x[1]))
    rows = []
    for (wh_id, br_id) in keys_sorted:
        rows.append({
            "warehouse_id": wh_id,
            "warehouse_name": warehouse_names.get(wh_id) or "—",
            "branch_id": br_id,
            "branch_name": branch_names.get(br_id) or "—",
            "total_nc": totals_nc.get((wh_id, br_id), 0),
            "total_pc": totals_pc.get((wh_id, br_id), 0),
        })
    # Příjmy zboží v daném měsíci – seskupené podle dodavatele (supplier)
    receipts = (
        GoodsReceipt.query.filter(
            GoodsReceipt.received_at >= month_start,
            GoodsReceipt.received_at < month_end,
        )
        .options(joinedload(GoodsReceipt.supplier), joinedload(GoodsReceipt.items).joinedload(GoodsReceiptItem.product))
        .order_by(GoodsReceipt.received_at)
        .all()
    )
    # supplier_id -> { supplier_name, receipts: [ { id, label, total_nc, total_pc }, ... ], total_nc, total_pc }
    receipt_by_supplier = defaultdict(lambda: {"supplier_name": "—", "receipts": [], "total_nc": 0.0, "total_pc": 0.0})
    for rec in receipts:
        sup_id = rec.supplier_id
        key = sup_id if sup_id is not None else -1
        if key not in receipt_by_supplier or receipt_by_supplier[key]["supplier_name"] == "—":
            receipt_by_supplier[key]["supplier_name"] = rec.supplier.supplier_name if rec.supplier else "Bez dodavatele"
        rec_nc = sum((it.product.nc or 0) * (it.quantity or 0) for it in rec.items if it.product)
        rec_pc = sum((it.product.pc or 0) * (it.quantity or 0) for it in rec.items if it.product)
        receipt_by_supplier[key]["receipts"].append({
            "id": rec.id,
            "label": f"Příjem #{rec.id}",
            "received_at": rec.received_at,
            "total_nc": rec_nc,
            "total_pc": rec_pc,
        })
        receipt_by_supplier[key]["total_nc"] += rec_nc
        receipt_by_supplier[key]["total_pc"] += rec_pc
    receipt_groups = []
    for sup_id in sorted(receipt_by_supplier.keys(), key=lambda x: (x == -1, receipt_by_supplier[x]["supplier_name"])):
        receipt_groups.append({
            "supplier_id": sup_id if sup_id != -1 else None,
            "supplier_name": receipt_by_supplier[sup_id]["supplier_name"],
            "receipts": receipt_by_supplier[sup_id]["receipts"],
            "total_nc": receipt_by_supplier[sup_id]["total_nc"],
            "total_pc": receipt_by_supplier[sup_id]["total_pc"],
        })
    # Zpětná kompatibilita: receipt_rows pro staré šablony (agregace po skladu) – nepoužíváme v nové šabloně
    receipt_rows = []
    years = list(range(datetime.now().year, datetime.now().year - 5, -1))
    months = list(range(1, 13))
    return render_template(
        "admin/monthly_overview.html",
        user=get_current_user(),
        rows=rows,
        receipt_rows=receipt_rows,
        receipt_groups=receipt_groups,
        year=year,
        month=month,
        years=years,
        months=months,
    )


@app.route("/admin/goods-receipts")
@login_required
@role_required("admin")
def admin_goods_receipts():
    """Seznam všech příjmů zboží (všechny sklady). Filtr podle dodavatele."""
    query = (
        GoodsReceipt.query
        .options(joinedload(GoodsReceipt.warehouse), joinedload(GoodsReceipt.supplier), selectinload(GoodsReceipt.items))
        .order_by(GoodsReceipt.received_at.desc())
    )
    supplier_id = request.args.get("supplier_id", type=int)
    if supplier_id is not None:
        query = query.filter(GoodsReceipt.supplier_id == supplier_id)
    receipts = query.all()
    suppliers = Supplier.query.order_by(Supplier.supplier_name).all()
    return render_template(
        "admin/goods_receipts_list.html",
        receipts=receipts,
        suppliers=suppliers,
        selected_supplier_id=supplier_id,
        user=get_current_user(),
    )


@app.route("/admin/goods-receipts/<int:receipt_id>")
@login_required
@role_required("admin")
def admin_goods_receipt_detail(receipt_id):
    """Detail příjmu zboží. Součty: total_nc = sum(nc × množství), total_pc = sum(pc × množství)."""
    receipt = GoodsReceipt.query.get_or_404(receipt_id)
    total_nc = sum((it.product.nc or 0) * (it.quantity or 0) for it in receipt.items if it.product)
    total_pc = sum((it.product.pc or 0) * (it.quantity or 0) for it in receipt.items if it.product)
    return render_template(
        "admin/goods_receipt_detail.html",
        receipt=receipt,
        total_nc=total_nc,
        total_pc=total_pc,
        user=get_current_user(),
    )


@app.route("/admin/goods-receipts/<int:receipt_id>/delete", methods=["POST"])
@login_required
@role_required("admin")
def admin_goods_receipt_delete(receipt_id):
    """Smazání příjmu zboží (cascade na položky)."""
    receipt = GoodsReceipt.query.get_or_404(receipt_id)
    rid = receipt.id
    db.session.delete(receipt)
    db.session.commit()
    audit_log("goods_receipt_deleted", "goods_receipt", rid, f"Admin smazal příjem #{rid}")
    flash("Příjem zboží byl smazán.")
    return redirect(url_for("admin_goods_receipts"))


@app.route("/admin/users", methods=["GET", "POST"])
@login_required
@role_required("admin")
def admin_users():
    flash("Správa uživatelů je v centrálním auth-system (Směrovač → Správa uživatelů).", "info")
    return redirect(url_for("admin_management"))


@app.route("/admin/users/<int:user_id>/edit", methods=["POST"])
@login_required
@role_required("admin")
def admin_user_edit(user_id):
    flash("Správa uživatelů je v centrálním auth-system (Směrovač → Správa uživatelů).", "info")
    return redirect(url_for("admin_management"))


@app.route("/admin/users/<int:user_id>/delete", methods=["POST"])
@login_required
@role_required("admin")
def admin_user_delete(user_id):
    flash("Správa uživatelů je v centrálním auth-system (Směrovač → Správa uživatelů).", "info")
    return redirect(url_for("admin_management"))


@app.route("/admin/branches", methods=["GET", "POST"])
@login_required
@role_required("admin")
def admin_branches():
    flash("Pobočky a sklady se spravují v centrálním systému (Směrovač → Správa uživatelů → Pobočky / Sklady).", "info")
    return redirect(url_for("admin_management"))


@app.route("/admin/branches/<int:branch_id>/edit", methods=["POST"])
@login_required
@role_required("admin")
def admin_branch_edit(branch_id):
    flash("Pobočky a sklady se spravují v centrálním systému (Směrovač → Správa uživatelů).", "info")
    return redirect(url_for("admin_management"))


@app.route("/admin/branches/<int:branch_id>/delete", methods=["POST"])
@login_required
@role_required("admin")
def admin_branch_delete(branch_id):
    flash("Pobočky a sklady se spravují v centrálním systému (Směrovač → Správa uživatelů).", "info")
    return redirect(url_for("admin_management"))


@app.route("/admin/warehouses", methods=["GET", "POST"])
@login_required
@role_required("admin")
def admin_warehouses():
    flash("Pobočky a sklady se spravují v centrálním systému (Směrovač → Správa uživatelů → Pobočky / Sklady).", "info")
    return redirect(url_for("admin_management"))


@app.route("/admin/warehouses/<int:warehouse_id>/edit", methods=["POST"])
@login_required
@role_required("admin")
def admin_warehouse_edit(warehouse_id):
    flash("Pobočky a sklady se spravují v centrálním systému (Směrovač → Správa uživatelů).", "info")
    return redirect(url_for("admin_management"))


@app.route("/admin/warehouses/<int:warehouse_id>/delete", methods=["POST"])
@login_required
@role_required("admin")
def admin_warehouse_delete(warehouse_id):
    flash("Pobočky a sklady se spravují v centrálním systému (Směrovač → Správa uživatelů).", "info")
    return redirect(url_for("admin_management"))


@app.route("/admin/products")
@login_required
@role_required("admin")
def admin_products():
    """Admin přehled produktů s filtry: skupina, značka, 10mg/20mg, interní. Stránkování po 50."""
    from sqlalchemy import or_
    q = request.args.get("q", "").strip()
    group_filter = request.args.get("group", "").strip() or None
    brand_filter = request.args.get("brand", "").strip() or None
    category_filter = request.args.get("category", "").strip() or None  # 10mg | 20mg
    internal_filter = request.args.get("internal", "").strip() or None  # "0" = běžné, "1" = interní, None = vše
    page = request.args.get("page", 1, type=int)
    per_page = 50
    if page < 1:
        page = 1
    base = Product.query.filter(Product.nazev != "[Vlastní – produkt mimo katalog]")
    if internal_filter == "1":
        base = base.filter(Product.is_internal == True)
    elif internal_filter == "0":
        base = base.filter(db.or_(Product.is_internal == False, Product.is_internal.is_(None)))
    if group_filter:
        base = base.filter(Product.naz_skup == group_filter)
    if category_filter and category_filter in PRODUCT_MG_OPTIONS:
        base = base.filter(Product.nazev.ilike(f"%{category_filter}%"))
    if brand_filter:
        base = base.filter(db.or_(Product.nazev.ilike(brand_filter + " %"), Product.nazev == brand_filter))
    if q:
        q_like = f"%{q}%"
        ean_in_extra = db.session.query(ProductEan.id).filter(
            ProductEan.product_id == Product.id, ProductEan.ean.ilike(q_like)
        ).exists()
        base = base.filter(
            or_(
                Product.nazev.ilike(q_like),
                Product.kod_zbozi.ilike(q_like),
                (Product.naz_skup.isnot(None)) & (Product.naz_skup.ilike(q_like)),
                (Product.ean.isnot(None)) & (Product.ean.ilike(q_like)),
                ean_in_extra,
            )
        )
    base = base.order_by(Product.naz_skup, Product.nazev)
    total = base.count()
    total_pages = max(1, (total + per_page - 1) // per_page) if total else 1
    if page > total_pages:
        page = total_pages
    products = base.offset((page - 1) * per_page).limit(per_page).all()
    # Dropdown možnosti z celého katalogu (ne jen z jedné stránky)
    catalog = Product.query.filter(Product.nazev != "[Vlastní – produkt mimo katalog]")
    if internal_filter == "1":
        catalog = catalog.filter(Product.is_internal == True)
    elif internal_filter == "0":
        catalog = catalog.filter(db.or_(Product.is_internal == False, Product.is_internal.is_(None)))
    groups = sorted({r[0] for r in catalog.with_entities(Product.naz_skup).distinct().all() if r and r[0]})
    nazev_for_brands = catalog.with_entities(Product.nazev).distinct().limit(15000).all()
    brands = sorted({_product_brand(r[0]) for r in nazev_for_brands if r and r[0] and _product_brand(r[0])})
    categories = PRODUCT_MG_OPTIONS.copy()
    return render_template(
        "admin_products.html",
        products=products,
        user=get_current_user(),
        search_q=q,
        groups=groups,
        group_filter=group_filter,
        brands=brands,
        brand_filter=brand_filter,
        categories=categories,
        category_filter=category_filter,
        internal_filter=internal_filter,
        page=page,
        total_pages=total_pages,
        total=total,
    )


@app.route("/admin/products-internal")
@login_required
@role_required("admin")
def admin_internal_products():
    """Interní produkty (kancelář) – samostatná stránka."""
    products = Product.query.filter(Product.is_internal == True).order_by(Product.naz_skup, Product.nazev).all()
    return render_template("admin_internal_products.html", products=products, user=get_current_user())


@app.route("/admin/products/<int:product_id>/set-internal", methods=["POST"])
@login_required
@role_required("admin")
def admin_product_set_internal(product_id):
    product = Product.query.get_or_404(product_id)
    val = request.form.get("internal")
    product.is_internal = str(val).strip().lower() in ("1", "true", "yes", "on")
    db.session.commit()
    audit_log("product_updated", "product", product.id, f"is_internal={product.is_internal}")
    if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.accept_mimetypes.best == "application/json":
        return jsonify({"ok": True, "is_internal": product.is_internal})
    flash("Produkt upraven (interní)." if product.is_internal else "Produkt upraven (běžný).")
    return redirect(request.referrer or url_for("admin_products"))


@app.route("/admin/products/<int:product_id>/add-ean", methods=["POST"])
@login_required
@role_required("admin")
def admin_product_add_ean(product_id):
    product = Product.query.get_or_404(product_id)
    ean = (request.form.get("ean") or "").strip()
    if not ean:
        flash("EAN je povinný.", "error")
        return redirect(request.referrer or url_for("admin_products"))
    if ProductEan.query.filter(ProductEan.ean == ean).first() or Product.query.filter(Product.ean == ean).first():
        flash("Tento EAN již patří jinému produktu nebo je hlavní EAN.", "error")
        return redirect(request.referrer or url_for("admin_products"))
    pe = ProductEan(product_id=product.id, ean=ean)
    db.session.add(pe)
    db.session.commit()
    audit_log("product_ean_added", "product", product.id, f"ean={ean}")
    flash("EAN přidán.")
    return redirect(request.referrer or url_for("admin_products"))


@app.route("/admin/products/<int:product_id>/remove-ean/<int:ean_id>", methods=["POST"])
@login_required
@role_required("admin")
def admin_product_remove_ean(product_id, ean_id):
    product = Product.query.get_or_404(product_id)
    pe = ProductEan.query.filter(ProductEan.id == ean_id, ProductEan.product_id == product.id).first()
    if pe:
        db.session.delete(pe)
        db.session.commit()
        audit_log("product_ean_removed", "product", product.id, f"ean={pe.ean}")
        flash("EAN odebrán.")
    return redirect(request.referrer or url_for("admin_products"))


def _norm(val):
    """Normalizuje hodnotu z buňky na řetězec (Excel může vracet float)."""
    if val is None:
        return None
    if isinstance(val, float):
        if val == int(val):
            val = int(val)
    s = str(val).strip()
    return s if s else None


def _normalize_header_name(s):
    """Převede název sloupce na ASCII pro spolehlivé mapování (odstraní diakritiku)."""
    if not s:
        return s
    s = str(s).strip().lower().replace(" ", "_").replace("-", "_")
    trans = {
        "á": "a", "č": "c", "ď": "d", "é": "e", "ě": "e", "í": "i", "ň": "n", "ó": "o",
        "ř": "r", "š": "s", "ť": "t", "ú": "u", "ů": "u", "ý": "y", "ž": "z",
    }
    for cz, ascii_ in trans.items():
        s = s.replace(cz, ascii_)
    return s


def _validate_katalog_xls_headers(path):
    """Ověří, že katalog.xlsx má hlavičky nazev nebo kod_zbozi a alespoň jeden sloupec EAN."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    headers_raw = [c.value for c in ws[1]]
    wb.close()
    headers = [_normalize_header_name(str(h or "")) for h in headers_raw]
    has_id = "nazev" in headers or "kod_zbozi" in headers
    ean_cols = [h for h in headers if h and (h == "ean" or (h.startswith("ean") and (len(h) == 3 or (len(h) > 3 and h[3:].replace("_", "").isdigit()))))]
    if not ean_cols:
        ean_cols = [h for h in headers if h and "ean" in h.lower()]
    return has_id and len(ean_cols) > 0


def _parse_katalog_xlsx(path):
    """Parsuje katalog.xlsx: první řádek hlavičky (normalizované). Formát: naz_skupiny, kod_zbozi, nazev, ean, mj, nc, pc.
    Jeden EAN na řádek; více řádků se stejným produktem = více EAN. Match produktu: kod_zbozi nebo nazev.
    Vrací též naz_skup, mj, nc, pc pro plný import produktů."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    headers_raw = [c.value for c in ws[1]]
    headers = [_normalize_header_name(str(h or "")) for h in headers_raw]
    if "naz_skupiny" in headers and "naz_skup" not in headers:
        headers = ["naz_skup" if h == "naz_skupiny" else h for h in headers]
    ean_cols = [i for i, h in enumerate(headers) if h and (h == "ean" or (h.startswith("ean") and (len(h) == 3 or (len(h) > 3 and h[3:].replace("_", "").isdigit()))))]
    if not ean_cols:
        ean_cols = [i for i, h in enumerate(headers) if h and "ean" in h.lower()]
    idx_nazev = next((i for i, h in enumerate(headers) if h == "nazev"), None)
    idx_kod = next((i for i, h in enumerate(headers) if h == "kod_zbozi"), None)
    # Skupina: sloupec naz_skup
    idx_naz_skup = next((i for i, h in enumerate(headers) if h == "naz_skup"), None) or next((i for i, h in enumerate(headers) if h == "naz_skupiny"), None)
    idx_mj = next((i for i, h in enumerate(headers) if h == "mj"), None)
    idx_nc = next((i for i, h in enumerate(headers) if h == "nc"), None)
    idx_pc = next((i for i, h in enumerate(headers) if h == "pc"), None)
    rows = []
    for row in ws.iter_rows(min_row=2):
        vals = [c.value for c in row]
        need_len = max((idx_nazev or 0), (idx_kod or 0), (idx_naz_skup or 0), (idx_mj or 0), (idx_nc or 0), (idx_pc or 0), *ean_cols)
        if len(vals) <= need_len:
            continue
        nazev = _norm(vals[idx_nazev]) if idx_nazev is not None and idx_nazev < len(vals) else None
        kod_zbozi = _norm(vals[idx_kod]) if idx_kod is not None and idx_kod < len(vals) else None
        naz_skup = _norm(vals[idx_naz_skup]) if idx_naz_skup is not None and idx_naz_skup < len(vals) else None
        mj = _norm(vals[idx_mj]) if idx_mj is not None and idx_mj < len(vals) else None
        nc = _pc_to_float(vals[idx_nc]) if idx_nc is not None and idx_nc < len(vals) else None
        pc = _pc_to_float(vals[idx_pc]) if idx_pc is not None and idx_pc < len(vals) else None
        eans = []
        for i in ean_cols:
            if i < len(vals):
                v = _norm(vals[i])
                if v:
                    eans.append(v)
        if not kod_zbozi and not nazev:
            continue
        rows.append({
            "kod_zbozi": kod_zbozi, "nazev": nazev, "eans": eans,
            "naz_skup": naz_skup, "mj": mj, "nc": nc, "pc": pc,
        })
    wb.close()
    return rows


def _pc_to_float(val):
    """Převede pc (řetězec nebo číslo) na float pro Product.pc."""
    if val is None:
        return None
    try:
        return float(str(val).strip().replace(",", "."))
    except (ValueError, TypeError):
        return None


def _product_field_empty(val):
    """True pokud je hodnota považována za prázdnou (None nebo prázdný řetězec)."""
    if val is None:
        return True
    if isinstance(val, (int, float)) and val == 0:
        return True
    return not str(val).strip()


def _update_product_only_missing(existing, name, sku, unit, group_name, pc_float, ean, nc_float=None):
    """Aktualizuje existující produkt: prázdná pole doplní; naz_skup se při importu vždy přepíše, pokud je v souboru."""
    if name and _product_field_empty(existing.nazev):
        existing.nazev = name
    if sku is not None and _product_field_empty(existing.kod_zbozi):
        existing.kod_zbozi = sku
    if unit is not None and _product_field_empty(existing.mj):
        existing.mj = unit
    if group_name is not None:
        existing.naz_skup = group_name
    if pc_float is not None and existing.pc is None:
        existing.pc = pc_float
    if ean is not None and _product_field_empty(existing.ean):
        existing.ean = ean
    if nc_float is not None and existing.nc is None:
        existing.nc = nc_float


def _import_csv(path):
    import csv
    added, updated, skipped, errors = 0, 0, 0, 0
    with open(path, "r", encoding="utf-8-sig") as f:
        first_line = f.readline()
        f.seek(0)
        # Podpora tabulátoru (TSV): název_skupiny\tnázev\tsku\tpc
        delimiter = "\t" if "\t" in first_line and "název" in first_line.lower() else ","
        reader = csv.DictReader(f, delimiter=delimiter)
        for row in reader:
            name = (row.get("název") or row.get("nazev") or row.get("name") or "").strip()
            if not name:
                skipped += 1
                continue
            try:
                sku = _norm(row.get("sku") or row.get("code") or row.get("kod") or row.get("kod_zbozi"))
                unit = _norm(row.get("ks") or row.get("mj"))  # jednotka: ks, ml, balení
                group_name = _norm(
                    row.get("naz_skup") or row.get("naz_skupiny") or row.get("název_skupiny") or row.get("nazev_skupiny") or row.get("group_name")
                )
                pc = _norm(row.get("pc"))
                ean = _norm(row.get("ean"))
                nc_float = _pc_to_float(row.get("nc"))
                # Párování: nejdřív podle EAN (Product.ean i ProductEan), pak kod_zbozi, pak název
                existing = None
                if ean:
                    existing = find_product_by_ean(ean)
                if not existing and sku:
                    existing = Product.query.filter_by(kod_zbozi=sku).first()
                if not existing:
                    existing = Product.query.filter(Product.nazev == name).first()
                if existing:
                    _update_product_only_missing(
                        existing, name, sku, unit, group_name, _pc_to_float(pc), ean, nc_float
                    )
                    updated += 1
                else:
                    db.session.add(Product(
                        nazev=name, kod_zbozi=sku, mj=unit, naz_skup=group_name,
                        pc=_pc_to_float(pc), ean=ean, nc=nc_float
                    ))
                    added += 1
                db.session.commit()
            except Exception:
                db.session.rollback()
                errors += 1
    return added, updated, skipped, errors


def _import_internal_csv(path):
    """Import do tabulky internal_products (stejný formát CSV jako běžné produkty)."""
    import csv
    added, updated, skipped, errors = 0, 0, 0, 0
    with open(path, "r", encoding="utf-8-sig") as f:
        first_line = f.readline()
        f.seek(0)
        delimiter = "\t" if "\t" in first_line and "název" in first_line.lower() else ","
        reader = csv.DictReader(f, delimiter=delimiter)
        for row in reader:
            name = (row.get("název") or row.get("nazev") or row.get("name") or "").strip()
            if not name:
                skipped += 1
                continue
            try:
                sku = _norm(row.get("sku") or row.get("code") or row.get("kod"))
                unit = _norm(row.get("ks"))
                group_name = _norm(row.get("naz_skup") or row.get("naz_skupiny") or row.get("název_skupiny") or row.get("nazev_skupiny") or row.get("group_name"))
                pc = _norm(row.get("pc"))
                existing = InternalProduct.query.filter_by(sku=sku).first() if sku else None
                if not existing:
                    existing = InternalProduct.query.filter(InternalProduct.name == name).first()
                if existing:
                    existing.name = name
                    existing.sku = sku
                    existing.unit = unit
                    existing.group_name = group_name
                    existing.pc = pc
                    updated += 1
                else:
                    db.session.add(InternalProduct(name=name, sku=sku, unit=unit, group_name=group_name, pc=pc))
                    added += 1
                db.session.commit()
            except Exception:
                db.session.rollback()
                errors += 1
    return added, updated, skipped, errors


def _excel_row_to_product(d):
    """Z řádku Excelu vrátí (name, sku, unit, group_name, pc, ean, nc_float) nebo None. Podporuje hlavičky nazev, kod_zbozi, ean, mj, nc, pc, naz_skupiny."""
    name = _norm(
        d.get("název") or d.get("nazev") or d.get("name") or d.get("col1")
    )
    if not name:
        return None
    sku = _norm(d.get("sku") or d.get("code") or d.get("kod") or d.get("kod_zbozi"))
    unit = _norm(d.get("ks") or d.get("mj"))
    if not unit and d.get("col2") is not None:
        unit = _norm(d.get("col2"))
    group_name = _norm(
        d.get("naz_skup") or d.get("naz_skupiny") or d.get("název_skupiny") or d.get("nazev_skupiny") or d.get("group_name") or d.get("col0")
    )
    pc = _norm(d.get("pc") or d.get("col3"))
    ean = _norm(d.get("ean"))
    nc_float = _pc_to_float(d.get("nc"))
    return (name, sku, unit, group_name, pc, ean, nc_float)


def _import_excel(path):
    added, updated, skipped, errors = 0, 0, 0, 0
    error_reasons = []

    def process_row(d, row_num):
        nonlocal added, updated, skipped, errors, error_reasons
        t = _excel_row_to_product(d)
        if not t:
            skipped += 1
            return
        name, sku, unit, group_name, pc, ean, nc_float = t
        try:
            # Párování: nejdřív podle EAN (Product.ean i ProductEan), pak kod_zbozi, pak název
            existing = None
            if ean:
                existing = find_product_by_ean(ean)
            if not existing and sku:
                existing = Product.query.filter_by(kod_zbozi=sku).first()
            if not existing:
                existing = Product.query.filter(Product.nazev == name).first()
            if existing:
                _update_product_only_missing(
                    existing, name, sku, unit, group_name, _pc_to_float(pc), ean, nc_float
                )
                updated += 1
            else:
                db.session.add(Product(
                    nazev=name, kod_zbozi=sku, mj=unit, naz_skup=group_name,
                    pc=_pc_to_float(pc), ean=ean, nc=nc_float
                ))
                added += 1
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            errors += 1
            error_reasons.append((row_num, str(e)))

    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        headers_raw = [c.value for c in ws[1]]
        headers = [str(h or "").strip().lower().replace(" ", "_").replace("-", "_") for h in headers_raw]
        ncols = len(headers)
        for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
            vals = [c.value for c in row]
            d = dict(zip(headers, vals))
            for i in range(4):
                d["col%d" % i] = vals[i] if i < len(vals) else None
            process_row(d, row_num)
        wb.close()
    except Exception:
        import xlrd
        wb = xlrd.open_workbook(path)
        sheet = wb.sheet_by_index(0)
        ncols = sheet.ncols
        headers = []
        for c in range(ncols):
            v = sheet.cell_value(0, c)
            headers.append(str(v or "").strip().lower().replace(" ", "_").replace("-", "_"))
        for r in range(1, sheet.nrows):
            d = {}
            for c in range(ncols):
                val = sheet.cell_value(r, c)
                if headers[c]:
                    d[headers[c]] = val
                if c < 4:
                    d["col%d" % c] = val
            process_row(d, r + 1)
    return added, updated, skipped, errors, error_reasons


@app.route("/admin/import", methods=["GET", "POST"])
@login_required
@role_required("admin")
def admin_import():
    if request.method == "POST":
        f = request.files.get("file")
        if not f or not f.filename:
            flash("Vyberte soubor.", "error")
            return redirect(url_for("admin_import"))
        ext = os.path.splitext(f.filename)[1].lower()
        path = os.path.join(app.config["UPLOAD_FOLDER"], f.filename)
        f.save(path)
        try:
            error_reasons = []
            if ext == ".csv":
                added, updated, skipped, errors = _import_csv(path)
            elif ext in (".xls", ".xlsx"):
                added, updated, skipped, errors, error_reasons = _import_excel(path)
            else:
                flash("Podporované formáty: CSV, XLS, XLSX.", "error")
                return redirect(url_for("admin_import"))
            if errors and error_reasons:
                for row_num, reason in error_reasons[:30]:
                    app.logger.warning("Import produktů řádek %s: %s", row_num, reason)
                if len(error_reasons) > 30:
                    app.logger.warning("Import: dalších %s řádků s chybou (viz výše)", len(error_reasons) - 30)
            msg = f"Import dokončen. Přidáno: {added}, aktualizováno: {updated}"
            if skipped:
                msg += f", přeskočeno (prázdný název): {skipped}"
            if errors:
                msg += f", chyba při zpracování: {errors} řádků"
            msg += "."
            flash(msg)
        except Exception as e:
            app.logger.exception("Admin import failed")
            flash("Neplatný formát souboru nebo chyba při čtení. Zkontrolujte, že soubor je CSV, XLS nebo XLSX s očekávanými sloupci (název, sku, …).", "error")
        finally:
            if os.path.exists(path):
                os.remove(path)
        return redirect(url_for("admin_import"))
    return render_template("admin_import.html", user=get_current_user())


@app.route("/admin/import-katalog-ean", methods=["GET", "POST"])
@login_required
@role_required("admin")
def admin_import_katalog_ean():
    """Import EAN z katalog.xlsx: přiřazení EAN k produktům podle kod_zbozi nebo nazev."""
    if request.method == "POST":
        f = request.files.get("file")
        if not f or not f.filename:
            flash("Vyberte soubor katalog.xlsx.", "error")
            return redirect(url_for("admin_import_katalog_ean"))
        ext = os.path.splitext(f.filename)[1].lower()
        if ext != ".xlsx":
            flash("Očekáván formát .xlsx.", "error")
            return redirect(url_for("admin_import_katalog_ean"))
        path = os.path.join(app.config["UPLOAD_FOLDER"], f.filename)
        f.save(path)
        try:
            if not _validate_katalog_xls_headers(path):
                flash("Neplatná struktura katalogu – očekávány sloupce nazev nebo kod_zbozi a alespoň jeden sloupec EAN.", "error")
                if os.path.exists(path):
                    os.remove(path)
                return redirect(url_for("admin_import_katalog_ean"))
            rows = _parse_katalog_xlsx(path)
        except Exception as e:
            app.logger.exception("Katalog EAN import parse failed")
            flash(f"Chyba čtení souboru: {e}", "error")
            if os.path.exists(path):
                os.remove(path)
            return redirect(url_for("admin_import_katalog_ean"))
        if os.path.exists(path):
            os.remove(path)
        imported = 0
        updated_ids = set()
        new_eans = 0
        skipped_rows = 0
        duplicate_eans = 0
        for row in rows:
            product = None
            if row.get("kod_zbozi"):
                product = Product.query.filter(Product.kod_zbozi == row["kod_zbozi"]).first()
            if not product and row.get("nazev"):
                product = Product.query.filter(Product.nazev == row["nazev"]).first()
            if not product:
                # Vytvoř nový produkt (plný import katalogu)
                nazev = (row.get("nazev") or "").strip() or None
                kod_zbozi = _norm(row.get("kod_zbozi"))
                if not nazev and not kod_zbozi:
                    skipped_rows += 1
                    app.logger.info(
                        "Katalog EAN: přeskočen řádek (bez názvu i kódu), kod_zbozi=%s nazev=%s",
                        row.get("kod_zbozi"),
                        row.get("nazev"),
                    )
                    continue
                product = Product(
                    nazev=nazev or "[Bez názvu]",
                    kod_zbozi=kod_zbozi,
                    naz_skup=row.get("naz_skup"),
                    mj=row.get("mj"),
                    nc=row.get("nc"),
                    pc=row.get("pc"),
                )
                db.session.add(product)
                db.session.flush()
                imported += 1
            else:
                # Aktualizuj existující produkt (pouze doplnění prázdných polí)
                _update_product_only_missing(
                    product,
                    row.get("nazev"),
                    row.get("kod_zbozi"),
                    row.get("mj"),
                    row.get("naz_skup"),
                    row.get("pc"),
                    None,
                    row.get("nc"),
                )
                updated_ids.add(product.id)
            for ean in row.get("eans") or []:
                ean = (ean or "").strip()
                if not ean:
                    continue
                existing_pe = ProductEan.query.filter(ProductEan.ean == ean).first()
                existing_p = Product.query.filter(Product.ean == ean).first()
                if existing_pe and existing_pe.product_id != product.id:
                    duplicate_eans += 1
                    app.logger.info("Katalog EAN: duplicitní EAN %s (již přiřazen jinému produktu)", ean)
                    continue
                if existing_p and existing_p.id != product.id:
                    duplicate_eans += 1
                    app.logger.info("Katalog EAN: duplicitní EAN %s (již u jiného produktu)", ean)
                    continue
                if existing_pe and existing_pe.product_id == product.id:
                    continue
                if existing_p and existing_p.id == product.id:
                    continue
                if product.ean is None or product.ean == "":
                    product.ean = ean
                    new_eans += 1
                else:
                    pe = ProductEan(product_id=product.id, ean=ean)
                    db.session.add(pe)
                    new_eans += 1
        db.session.commit()
        msg = f"Import katalogu dokončen: nové produkty {imported}, aktualizované {len(updated_ids)}, nové EAN {new_eans}"
        if skipped_rows:
            msg += f", přeskočené řádky {skipped_rows}"
        if duplicate_eans:
            msg += f", duplicitní EAN (u jiného produktu) {duplicate_eans}"
        msg += "."
        flash(msg)
        return redirect(url_for("admin_import_katalog_ean"))
    return render_template("admin/import_katalog_ean.html", user=get_current_user())


def _import_orders_csv(path):
    """Import objednávek z CSV: branch_code, product_id, quantity. Řádky se stejným branch_code = jedna objednávka."""
    import csv
    from collections import defaultdict
    created = 0
    errors = []
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            raise ValueError("Prázdný soubor nebo chybějící hlavička")
        cols = [c.strip().lower().replace(" ", "_") for c in reader.fieldnames]
        rows_by_branch = defaultdict(list)
        for row in reader:
            row = {cols[i]: (v.strip() if isinstance(v, str) else v) for i, v in enumerate(row.values()) if i < len(cols)}
            branch_code = (row.get("branch_code") or row.get("pobocka") or "").strip()
            pid_raw = row.get("product_id") or row.get("product_id") or ""
            qty = row.get("quantity")
            if not branch_code:
                errors.append("Chybějící branch_code")
                continue
            try:
                qty = float(qty) if qty not in (None, "") else 0
            except (TypeError, ValueError):
                errors.append(f"Neplatné množství: {qty}")
                continue
            if qty <= 0:
                continue
            rows_by_branch[branch_code].append({"product_id": pid_raw, "quantity": qty})
    branch_by_code = {b.code: b for b in Branch.query.all() if b.code}
    placeholder = _get_or_create_custom_placeholder_product()
    user = get_current_user()
    for branch_code, rows in rows_by_branch.items():
        branch = branch_by_code.get(branch_code)
        if not branch:
            errors.append(f"Pobočka s kódem '{branch_code}' neexistuje")
            continue
        order = Order(branch_id=branch.id, status="pending", created_by_id=user.id if user else None, order_type="normal")
        db.session.add(order)
        db.session.flush()
        for r in rows:
            pid_raw = (r["product_id"] or "").strip()
            if pid_raw.isdigit():
                product = Product.query.get(int(pid_raw))
                if not product:
                    errors.append(f"Produkt id {pid_raw} neexistuje")
                    continue
                item = OrderItem(order_id=order.id, product_id=product.id, ordered_quantity=r["quantity"])
            else:
                item = OrderItem(order_id=order.id, product_id=placeholder.id, ordered_quantity=r["quantity"], custom_product_name=pid_raw or None)
            db.session.add(item)
        db.session.commit()
        audit_log("order_created", "order", order.id, f"Import objednávky pro pobočku {branch_code}")
        created += 1
    return created, errors


def _import_vydejky_csv(path):
    """Import výdejek z CSV: order_id, order_item_id, shipped_quantity. Aktualizuje odeslané množství.
    Položky objednávky, které v importu nejsou, se označí jako sklad nemá (unavailable=True)."""
    import csv
    from collections import defaultdict
    updated = 0
    errors = []
    # order_id -> set of order_item_id, které jsou v importu
    order_item_ids_in_import = defaultdict(set)
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            raise ValueError("Prázdný soubor nebo chybějící hlavička")
        cols = [c.strip().lower().replace(" ", "_") for c in reader.fieldnames]
        for row in reader:
            row = {cols[i]: (v.strip() if isinstance(v, str) else v) for i, v in enumerate(row.values()) if i < len(cols)}
            oid = row.get("order_id")
            iid = row.get("order_item_id") or row.get("item_id")
            shipped = row.get("shipped_quantity")
            if not oid or not iid:
                errors.append("Chybějící order_id nebo order_item_id")
                continue
            try:
                oid, iid = int(oid), int(iid)
                shipped = float(shipped) if shipped not in (None, "") else 0
            except (TypeError, ValueError):
                errors.append(f"Neplatné číslo: order_id={oid}, item_id={iid}, shipped={shipped}")
                continue
            if shipped < 0:
                continue
            item = OrderItem.query.filter_by(id=iid, order_id=oid).first()
            if not item:
                errors.append(f"Položka order_item_id={iid} v objednávce {oid} neexistuje")
                continue
            item.shipped_quantity = shipped
            order_item_ids_in_import[oid].add(iid)
            order = Order.query.get(oid)
            if order:
                _update_order_status_from_items(order)
            updated += 1
    # Položky v objednávce, které v importu nebyly → sklad nemá
    for oid, item_ids_in_import in order_item_ids_in_import.items():
        order = Order.query.get(oid)
        if not order:
            continue
        for item in order.items:
            if item.id not in item_ids_in_import:
                item.unavailable = True
    db.session.commit()
    return updated, errors


@app.route("/admin/import-orders", methods=["POST"])
@login_required
@role_required("admin")
def admin_import_orders():
    f = request.files.get("file")
    if not f or not f.filename:
        flash("Vyberte soubor.", "error")
        return redirect(url_for("admin_import"))
    ext = os.path.splitext(f.filename)[1].lower()
    if ext != ".csv":
        flash("Import objednávek podporuje pouze CSV.", "error")
        return redirect(url_for("admin_import"))
    path = os.path.join(app.config["UPLOAD_FOLDER"], f.filename)
    f.save(path)
    try:
        created, errors = _import_orders_csv(path)
        msg = f"Import objednávek: vytvořeno {created} objednávek."
        if errors:
            msg += " Chyby: " + "; ".join(errors[:10])
            if len(errors) > 10:
                msg += f" (+{len(errors) - 10} dalších)"
        flash(msg)
    except Exception as e:
        flash(f"Chyba importu objednávek: {e}", "error")
    finally:
        if os.path.exists(path):
            os.remove(path)
    return redirect(url_for("admin_import"))


@app.route("/admin/import-vydejky", methods=["POST"])
@login_required
@role_required("admin")
def admin_import_vydejky():
    f = request.files.get("file")
    if not f or not f.filename:
        flash("Vyberte soubor.", "error")
        return redirect(url_for("admin_import"))
    ext = os.path.splitext(f.filename)[1].lower()
    if ext != ".csv":
        flash("Import výdejek podporuje pouze CSV.", "error")
        return redirect(url_for("admin_import"))
    path = os.path.join(app.config["UPLOAD_FOLDER"], f.filename)
    f.save(path)
    try:
        updated, errors = _import_vydejky_csv(path)
        msg = f"Import výdejek: aktualizováno {updated} položek."
        if errors:
            msg += " Chyby: " + "; ".join(errors[:10])
            if len(errors) > 10:
                msg += f" (+{len(errors) - 10} dalších)"
        flash(msg)
    except Exception as e:
        flash(f"Chyba importu výdejek: {e}", "error")
    finally:
        if os.path.exists(path):
            os.remove(path)
    return redirect(url_for("admin_import"))


@app.route("/admin/import-internal", methods=["POST"])
@login_required
@role_required("admin")
def admin_import_internal():
    """Import interních produktů do tabulky internal_products (CSV, stejný formát jako běžné produkty)."""
    f = request.files.get("file")
    if not f or not f.filename:
        flash("Vyberte soubor.", "error")
        return redirect(url_for("admin_import"))
    ext = os.path.splitext(f.filename)[1].lower()
    if ext != ".csv":
        flash("Import interních produktů podporuje pouze CSV.", "error")
        return redirect(url_for("admin_import"))
    path = os.path.join(app.config["UPLOAD_FOLDER"], f.filename)
    f.save(path)
    try:
        added, updated, skipped, errors = _import_internal_csv(path)
        msg = f"Import interních produktů: přidáno {added}, aktualizováno {updated}"
        if skipped:
            msg += f", přeskočeno: {skipped}"
        if errors:
            msg += f", chyba: {errors} řádků"
        msg += "."
        flash(msg)
    except Exception as e:
        flash(f"Chyba importu interních produktů: {e}", "error")
    finally:
        if os.path.exists(path):
            os.remove(path)
    return redirect(url_for("admin_import"))


@app.route("/warehouse")
@login_required
@role_required("warehouse")
def warehouse_dashboard():
    from sqlalchemy import func
    total_orders = Order.query.count()
    by_status = (
        db.session.query(Order.status, func.count(Order.id))
        .group_by(Order.status)
        .all()
    )
    status_counts = {s: c for s, c in by_status}
    top_products = (
        db.session.query(Product.id, Product.nazev, Product.kod_zbozi, Product.mj, func.sum(OrderItem.ordered_quantity).label("total"))
        .join(OrderItem, OrderItem.product_id == Product.id)
        .group_by(Product.id, Product.nazev, Product.kod_zbozi, Product.mj)
        .order_by(func.sum(OrderItem.ordered_quantity).desc())
        .limit(12)
        .all()
    )
    recent_orders = Order.query.order_by(Order.created_at.desc()).limit(4).all()
    return render_template(
        "warehouse_dashboard.html",
        user=get_current_user(),
        total_orders=total_orders,
        status_counts=status_counts,
        top_products=top_products,
        recent_orders=recent_orders,
        status_cz=status_cz,
    )


@app.route("/warehouse/goods-receipts")
@login_required
@role_required("warehouse", "admin")
def warehouse_goods_receipts():
    """Seznam příjmů zboží od dodavatele (aktuální sklad)."""
    wid = get_current_warehouse_id()
    q = GoodsReceipt.query
    if wid:
        q = q.filter(GoodsReceipt.warehouse_id == wid)
    receipts = q.order_by(GoodsReceipt.received_at.desc()).all()
    return render_template(
        "warehouse_goods_receipts.html",
        receipts=receipts,
        user=get_current_user(),
    )


@app.route("/warehouse/branches")
@login_required
@role_required("warehouse", "admin")
def warehouse_branches():
    """Seznam odběratelů (poboček) – pouze pro čtení pro roli sklad."""
    branches = Branch.query.order_by(Branch.name).all()
    return render_template(
        "warehouse_branches.html",
        branches=branches,
        user=get_current_user(),
    )


@app.route("/warehouse/goods-receipts/<int:receipt_id>")
@login_required
@role_required("warehouse", "admin")
def warehouse_goods_receipt_detail(receipt_id):
    """Detail příjmu zboží pro sklad – stejné údaje jako admin (ID, dodavatel, datum, položky, nc, pc)."""
    receipt = GoodsReceipt.query.get_or_404(receipt_id)
    wid = get_current_warehouse_id()
    user = get_current_user()
    if wid and receipt.warehouse_id != wid and (not user or user.role != "admin"):
        flash("Tento příjem nepatří do vašeho skladu.", "error")
        return redirect(url_for("warehouse_goods_receipts"))
    total_nc = sum((it.product.nc or 0) * (it.quantity or 0) for it in receipt.items if it.product)
    total_pc = sum((it.product.pc or 0) * (it.quantity or 0) for it in receipt.items if it.product)
    return render_template(
        "warehouse_goods_receipt_detail.html",
        receipt=receipt,
        total_nc=total_nc,
        total_pc=total_pc,
        user=user,
    )


@app.route("/warehouse/goods-receipts/new", methods=["GET", "POST"])
@login_required
@role_required("warehouse", "admin")
def warehouse_goods_receipt_new():
    """Vytvoření nového příjmu zboží, volitelný import XLSX."""
    wid = get_current_warehouse_id()
    if not wid:
        flash("Nemáte přiřazen sklad.", "error")
        return redirect(url_for("warehouse_dashboard"))
    if request.method == "POST":
        supplier_id = None
        supplier_code = (request.form.get("supplier_code") or "").strip()
        if supplier_code:
            sup = Supplier.query.filter(Supplier.supplier_id == supplier_code).first()
            if sup:
                supplier_id = sup.id
        receipt = GoodsReceipt(
            warehouse_id=wid,
            supplier_id=supplier_id,
            created_by_id=get_current_user().id if get_current_user() else None,
            note=None,
        )
        db.session.add(receipt)
        db.session.flush()
        added = 0
        skipped = 0
        path = None
        f = request.files.get("file")
        if f and f.filename:
            ext = os.path.splitext(f.filename)[1].lower()
            if ext in (".xls", ".xlsx"):
                path = os.path.join(app.config["UPLOAD_FOLDER"], f.filename)
                f.save(path)
                try:
                    if _validate_warehouse_xls_headers(path):
                        rows = _parse_warehouse_xls(path)
                        product_quantities = {}
                        for row in rows:
                            product = find_product_by_ean_or_code(row.get("ean"), row.get("kod_zbozi"))
                            if not product:
                                skipped += 1
                                app.logger.info(
                                    "Goods receipt import: produkt nenalezen pro ean=%s kod_zbozi=%s",
                                    row.get("ean"), row.get("kod_zbozi")
                                )
                                continue
                            qty = row.get("quantity") or 1.0
                            product_quantities[product.id] = product_quantities.get(product.id, 0) + qty
                        for product_id, qty in product_quantities.items():
                            db.session.add(GoodsReceiptItem(goods_receipt_id=receipt.id, product_id=product_id, quantity=qty))
                            added += 1
                    else:
                        flash("Neplatná struktura XLS – očekávány sloupce EAN nebo kod_zbozi.", "error")
                except Exception as e:
                    db.session.rollback()
                    app.logger.exception("Goods receipt import failed")
                    flash(f"Chyba importu příjmu: {e}", "error")
                    if path and os.path.exists(path):
                        os.remove(path)
                    return redirect(url_for("warehouse_goods_receipts"))
                finally:
                    if path and os.path.exists(path):
                        os.remove(path)
        db.session.commit()
        audit_log("goods_receipt_created", "goods_receipt", receipt.id, f"Příjem #{receipt.id}, položek {added}, přeskočeno {skipped}")
        flash(f"Příjem zboží #{receipt.id} vytvořen." + (f" Import: {added} položek, přeskočeno (produkt nenalezen v katalogu): {skipped}." if added or skipped else ""))
        return redirect(url_for("warehouse_goods_receipts"))
    suppliers = Supplier.query.order_by(Supplier.supplier_name).all()
    return render_template("warehouse_goods_receipt_new.html", user=get_current_user(), suppliers=suppliers)


@app.route("/warehouse/suppliers/by-code")
@login_required
@role_required("warehouse", "admin")
def warehouse_suppliers_by_code():
    """Vrací JSON s názvem dodavatele pro daný kód (pro předvyplnění při vytváření příjmu)."""
    code = (request.args.get("code") or "").strip()
    if not code:
        return jsonify({"supplier_name": None, "supplier_id": None})
    sup = Supplier.query.filter(Supplier.supplier_id == code).first()
    if not sup:
        return jsonify({"supplier_name": None, "supplier_id": None})
    return jsonify({"supplier_name": sup.supplier_name, "supplier_id": sup.supplier_id})


@app.route("/warehouse/suppliers", methods=["GET", "POST"])
@login_required
@role_required("warehouse", "admin")
def warehouse_suppliers():
    """Seznam dodavatelů a vytvoření nového."""
    if request.method == "POST":
        sid = (request.form.get("supplier_id") or "").strip()
        name = (request.form.get("supplier_name") or "").strip()
        if not sid or not name:
            flash("Vyplňte kód i název dodavatele.", "error")
        elif Supplier.query.filter(Supplier.supplier_id == sid).first():
            flash(f"Dodavatel s kódem „{sid}“ již existuje.", "error")
        else:
            s = Supplier(supplier_id=sid, supplier_name=name)
            db.session.add(s)
            db.session.commit()
            flash(f"Dodavatel „{name}“ byl přidán.", "success")
            return redirect(url_for("warehouse_suppliers"))
    suppliers = Supplier.query.order_by(Supplier.supplier_name).all()
    warehouses = {w.id: w for w in Warehouse.query.all()}
    edit_id = request.args.get("edit", type=int)
    edit_supplier = Supplier.query.get(edit_id) if edit_id else None
    return render_template(
        "warehouse_suppliers.html",
        suppliers=suppliers,
        warehouses=warehouses,
        edit_supplier=edit_supplier,
        user=get_current_user(),
    )


@app.route("/warehouse/suppliers/<int:supplier_id>/edit", methods=["POST"])
@login_required
@role_required("admin")
def warehouse_supplier_edit(supplier_id):
    """Admin: úprava dodavatele (kód, název, ulice, obec)."""
    s = Supplier.query.get_or_404(supplier_id)
    new_code = (request.form.get("supplier_id") or "").strip()
    new_name = (request.form.get("supplier_name") or "").strip()
    if not new_code or not new_name:
        flash("Vyplňte kód i název dodavatele.", "error")
        return redirect(url_for("warehouse_suppliers"))
    if new_code != s.supplier_id and Supplier.query.filter(Supplier.supplier_id == new_code).first():
        flash(f"Dodavatel s kódem „{new_code}“ již existuje.", "error")
        return redirect(url_for("warehouse_suppliers"))
    s.supplier_id = new_code
    s.supplier_name = new_name
    s.street = (request.form.get("street") or "").strip() or None
    s.municipality = (request.form.get("municipality") or "").strip() or None
    db.session.commit()
    flash("Dodavatel byl upraven.")
    return redirect(url_for("warehouse_suppliers"))


@app.route("/warehouse/suppliers/<int:supplier_id>/delete", methods=["POST"])
@login_required
@role_required("admin")
def warehouse_supplier_delete(supplier_id):
    """Admin: smazání dodavatele. U objednávek a příjmů zboží se supplier_id nastaví na NULL."""
    s = Supplier.query.get_or_404(supplier_id)
    Order.query.filter(Order.supplier_id == s.id).update({Order.supplier_id: None})
    GoodsReceipt.query.filter(GoodsReceipt.supplier_id == s.id).update({GoodsReceipt.supplier_id: None})
    db.session.delete(s)
    db.session.commit()
    flash("Dodavatel byl smazán.")
    return redirect(url_for("warehouse_suppliers"))


def _normalize_supplier_import_header(h):
    """Normalizuje název sloupce pro import dodavatelů: bez diakritiky, lowercase, mezery podtržítko."""
    if not h or not isinstance(h, str):
        return ""
    s = (h.strip().lower()
         .replace("á", "a").replace("č", "c").replace("ď", "d").replace("é", "e").replace("ě", "e")
         .replace("í", "i").replace("ň", "n").replace("ó", "o").replace("ř", "r").replace("š", "s")
         .replace("ť", "t").replace("ú", "u").replace("ů", "u").replace("ý", "y").replace("ž", "z"))
    return s.replace(" ", "_").replace("-", "_")


@app.route("/warehouse/suppliers/import", methods=["GET", "POST"])
@login_required
@role_required("admin")
def warehouse_suppliers_import():
    """Admin: import dodavatelů z .xlsx. Hlavička: Číslo firmy | Název firmy | Ulice | Obec."""
    if request.method == "GET":
        return redirect(url_for("warehouse_suppliers"))
    f = request.files.get("file")
    if not f or not f.filename:
        flash("Vyberte soubor .xlsx.", "error")
        return redirect(url_for("warehouse_suppliers"))
    if os.path.splitext(f.filename)[1].lower() != ".xlsx":
        flash("Povolený formát je pouze .xlsx. Hlavička: Číslo firmy | Název firmy | Ulice | Obec.", "error")
        return redirect(url_for("warehouse_suppliers"))
    path = os.path.join(app.config["UPLOAD_FOLDER"], f.filename)
    try:
        f.save(path)
    except Exception as e:
        app.logger.exception("Supplier import: uložení souboru selhalo")
        flash("Soubor se nepodařilo nahrát.", "error")
        return redirect(url_for("warehouse_suppliers"))
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        headers_raw = [c.value for c in ws[1]]
        rows_data = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
        headers = [_normalize_supplier_import_header(str(h or "")) for h in headers_raw]
        col_supplier_id = None
        col_supplier_name = None
        col_street = None
        col_municipality = None
        for i, h in enumerate(headers):
            if h in ("cislo_firmy", "cislo_firma", "supplier_id", "kod", "ico"):
                col_supplier_id = i
            elif h in ("nazev_firmy", "nazev_firma", "supplier_name", "nazev"):
                col_supplier_name = i
            elif h in ("ulice", "street"):
                col_street = i
            elif h in ("obec", "municipality", "mesto", "city"):
                col_municipality = i
        if col_supplier_id is None:
            for i, h in enumerate(headers):
                if "cislo" in h or ("firma" in h and "nazev" not in h):
                    col_supplier_id = i
                    break
        if col_supplier_name is None:
            for i, h in enumerate(headers):
                if "nazev" in h and "firmy" in h:
                    col_supplier_name = i
                    break
        if col_supplier_id is None or col_supplier_name is None:
            flash("V souboru chybí sloupce „Číslo firmy“ a „Název firmy“. Očekávaná hlavička: Číslo firmy | Název firmy | Ulice | Obec.", "error")
            if os.path.exists(path):
                os.remove(path)
            return redirect(url_for("warehouse_suppliers"))
        created = 0
        updated = 0
        def _cell(row, col):
            if col is None or col >= len(row):
                return None
            return row[col]

        for row in rows_data:
            sid = (_cell(row, col_supplier_id) and str(_cell(row, col_supplier_id)).strip()) or ""
            name = (_cell(row, col_supplier_name) and str(_cell(row, col_supplier_name)).strip()) or ""
            if not sid or not name:
                continue
            street = _cell(row, col_street)
            street = (str(street).strip() if street is not None else None) or None
            if street == "None":
                street = None
            municipality = _cell(row, col_municipality)
            municipality = (str(municipality).strip() if municipality is not None else None) or None
            if municipality == "None":
                municipality = None
            existing = Supplier.query.filter(Supplier.supplier_id == sid).first()
            if existing:
                existing.supplier_name = name
                existing.street = street
                existing.municipality = municipality
                updated += 1
            else:
                db.session.add(Supplier(supplier_id=sid, supplier_name=name, street=street, municipality=municipality))
                created += 1
        db.session.commit()
        flash(f"Import dokončen: přidáno {created} dodavatelů, aktualizováno {updated}.", "success")
    except Exception as e:
        db.session.rollback()
        app.logger.exception("Supplier import failed: %s", e)
        flash(f"Chyba importu: {e}", "error")
    finally:
        if path and os.path.exists(path):
            try:
                os.remove(path)
            except OSError:
                pass
    return redirect(url_for("warehouse_suppliers"))


@app.route("/warehouse/orders")
@login_required
@role_required("warehouse")
def warehouse_orders():
    wid = get_current_warehouse_id()
    status_filter = request.args.get("status", "").strip()
    branch_filter = request.args.get("branch_id", type=int)
    page = request.args.get("page", 1, type=int)
    per_page = 50
    if page < 1:
        page = 1
    q = Order.query
    if wid:
        # Sklad vidí: přiřazené svému skladu NEBO nepřiřazené (čekající od pobočky / staré bez warehouse_id)
        q = q.filter(
            db.or_(
                Order.warehouse_id == wid,
                Order.warehouse_id.is_(None),
            )
        )
    else:
        q = q.filter(False)
    if status_filter:
        if status_filter == "shipped":
            has_incomplete = db.session.query(OrderItem.id).filter(
                OrderItem.order_id == Order.id,
                db.func.coalesce(OrderItem.shipped_quantity, 0) < OrderItem.ordered_quantity,
            ).exists()
            q = q.filter(
                db.or_(
                    Order.status.in_(["shipped", "verified", "error"]),
                    (Order.status == "partially_shipped") & ~has_incomplete,
                )
            )
        else:
            q = q.filter(Order.status == status_filter)
    if branch_filter:
        q = q.filter(Order.branch_id == branch_filter)
    q = q.order_by(Order.created_at.desc())
    total = q.count()
    total_pages = max(1, (total + per_page - 1) // per_page) if total else 1
    if page > total_pages:
        page = total_pages
    orders = q.offset((page - 1) * per_page).limit(per_page).all()
    branches = Branch.query.order_by(Branch.name).all()
    order_check_status_map = {o.id: (getattr(o, "check_status", None) or "") for o in orders}
    return render_template(
        "warehouse_orders.html",
        orders=orders,
        user=get_current_user(),
        status_filter=status_filter,
        branch_filter=branch_filter,
        branches=branches,
        page=page,
        total_pages=total_pages,
        total=total,
        has_warehouse=bool(wid),
        order_check_status_map=order_check_status_map,
    )


@app.route("/warehouse/orders/new", methods=["GET", "POST"])
@login_required
@role_required("warehouse", "admin")
def warehouse_order_new():
    """Sklad vytvoří objednávku a přiřadí ji pobočce (created_by_warehouse_id)."""
    branches = Branch.query.order_by(Branch.name).all()
    suppliers = Supplier.query.order_by(Supplier.supplier_name).all()
    base_products = Product.query.filter(
        Product.nazev != "[Vlastní – produkt mimo katalog]",
        db.or_(Product.is_internal == False, Product.is_internal.is_(None)),
    ).order_by(Product.naz_skup, Product.nazev).all()
    if request.method == "POST":
        branch_id = request.form.get("branch_id", type=int)
        branch = Branch.query.get(branch_id) if branch_id else None
        if not branch:
            flash("Zvolte pobočku.", "error")
            return render_template("warehouse_order_new.html", branches=branches, suppliers=suppliers, products=base_products, user=get_current_user())
        supplier_id = request.form.get("supplier_id", type=int) or None
        user = get_current_user()
        create_from_file = request.form.get("create_from_file") == "1"
        # „Vytvořit z převodky“: pouze branch + supplier + file, žádný grid → malé tělo požadavku (žádný 413)
        if create_from_file:
            f = request.files.get("file")
            if not f or not f.filename or os.path.splitext(f.filename)[1].lower() not in (".xls", ".xlsx"):
                flash("Pro vytvoření z převodky vyberte soubor .xls nebo .xlsx.", "error")
                return render_template("warehouse_order_new.html", branches=branches, suppliers=suppliers, products=base_products, user=get_current_user())
            order = Order(
                branch_id=branch.id,
                status="pending",
                created_by_id=None,
                order_type="normal",
                created_by_warehouse_id=user.id if user else None,
                supplier_id=supplier_id,
            )
            wid = get_current_warehouse_id()
            if wid:
                order.warehouse_id = wid
            db.session.add(order)
            db.session.flush()
            path = os.path.join(app.config["UPLOAD_FOLDER"], f.filename)
            try:
                f.save(path)
            except Exception as e:
                db.session.rollback()
                app.logger.exception("Warehouse order from XLS: uložení souboru selhalo")
                flash("Soubor se nepodařilo nahrát.", "error")
                return redirect(url_for("warehouse_orders"))
            try:
                if not _validate_warehouse_xls_headers(path):
                    db.session.rollback()
                    flash("Neplatná struktura XLS – očekávány sloupce EAN nebo kód zboží (kod_zbozi).", "error")
                    if os.path.exists(path):
                        os.remove(path)
                    return render_template("warehouse_order_new.html", branches=branches, suppliers=suppliers, products=base_products, user=get_current_user())
                rows = _parse_warehouse_xls(path)
                if not rows:
                    db.session.rollback()
                    flash("V souboru nebyly nalezeny žádné řádky s EAN nebo kódem zboží.", "error")
                    if os.path.exists(path):
                        os.remove(path)
                    return render_template("warehouse_order_new.html", branches=branches, suppliers=suppliers, products=base_products, user=get_current_user())
                product_quantities = {}
                import_skipped = 0
                for row in rows:
                    product = find_product_by_ean_or_code(row.get("ean"), row.get("kod_zbozi"))
                    if not product:
                        import_skipped += 1
                        continue
                    qty = row.get("quantity") or 1.0
                    product_quantities[product.id] = product_quantities.get(product.id, 0) + qty
                for product_id, qty in product_quantities.items():
                    db.session.add(OrderItem(
                        order_id=order.id,
                        product_id=product_id,
                        ordered_quantity=qty,
                        shipped_quantity=qty,
                    ))
                db.session.commit()
                _update_order_status_from_items(order)
                audit_log("order_created", "order", order.id, f"Sklad vytvořil objednávku z převodky pro pobočku {branch.name} (#{branch.id})")
                flash(f"Objednávka #{order.id} pro pobočku {branch.name} vytvořena z převodky. Přidáno {len(product_quantities)} položek, přeskočeno (produkt nenalezen): {import_skipped}.", "success")
                if import_skipped and not product_quantities:
                    flash("Žádný produkt z XLS nebyl nalezen v katalogu. Zkontrolujte EAN a kód zboží.", "error")
            except Exception as e:
                db.session.rollback()
                app.logger.exception("Import XLS on order create from file failed: %s", e)
                flash("Soubor se nepodařilo přečíst nebo má neplatný formát.", "error")
            finally:
                if path and os.path.exists(path):
                    try:
                        os.remove(path)
                    except OSError:
                        pass
            return redirect(url_for("warehouse_order_detail", order_id=order.id))

        # Klasický formulář: grid produktů + volitelný XLS
        order = Order(
            branch_id=branch.id,
            status="pending",
            created_by_id=None,
            order_type="normal",
            created_by_warehouse_id=user.id if user else None,
            supplier_id=supplier_id,
        )
        wid = get_current_warehouse_id()
        if wid:
            order.warehouse_id = wid
        db.session.add(order)
        db.session.flush()
        added = 0
        for p in base_products:
            key = f"qty_{p.id}"
            try:
                qty = request.form.get(key, type=float)
            except (TypeError, ValueError):
                qty = None
            if qty is not None and qty > 0:
                db.session.add(OrderItem(order_id=order.id, product_id=p.id, ordered_quantity=qty))
                added += 1
        has_xls = False
        f_check = request.files.get("file")
        if f_check and f_check.filename and os.path.splitext(f_check.filename)[1].lower() in (".xls", ".xlsx"):
            has_xls = True
        if added == 0 and not has_xls:
            db.session.rollback()
            flash("Přidejte alespoň jednu položku s množstvím nebo nahrajte soubor XLS.", "error")
            return render_template("warehouse_order_new.html", branches=branches, suppliers=suppliers, products=base_products, user=get_current_user())
        db.session.commit()
        audit_log("order_created", "order", order.id, f"Sklad vytvořil objednávku pro pobočku {branch.name} (#{branch.id})")
        flash(f"Objednávka #{order.id} pro pobočku {branch.name} vytvořena.")
        # Volitelný import z XLS: přidat položky podle EAN/kod_zbozi, množství z XLS
        f = request.files.get("file")
        if f and f.filename:
            ext = os.path.splitext(f.filename)[1].lower()
            if ext in (".xls", ".xlsx"):
                path = os.path.join(app.config["UPLOAD_FOLDER"], f.filename)
                try:
                    f.save(path)
                except Exception as e:
                    app.logger.exception("Warehouse order: uložení souboru selhalo")
                    flash("Soubor se nepodařilo nahrát. Zkuste menší soubor nebo zkontrolujte oprávnění.", "error")
                    return redirect(url_for("warehouse_order_detail", order_id=order.id))
                try:
                    if not _validate_warehouse_xls_headers(path):
                        flash("Neplatná struktura XLS – očekávány sloupce EAN, kód zboží (kod_zbozi), Code nebo SKU.", "error")
                    else:
                        rows = _parse_warehouse_xls(path)
                        app.logger.info("Warehouse order import: přečteno %s řádků z XLS", len(rows))
                        if not rows:
                            flash("V souboru nebyly nalezeny žádné řádky s EAN nebo kódem zboží.", "error")
                        else:
                            product_quantities = {}
                            import_skipped = 0
                            for row in rows:
                                product = find_product_by_ean_or_code(row.get("ean"), row.get("kod_zbozi"))
                                if not product:
                                    import_skipped += 1
                                    continue
                                qty = row.get("quantity") or 1.0
                                product_quantities[product.id] = product_quantities.get(product.id, 0) + qty
                            for product_id, qty in product_quantities.items():
                                existing = OrderItem.query.filter_by(order_id=order.id, product_id=product_id).first()
                                if existing:
                                    existing.ordered_quantity = (existing.ordered_quantity or 0) + qty
                                    existing.shipped_quantity = (existing.shipped_quantity or 0) + qty
                                else:
                                    db.session.add(OrderItem(
                                        order_id=order.id,
                                        product_id=product_id,
                                        ordered_quantity=qty,
                                        shipped_quantity=qty,
                                    ))
                            db.session.commit()
                            _update_order_status_from_items(order)
                            import_added = len(product_quantities)
                            app.logger.info("Warehouse order import: přidáno %s produktů, přeskočeno %s", import_added, import_skipped)
                            if import_added or import_skipped:
                                audit_log("order_import_xls", "order", order.id, f"Import při vytvoření: položek {import_added}, přeskočeno {import_skipped}")
                                flash(f"Import z XLS: přidáno/aktualizováno {import_added} položek, přeskočeno (produkt nenalezen): {import_skipped}.", "message")
                            if import_added == 0 and import_skipped > 0:
                                flash("Žádný produkt z XLS nebyl nalezen v katalogu. Zkontrolujte EAN a kód zboží v souboru.", "error")
                except Exception as e:
                    db.session.rollback()
                    app.logger.exception("Import XLS on order create failed: %s", e)
                    flash("Soubor se nepodařilo přečíst nebo má neplatný formát. Zkontrolujte, že jde o XLS/XLSX se sloupci EAN nebo kód zboží a množství.", "error")
                finally:
                    if path and os.path.exists(path):
                        try:
                            os.remove(path)
                        except OSError:
                            pass
        return redirect(url_for("warehouse_order_detail", order_id=order.id))
    return render_template("warehouse_order_new.html", branches=branches, suppliers=suppliers, products=base_products, user=get_current_user())


@app.route("/warehouse/order/<int:order_id>")
@login_required
@role_required("warehouse", "admin")
def warehouse_order_detail(order_id):
    order = Order.query.get_or_404(order_id)
    total_qty = sum((i.ordered_quantity or 0) for i in order.items)
    shipped_qty = sum((i.shipped_quantity or 0) for i in order.items)
    order_audit_log = (
        AuditLog.query.filter_by(entity_type="order", entity_id=order.id)
        .order_by(AuditLog.created_at.desc())
        .limit(30)
        .all()
    )
    order_total_selling = _order_total_selling(order)
    order_check_status = getattr(order, "check_status", None) or ""
    check_ok_items, check_error_items = _order_check_results(order) if order_check_status in ("verified", "error") else ([], [])
    item_scanned_qty = _order_item_scanned_qtys(order)
    return render_template(
        "warehouse_order_detail.html",
        order=order,
        user=get_current_user(),
        status_cz=status_cz,
        total_qty=total_qty,
        shipped_qty=shipped_qty,
        order_audit_log=order_audit_log,
        order_total_selling=order_total_selling,
        check_ok_items=check_ok_items,
        check_error_items=check_error_items,
        order_check_status=order_check_status,
        item_scanned_qty=item_scanned_qty,
    )


def _parse_quantity_from_row(d, headers_hint=None):
    """Z řádku (dict) vybere množství ze sloupců mnozstvi, ks, quantity, pocet (normalizované hlavičky). Vrací float >= 0, výchozí 1."""
    for key in ("mnozstvi", "ks", "quantity", "pocet", "množství"):
        val = d.get(key)
        if val is None:
            continue
        try:
            if isinstance(val, (int, float)):
                q = float(val)
            else:
                s = str(val).strip().replace(",", ".")
                q = float(s) if s else 0
            if q >= 0:
                return q
        except (ValueError, TypeError):
            pass
    return 1.0


def _ean_from_row(d):
    """Z řádku (dict s normalizovanými klíči) vybere EAN – sloupec ean nebo první ean1, ean2, …"""
    ean = _norm(d.get("ean"))
    if ean:
        return ean
    for k in sorted(d.keys() or []):
        if k and (k == "ean1" or (k.startswith("ean") and len(k) > 3 and k[3:].replace("_", "").isdigit())):
            v = _norm(d.get(k))
            if v:
                return v
    return None


def _kod_zbozi_from_row(d):
    """Z řádku vybere kód zboží – kod_zbozi, kod, code, sku (normalizované hlavičky)."""
    return (
        _norm(d.get("kod_zbozi"))
        or _norm(d.get("kod"))
        or _norm(d.get("code"))
        or _norm(d.get("sku"))
    )


def _parse_warehouse_xls(path):
    """Parsuje .xls/.xlsx z externího skladu. Podporované sloupce: naz_skupiny/naz_skup, kod_zbozi, nazev, ean, mnozstvi/ks/quantity, pc/pc_bez.
    Hlavičky se normalizují (odstraní diakritika). Řádek musí mít alespoň ean nebo kod_zbozi. Vrací seznam dict včetně klíče quantity."""
    def _naz_skup(d):
        return _norm(d.get("naz_skupiny") or d.get("naz_skup"))

    def _pc_val(d):
        return _norm(d.get("pc") or d.get("pc_bez"))

    rows = []
    ext = os.path.splitext(path)[1].lower()
    try:
        if ext == ".xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True)
            ws = wb.active
            headers_raw = [c.value for c in ws[1]]
            headers = [_normalize_header_name(str(h or "")) for h in headers_raw]
            for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
                vals = [c.value for c in row]
                d = dict(zip(headers, vals))
                ean = _ean_from_row(d)
                kod_zbozi = _kod_zbozi_from_row(d)
                if not ean and not kod_zbozi:
                    app.logger.info("Warehouse XLS řádek %s přeskočen: chybí EAN i kod_zbozi", row_num)
                    continue
                qty = _parse_quantity_from_row(d)
                rows.append({
                    "naz_skup": _naz_skup(d),
                    "kod_zbozi": kod_zbozi,
                    "nazev": _norm(d.get("nazev")),
                    "ean": ean,
                    "pc": _pc_val(d),
                    "quantity": qty,
                })
            wb.close()
        else:
            import xlrd
            wb = xlrd.open_workbook(path)
            sheet = wb.sheet_by_index(0)
            headers = []
            for c in range(sheet.ncols):
                v = sheet.cell_value(0, c)
                headers.append(_normalize_header_name(str(v or "")))
            for r in range(1, sheet.nrows):
                row_num = r + 1
                d = {}
                for c in range(sheet.ncols):
                    if c < len(headers) and headers[c]:
                        d[headers[c]] = sheet.cell_value(r, c)
                ean = _ean_from_row(d)
                kod_zbozi = _kod_zbozi_from_row(d)
                if not ean and not kod_zbozi:
                    app.logger.info("Warehouse XLS řádek %s přeskočen: chybí EAN i kod_zbozi", row_num)
                    continue
                qty = _parse_quantity_from_row(d)
                rows.append({
                    "naz_skup": _naz_skup(d),
                    "kod_zbozi": kod_zbozi,
                    "nazev": _norm(d.get("nazev")),
                    "ean": ean,
                    "pc": _pc_val(d),
                    "quantity": qty,
                })
    except Exception:
        raise
    return rows


def _validate_warehouse_xls_headers(path):
    """Ověří, že soubor má v první řádce alespoň jeden ze sloupců ean, kod_zbozi, kod, code, sku (po normalizaci hlaviček)."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        headers_raw = [c.value for c in ws[1]]
        wb.close()
        headers = [_normalize_header_name(str(h or "")) for h in headers_raw]
    else:
        import xlrd
        wb = xlrd.open_workbook(path)
        sheet = wb.sheet_by_index(0)
        headers = [_normalize_header_name(str(sheet.cell_value(0, c) or "")) for c in range(sheet.ncols)]
    return (
        "ean" in headers
        or "kod_zbozi" in headers
        or "kod" in headers
        or "code" in headers
        or "sku" in headers
        or any(h and h.startswith("ean") for h in headers)
    )


@app.route("/warehouse/order/<int:order_id>/import-xls", methods=["POST"])
@login_required
@role_required("warehouse", "admin")
def warehouse_order_import_xls(order_id):
    """Import převodky: množství z XLS doplní odeslané množství u existujících položek; chybějící produkty se přidají do objednávky."""
    order = Order.query.get_or_404(order_id)
    f = request.files.get("file")
    if not f or not f.filename:
        flash("Vyberte soubor.", "error")
        return redirect(url_for("warehouse_order_detail", order_id=order_id))
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in (".xls", ".xlsx"):
        flash("Povolené formáty: .xls, .xlsx.", "error")
        return redirect(url_for("warehouse_order_detail", order_id=order_id))
    path = os.path.join(app.config["UPLOAD_FOLDER"], f.filename)
    f.save(path)
    try:
        if not _validate_warehouse_xls_headers(path):
            flash("Neplatná struktura souboru – očekávány sloupce EAN nebo kód zboží (kod_zbozi).", "error")
            return redirect(url_for("warehouse_order_detail", order_id=order_id))
        rows = _parse_warehouse_xls(path)
        if not rows:
            flash("V souboru nebyly nalezeny žádné řádky s EAN nebo kódem zboží.", "error")
            return redirect(url_for("warehouse_order_detail", order_id=order_id))
        # Agregace podle produktu (součet množství za stejný produkt v XLS)
        product_quantities = {}
        skipped = 0
        for row in rows:
            product = find_product_by_ean_or_code(row.get("ean"), row.get("kod_zbozi"))
            if not product:
                skipped += 1
                continue
            qty = row.get("quantity") or 1.0
            product_quantities[product.id] = product_quantities.get(product.id, 0) + qty
        updated = 0  # počet existujících položek, u kterých bylo doplněno odeslané množství
        added = 0    # počet nových položek přidaných do objednávky
        for product_id, qty in product_quantities.items():
            existing_item = OrderItem.query.filter_by(order_id=order.id, product_id=product_id).first()
            if existing_item:
                existing_item.shipped_quantity = qty
                updated += 1
            else:
                db.session.add(OrderItem(order_id=order.id, product_id=product_id, ordered_quantity=qty, shipped_quantity=qty))
                added += 1
        _update_order_status_from_items(order)
        db.session.commit()
        msg = f"Převodka načtena. Doplněno odeslané množství u {updated} položek, přidáno {added} nových položek z převodky."
        if skipped:
            msg += f" Přeskočeno (produkt nenalezen v katalogu): {skipped}."
        audit_log("order_import_xls", "order", order.id, f"Převodka: doplněno odeslané u {updated} položek, přidáno {added} nových, přeskočeno {skipped}")
        flash(msg)
    except Exception as e:
        db.session.rollback()
        app.logger.exception("Import XLS failed")
        flash(f"Chyba importu: {e}", "error")
    finally:
        if os.path.exists(path):
            os.remove(path)
    return redirect(url_for("warehouse_order_detail", order_id=order_id))


def _order_item_scanned_qtys(order):
    """Vrátí dict { order_item_id: celkové naskenované množství } pro všechny položky objednávky (pro sloupec Dorazilo produktů)."""
    try:
        result = {}
        for item in order.items:
            total = (
                db.session.query(sql_func.coalesce(sql_func.sum(OrderItemCheck.scanned_quantity), 0))
                .filter(OrderItemCheck.order_item_id == item.id)
                .scalar()
                or 0
            )
            result[item.id] = float(total)
        return result
    except Exception:
        app.logger.exception("_order_item_scanned_qtys failed for order %s", order.id)
        return {}


def _order_check_errors(order):
    """Vrátí seznam položek, u kterých kontrola nesedí (skenované množství != očekávané). Pro zobrazení „co bylo špatně“."""
    _, errors = _order_check_results(order)
    return errors


def _order_check_results(order):
    """Vrátí (ok_items, error_items): položky kde sken sedí a kde nesedí. Pro zobrazení výsledku kontroly (co bylo dobře/špatně)."""
    try:
        ok_items = []
        error_items = []
        for item in order.items:
            total_scanned = (
                db.session.query(sql_func.coalesce(sql_func.sum(OrderItemCheck.scanned_quantity), 0))
                .filter(OrderItemCheck.order_item_id == item.id)
                .scalar()
                or 0
            )
            expected = item.shipped_quantity or 0
            scanned_f = float(total_scanned)
            expected_f = float(expected)
            row = {"item": item, "expected": expected_f, "scanned": scanned_f}
            if abs(scanned_f - expected_f) >= 0.001:
                error_items.append(row)
            elif scanned_f > 0:
                ok_items.append(row)
        return (ok_items, error_items)
    except Exception:
        app.logger.exception("_order_check_results failed for order %s", order.id)
        return ([], [])


def _order_check_view(order_id, redirect_back_route, for_branch=False):
    """Společná logika pro stránku kontroly objednávky (používá sklad i pobočka).
    Pro pobočku se zobrazují jen položky s odeslaným množstvím > 0 a které sklad má (ne unavailable)."""
    order = Order.query.get_or_404(order_id)
    if order.status not in ("shipped", "partially_shipped", "verified", "error"):
        flash("Kontrola je k dispozici pouze u odeslaných nebo částečně odeslaných objednávek.", "error")
        return redirect(url_for(redirect_back_route, order_id=order_id))
    if for_branch:
        check_items = [i for i in order.items if (i.shipped_quantity or 0) > 0 and not i.unavailable]
    else:
        check_items = None
    item_results = {}
    items_to_show = check_items if check_items is not None else order.items
    for item in items_to_show:
        last_check = OrderItemCheck.query.filter_by(order_item_id=item.id).order_by(OrderItemCheck.created_at.desc()).first()
        if last_check:
            item_results[item.id] = last_check
    return render_template(
        "warehouse_order_check.html",
        order=order,
        item_results=item_results,
        check_items=check_items,
        user=get_current_user(),
        status_cz=status_cz,
        check_for_branch=for_branch,
    )


@app.route("/warehouse/order/<int:order_id>/check")
@login_required
@role_required("warehouse", "admin")
def warehouse_order_check(order_id):
    """Stránka kontroly objednávky čtečkou (sklad)."""
    return _order_check_view(order_id, "warehouse_order_detail", for_branch=False)


@app.route("/branch/order/<int:order_id>/check")
@login_required
@role_required("branch")
def branch_order_check(order_id):
    """Stránka kontroly objednávky čtečkou (pobočka)."""
    order = Order.query.get_or_404(order_id)
    if order.branch_id != get_current_branch_id():
        flash("Objednávka nepatří vaší pobočce.", "error")
        return redirect(url_for("branch_orders"))
    return _order_check_view(order_id, "branch_order_detail", for_branch=True)


@app.route("/warehouse/order/<int:order_id>/check/scan", methods=["POST"])
@login_required
@role_required("warehouse", "admin", "branch")
def warehouse_order_check_scan(order_id):
    """Zpracuje jeden sken ve formátu množství*EAN (např. 3*8591234567890). Vrací JSON. Přístup: sklad, admin, pobočka (jen k objednávkám své pobočky)."""
    order = Order.query.get_or_404(order_id)
    user = get_current_user()
    if user and user.role == "branch" and get_current_branch_id() != order.branch_id:
        return jsonify({"ok": False, "error": "Objednávka nepatří vaší pobočce."}), 403
    if order.status not in ("shipped", "partially_shipped", "verified", "error"):
        return jsonify({"ok": False, "error": "Kontrola není k dispozici."}), 400
    raw = (request.form.get("scan") or request.json.get("scan") or request.get_data(as_text=True) or "").strip()
    if "*" in raw:
        parts = raw.split("*", 1)
        try:
            scanned_qty = float(parts[0].strip())
        except (ValueError, IndexError):
            return jsonify({"ok": False, "error": "Neplatné množství"}), 400
        ean = (parts[1].strip() if len(parts) > 1 else "").strip()
    else:
        # Bez * = celý vstup je EAN, množství 1 kus
        ean = raw
        scanned_qty = 1.0
    if not ean:
        return jsonify({"ok": False, "error": "Chybí EAN"}), 400
    product = find_product_by_ean(ean)
    if not product:
        return jsonify({"ok": False, "error": f"Produkt s EAN {ean} nenalezen"}), 404
    order_item = OrderItem.query.filter(OrderItem.order_id == order.id, OrderItem.product_id == product.id).first()
    if not order_item:
        return jsonify({"ok": False, "error": "Tento produkt není v objednávce"}), 404
    if user and user.role == "branch":
        if order_item.unavailable or (order_item.shipped_quantity or 0) <= 0:
            return jsonify({"ok": False, "error": "Tato položka se v kontrole neověřuje (nebyla odeslána nebo sklad nemá)."}), 400
    expected = order_item.shipped_quantity if order_item.shipped_quantity is not None else 0
    from sqlalchemy import func
    total_already = db.session.query(func.coalesce(func.sum(OrderItemCheck.scanned_quantity), 0)).filter(
        OrderItemCheck.order_item_id == order_item.id
    ).scalar() or 0
    total_after = total_already + (scanned_qty or 0)
    result = "correct" if abs(total_after - expected) < 0.001 else "incorrect"
    check = OrderItemCheck(
        order_item_id=order_item.id,
        scanned_quantity=scanned_qty,
        expected_quantity=expected,
        result=result,
    )
    db.session.add(check)
    db.session.commit()
    audit_log("order_check_scan", "order_item", order_item.id, f"Scan: {scanned_qty}*{ean} -> {result} (celkem {total_after})")
    # Pro pobočku se výsledek kontroly vyhodnocuje jen z položek s odesl. množstvím > 0 a které sklad má
    if user and user.role == "branch":
        all_items_with_product = [i for i in order.items if i.product_id and i.product and (i.product.ean or i.product.extra_eans.count() > 0) and (i.shipped_quantity or 0) > 0 and not i.unavailable]
    else:
        all_items_with_product = [i for i in order.items if i.product_id and i.product and (i.product.ean or i.product.extra_eans.count() > 0)]
    item_totals = {}
    for it in order.items:
        s = db.session.query(func.coalesce(func.sum(OrderItemCheck.scanned_quantity), 0)).filter(
            OrderItemCheck.order_item_id == it.id
        ).scalar() or 0
        item_totals[it.id] = s
    all_checked = all(item_totals.get(it.id, 0) > 0 for it in all_items_with_product)
    all_correct = all(
        abs((item_totals.get(it.id) or 0) - (it.shipped_quantity or 0)) < 0.001
        for it in all_items_with_product
    )
    if all_checked:
        if all_correct:
            order.check_status = "verified"
            db.session.commit()
            audit_log("order_verified", "order", order.id, "Všechny položky zkontrolovány správně")
        else:
            order.check_status = "error"
            db.session.commit()
            audit_log("order_check_error", "order", order.id, "Kontrola odhalila chybu")
    return jsonify({
        "ok": True,
        "order_item_id": order_item.id,
        "scanned_quantity": scanned_qty,
        "expected_quantity": expected,
        "result": result,
        "order_status": order.status,
        "order_check_status": getattr(order, "check_status", None),
        "all_checked": all_checked,
        "all_correct": all_correct,
        "already_scanned": total_already > 0,
        "total_scanned_for_item": total_after,
    })


@app.route("/warehouse/order/<int:order_id>/status", methods=["POST"])
@login_required
@role_required("warehouse", "admin")
def warehouse_order_status(order_id):
    """Změna stavu objednávky. verified/error nelze nastavit ručně – vyhodnotí se automaticky z kontroly pobočky."""
    order = Order.query.get_or_404(order_id)
    status = request.form.get("status")
    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
    if status in ("verified", "error"):
        if is_ajax:
            return jsonify({"ok": False, "error": "Stav „Zkontrolováno“ a „Chyba kontroly“ se nastavuje automaticky podle kontroly pobočky."}), 400
        flash("Stav „Zkontrolováno“ a „Chyba kontroly“ se nastavuje automaticky podle kontroly pobočky.", "error")
        return redirect(url_for("warehouse_order_detail", order_id=order_id))
    if status in ("pending", "partially_shipped", "shipped"):
        order.status = status
        db.session.commit()
        audit_log("order_status", "order", order.id, f"Objednávka #{order_id} → {status}")
        if is_ajax:
            return jsonify({"ok": True, "order_status": order.status})
        flash("Stav objednávky změněn.")
    elif is_ajax:
        return jsonify({"ok": False}), 400
    return redirect(url_for("warehouse_order_detail", order_id=order_id))


def _update_order_status_from_items(order):
    """Nastaví stav objednávky podle odeslaných množství: všechny položky odeslány -> shipped, jinak částečně -> partially_shipped."""
    if not order.items:
        return
    all_shipped = all(
        (it.shipped_quantity or 0) >= (it.ordered_quantity or 0)
        for it in order.items
    )
    any_shipped = any((it.shipped_quantity or 0) > 0 for it in order.items)
    if all_shipped:
        order.status = "shipped"
    elif any_shipped:
        order.status = "partially_shipped"


def _order_totals(order):
    """Vrátí (total_ordered, total_shipped) pro objednávku."""
    total_o = sum((it.ordered_quantity or 0) for it in order.items)
    total_s = sum((it.shipped_quantity or 0) for it in order.items)
    return total_o, total_s


def _order_total_selling(order):
    """Celková prodejní cena objednávky: SUM(pc * shipped_quantity) přes položky s produktem (jen odeslané množství)."""
    total = 0
    for it in order.items:
        if it.product and it.product.pc is not None:
            total += (it.product.pc or 0) * (it.shipped_quantity or 0)
    return total


@app.route("/warehouse/order/<int:order_id>/item/<int:item_id>/shipped", methods=["POST"])
@login_required
@role_required("warehouse", "admin")
def warehouse_item_shipped(order_id, item_id):
    item = OrderItem.query.filter_by(id=item_id, order_id=order_id).first_or_404()
    shipped = request.form.get("shipped_quantity", type=float)
    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
    if shipped is not None and shipped >= 0:
        item.shipped_quantity = shipped
        order = Order.query.get(order_id)
        _update_order_status_from_items(order)
        db.session.commit()
        audit_log("order_item_shipped", "order_item", item.id, f"Objednávka #{order_id}, odesláno {shipped}")
        audit_log("order_item_shipped", "order", order_id, f"Položka {item.display_name()}: odesláno {shipped}")
        if is_ajax:
            total_o, total_s = _order_totals(order)
            return jsonify({
                "ok": True,
                "shipped_quantity": item.shipped_quantity,
                "order_status": order.status,
                "order_total_ordered": total_o,
                "order_total_shipped": total_s,
            })
        flash("Dodané množství upraveno.")
    elif is_ajax:
        return jsonify({"ok": False, "error": "Neplatné množství."}), 400
    return redirect(url_for("warehouse_order_detail", order_id=order_id))


@app.route("/warehouse/order/<int:order_id>/item/<int:item_id>/unavailable", methods=["POST"])
@login_required
@role_required("warehouse", "admin")
def warehouse_item_unavailable(order_id, item_id):
    item = OrderItem.query.filter_by(id=item_id, order_id=order_id).first_or_404()
    item.unavailable = True
    item.shipped_quantity = 0
    db.session.commit()
    audit_log("order_item_unavailable", "order_item", item.id, f"Objednávka #{order_id}, položka {item.display_name()}")
    audit_log("order_item_unavailable", "order", order_id, f"Položka {item.display_name()} – Nemám skladem")
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        total_o, total_s = _order_totals(Order.query.get(order_id))
        return jsonify({"ok": True, "unavailable": True, "order_total_shipped": total_s, "order_total_ordered": total_o})
    flash("Položka označena jako „Nemám na skladě“.")
    return redirect(url_for("warehouse_order_detail", order_id=order_id))


@app.route("/warehouse/order/<int:order_id>/item/<int:item_id>/available", methods=["POST"])
@login_required
@role_required("warehouse", "admin")
def warehouse_item_available(order_id, item_id):
    item = OrderItem.query.filter_by(id=item_id, order_id=order_id).first_or_404()
    item.unavailable = False
    db.session.commit()
    audit_log("order_item_available", "order_item", item.id, f"Objednávka #{order_id}, položka {item.display_name()}")
    audit_log("order_item_available", "order", order_id, f"Položka {item.display_name()} – Má skladem")
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return jsonify({"ok": True, "unavailable": False})
    flash("Položka znovu označena jako „Má skladem“.")
    return redirect(url_for("warehouse_order_detail", order_id=order_id))


@app.route("/warehouse/order/<int:order_id>/create-order-from-unavailable", methods=["POST"])
@login_required
@role_required("admin")
def warehouse_order_create_from_unavailable(order_id):
    """Vytvoří novou objednávku pro stejnou pobočku z položek označených jako „Nemám“."""
    order = Order.query.get_or_404(order_id)
    unavailable_items = [i for i in order.items if i.unavailable and i.product_id]
    if not unavailable_items:
        flash("Žádné položky označené jako „Nemám“. Označte nejprve položky tlačítkem Nemám.", "error")
        return redirect(url_for("warehouse_order_detail", order_id=order_id))
    user = get_current_user()
    new_order = Order(
        branch_id=order.branch_id,
        status="pending",
        order_type="normal",
        created_by_warehouse_id=user.id if user else None,
        supplier_id=order.supplier_id,
    )
    # nová objednávka je přiřazena stejnému skladu jako původní, nebo aktuálnímu skladu admina
    if order.warehouse_id:
        new_order.warehouse_id = order.warehouse_id
    else:
        wid_new = get_current_warehouse_id()
        if wid_new:
            new_order.warehouse_id = wid_new
    db.session.add(new_order)
    db.session.flush()
    for it in unavailable_items:
        db.session.add(OrderItem(
            order_id=new_order.id,
            product_id=it.product_id,
            ordered_quantity=it.ordered_quantity or 0,
            shipped_quantity=None,
            branch_note=it.branch_note,
        ))
    db.session.commit()
    audit_log("order_created_from_unavailable", "order", new_order.id, f"Z objednávky #{order_id}: {len(unavailable_items)} položek (Nemám) → nová objednávka #{new_order.id}")
    flash(f"Nová objednávka #{new_order.id} vytvořena z {len(unavailable_items)} nedostupných položek. Stejná pobočka: {order.branch.name if order.branch else '—'}.")
    return redirect(url_for("warehouse_order_detail", order_id=new_order.id))


@app.route("/warehouse/order/<int:order_id>/item/<int:item_id>/note", methods=["POST"])
@login_required
@role_required("warehouse", "admin")
def warehouse_item_note(order_id, item_id):
    item = OrderItem.query.filter_by(id=item_id, order_id=order_id).first_or_404()
    note = request.form.get("warehouse_note", "").strip() or None
    item.warehouse_note = note
    db.session.commit()
    if note:
        audit_log("order_item_note", "order_item", item.id, f"Poznámka skladu k položce (objednávka #{order_id})")
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return jsonify({"ok": True, "warehouse_note": note or ""})
    flash("Poznámka skladu uložena." if note else "Poznámka skladu odebrána.")
    return redirect(url_for("warehouse_order_detail", order_id=order_id))


with app.app_context():
    from sqlalchemy import text
    db.create_all()
    # Migrace: vytvořit user_branches, pokud chybí (staré DB nebo create_all ji neudělal)
    try:
        db.session.execute(text("""
            CREATE TABLE IF NOT EXISTS user_branches (
                user_id INTEGER NOT NULL,
                branch_id INTEGER NOT NULL,
                PRIMARY KEY (user_id, branch_id),
                FOREIGN KEY (user_id) REFERENCES users (id),
                FOREIGN KEY (branch_id) REFERENCES branches (id)
            )
        """))
        db.session.commit()
    except Exception:
        db.session.rollback()
    try:
        db.session.execute(text("""
            CREATE TABLE IF NOT EXISTS audit_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                user_id INTEGER REFERENCES users(id),
                username VARCHAR(80) NOT NULL,
                action VARCHAR(80) NOT NULL,
                entity_type VARCHAR(50),
                entity_id INTEGER,
                details TEXT
            )
        """))
        db.session.commit()
    except Exception:
        db.session.rollback()
    # Doplnit nové sloupce v order_items / products, pokud ještě neexistují (pro existující DB)
    for sql in (
        "ALTER TABLE order_items ADD COLUMN branch_note VARCHAR(255)",
        "ALTER TABLE order_items ADD COLUMN unavailable INTEGER DEFAULT 0",
        "ALTER TABLE order_items ADD COLUMN custom_product_name VARCHAR(255)",
        "ALTER TABLE products ADD COLUMN group_name VARCHAR(255)",
        "ALTER TABLE products ADD COLUMN pc VARCHAR(100)",
        "ALTER TABLE products ADD COLUMN unit VARCHAR(50)",
        "ALTER TABLE products ADD COLUMN is_internal INTEGER DEFAULT 0",
        "ALTER TABLE orders ADD COLUMN created_by_id INTEGER REFERENCES users(id)",
        "ALTER TABLE orders ADD COLUMN order_type VARCHAR(30) DEFAULT 'normal'",
        "ALTER TABLE orders ADD COLUMN created_by_warehouse_id INTEGER REFERENCES users(id)",
    ):
        try:
            db.session.execute(text(sql))
            db.session.commit()
        except Exception:
            db.session.rollback()
    # Tabulka warehouses a orders.warehouse_id
    try:
        db.session.execute(text("""
            CREATE TABLE IF NOT EXISTS warehouses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name VARCHAR(255) NOT NULL
            )
        """))
        db.session.commit()
    except Exception:
        db.session.rollback()
    try:
        db.session.execute(text("ALTER TABLE orders ADD COLUMN warehouse_id INTEGER REFERENCES warehouses(id)"))
        db.session.commit()
    except Exception:
        db.session.rollback()
    # Tabulka order_item_checks (kontrola objednávky čtečkou)
    try:
        db.session.execute(text("""
            CREATE TABLE IF NOT EXISTS order_item_checks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_item_id INTEGER NOT NULL REFERENCES order_items(id),
                scanned_quantity REAL NOT NULL,
                expected_quantity REAL NOT NULL,
                result VARCHAR(20) NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """))
        db.session.commit()
    except Exception:
        db.session.rollback()
    # Nové schéma produktů: naz_skup, kod_zbozi, nazev, ean, mj, nc, pc_float
    for sql in (
        "ALTER TABLE products ADD COLUMN naz_skup VARCHAR(100)",
        "ALTER TABLE products ADD COLUMN kod_zbozi VARCHAR(100)",
        "ALTER TABLE products ADD COLUMN nazev VARCHAR(255)",
        "ALTER TABLE products ADD COLUMN ean VARCHAR(50)",
        "ALTER TABLE products ADD COLUMN mj VARCHAR(20)",
        "ALTER TABLE products ADD COLUMN nc REAL",
        "ALTER TABLE products ADD COLUMN pc_float REAL",
    ):
        try:
            db.session.execute(text(sql))
            db.session.commit()
        except Exception:
            db.session.rollback()
    # Zkopírovat data ze starých sloupců do nových (pokud tabulka má ještě name/sku/...)
    try:
        r = db.session.execute(text("PRAGMA table_info(products)"))
        cols = [row[1] for row in r.fetchall()]
        if "name" in cols and "nazev" in cols:
            db.session.execute(text(
                "UPDATE products SET nazev=name, kod_zbozi=sku, naz_skup=group_name, mj=unit WHERE id IS NOT NULL"
            ))
            db.session.execute(text(
                "UPDATE products SET pc_float=CAST(pc AS REAL) WHERE pc IS NOT NULL AND TRIM(COALESCE(pc,'')) != ''"
            ))
            db.session.commit()
    except Exception:
        db.session.rollback()
    # Tabulka internal_products (samostatná od products) – před přidáním FK z order_items
    try:
        db.session.execute(text("""
            CREATE TABLE IF NOT EXISTS internal_products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name VARCHAR(255) NOT NULL,
                sku VARCHAR(100),
                unit VARCHAR(50),
                group_name VARCHAR(255),
                pc VARCHAR(100)
            )
        """))
        db.session.commit()
    except Exception:
        db.session.rollback()
    for sql in (
        "ALTER TABLE order_items ADD COLUMN internal_product_id INTEGER REFERENCES internal_products(id)",
        "ALTER TABLE branch_carts ADD COLUMN cart_internal_json TEXT DEFAULT '{}'",
    ):
        try:
            db.session.execute(text(sql))
            db.session.commit()
        except Exception:
            db.session.rollback()
    # Migrace: naplnit user_branches z existujícího branch_id
    try:
        for u in User.query.filter(User.branch_id.isnot(None)):
            b = Branch.query.get(u.branch_id)
            if b and b not in list(u.branches):
                u.branches.append(b)
        db.session.commit()
    except Exception:
        db.session.rollback()
    # Seed admin jen když neexistuje; při Gunicornu (více workery) může běžet
    # souběžně – ošetříme race condition pomocí try/except.
    if not User.query.filter_by(username="admin").first():
        try:
            u = User(username="admin", role="admin")
            u.set_password("admin")
            db.session.add(u)
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            # Jiný worker už admina vložil – OK
            pass


def _seed_admin():
    """Jednorázově vytvoří admina (pro flask seed-admin nebo ruční volání)."""
    if User.query.filter_by(username="admin").first():
        return False
    u = User(username="admin", role="admin")
    u.set_password("admin")
    db.session.add(u)
    db.session.commit()
    return True


@app.cli.command("seed-admin")
def seed_admin_cmd():
    """Vytvoří výchozího admina (username: admin, heslo: admin). Spusť jednou: flask seed-admin"""
    if _seed_admin():
        print("Admin uživatel vytvořen (admin / admin).")
    else:
        print("Admin uživatel již existuje.")


def _reset_db(seed_admin=True):
    """Smaže všechny tabulky, znovu je vytvoří. Po resetu vždy vytvoří výchozího admina (admin/admin)."""
    db.drop_all()
    db.create_all()
    if seed_admin:
        _seed_admin()
    return True


@app.cli.command("reset-db")
@click.option("--no-seed", is_flag=True, help="Po resetu nevytvářet výchozího admina")
def reset_db_cmd(no_seed):
    """Smaže a znovu vytvoří celou databázi. POZOR: smaže všechna data! Vytvoří admin/admin."""
    with app.app_context():
        _reset_db(seed_admin=not no_seed)
    click.echo("Databáze byla resetována." + ("" if no_seed else " Admin vytvořen (admin/admin)."))


@app.cli.command("check-ean-consistency")
def check_ean_consistency_cmd():
    """Kontrola konzistence EAN: duplicitní EAN v ProductEan a Product.ean. Výstup pouze report, žádné mazání."""
    with app.app_context():
        from sqlalchemy import func
        dup_pe = (
            db.session.query(ProductEan.ean, func.count(ProductEan.id).label("cnt"))
            .group_by(ProductEan.ean)
            .having(func.count(ProductEan.id) > 1)
            .all()
        )
        dup_p = (
            db.session.query(Product.ean, func.count(Product.id).label("cnt"))
            .filter(Product.ean.isnot(None), Product.ean != "")
            .group_by(Product.ean)
            .having(func.count(Product.id) > 1)
            .all()
        )
        if dup_pe or dup_p:
            click.echo("Nalezeny duplicitní EAN:")
            for ean, cnt in dup_pe:
                click.echo("  ProductEan: EAN %s v %s záznamech" % (ean, cnt))
            for ean, cnt in dup_p:
                click.echo("  Product.ean: EAN %s u %s produktů" % (ean, cnt))
        else:
            click.echo("OK: Žádné duplicitní EAN v ProductEan ani v Product.ean.")


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=True)
