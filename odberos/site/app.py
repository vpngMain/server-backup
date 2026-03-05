from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_wtf import FlaskForm
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from wtforms import StringField, BooleanField, DateField, FloatField, TextAreaField, PasswordField, SelectField
from wtforms.validators import DataRequired, Optional, Regexp, Length
from datetime import datetime, date
try:
    from zoneinfo import ZoneInfo
    from zoneinfo import ZoneInfoNotFoundError
except ImportError:
    ZoneInfo = None
    ZoneInfoNotFoundError = Exception  # fallback pro starší Python
import os
import re
import csv

# Načtení .env při spuštění přes python app.py (volitelné: pip install python-dotenv)
try:
    from dotenv import load_dotenv
    _site_dir = os.path.dirname(os.path.abspath(__file__))
    _env_path = os.path.join(_site_dir, '.env')
    if os.path.isfile(_env_path):
        load_dotenv(_env_path)
except ImportError:
    pass
import io
from collections import defaultdict
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.exceptions import HTTPException
from flask import Response

app = Flask(__name__)

# Konfigurace – umožní nastavení přes proměnné prostředí na PythonAnywhere,
# ale zachová stávající SQLite soubor a fungování aplikace.
SECRET_KEY = os.environ.get('SECRET_KEY', 'THISASIAINSDNSUIAMO<SMINFDBAUENIMPOX<MKDAOBVISWICJO@@@#$#@!')
if SECRET_KEY == 'your-secret-key' and os.environ.get('FLASK_ENV') != 'development':
    import warnings
    warnings.warn(
        "⚠️ VAROVÁNÍ: Používá se defaultní SECRET_KEY! "
        "V produkci nastavte SECRET_KEY jako environment variable! "
        "Aplikace může být zranitelná!",
        UserWarning
    )
    app.logger.warning("⚠️ KRITICKÉ: Používá se defaultní SECRET_KEY! Nastavte SECRET_KEY v produkci!")

app.config['SECRET_KEY'] = SECRET_KEY
_db_url = (os.environ.get('DATABASE_URL') or '').strip()
if not _db_url or len(_db_url) < 10:
    _db_url = 'sqlite:///odbery.db'
elif _db_url.startswith('postgres://'):
    _db_url = 'postgresql://' + _db_url[11:]  # Neon vrací postgres://, SQLAlchemy chce postgresql://
app.config['SQLALCHEMY_DATABASE_URI'] = _db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True,
    'connect_args': {'timeout': 15, 'check_same_thread': False} if 'sqlite' in _db_url.lower() else {},
}
app.config['WTF_CSRF_ENABLED'] = True  # CSRF ochrana zapnuta
app.config['WTF_CSRF_TIME_LIMIT'] = 3600  # 1 hodina

# Session security pro produkci
# V produkci Secure=True (HTTPS). Pro Tailscale pouze HTTP: ALLOW_HTTP_SESSION=1
app.config['SESSION_COOKIE_SECURE'] = (
    os.environ.get('FLASK_ENV') == 'production' and os.environ.get('ALLOW_HTTP_SESSION') != '1'
)
app.config['SESSION_COOKIE_HTTPONLY'] = True  # Ochrana před XSS
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'  # CSRF ochrana
app.config['PERMANENT_SESSION_LIFETIME'] = 3600  # 1 hodina

# Logging konfigurace
if not app.debug:
    import logging
    from logging.handlers import RotatingFileHandler

    app.logger.setLevel(logging.INFO)
    try:
        if not os.path.exists('logs'):
            os.mkdir('logs')
        file_handler = RotatingFileHandler('logs/app.log', maxBytes=10240000, backupCount=10)
        file_handler.setFormatter(logging.Formatter(
            '%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]'
        ))
        file_handler.setLevel(logging.INFO)
        app.logger.addHandler(file_handler)
    except (OSError, PermissionError):
        pass  # bez zápisu do souboru (např. nemáme oprávnění) – log jde jen do stderr
    app.logger.info('Aplikace spuštěna')

# Gzip komprese odpovědí – menší přenos přes síť (Tailscale, WiFi)
try:
    from flask_compress import Compress
    Compress(app)
except ImportError:
    pass

db = SQLAlchemy(app)


@app.teardown_appcontext
def shutdown_session(exception=None):
    """Uzavře DB session po každém requestu – brání úniku připojení a zajišťuje stabilitu."""
    db.session.remove()
    if exception:
        db.session.rollback()


login_manager = LoginManager(app)
login_manager.login_view = 'admin_login'
login_manager.session_protection = 'strong'  # Silnější ochrana session


def _cesky_datum(d, fmt='%d.%m.%Y'):
    """Převede date/datetime na řetězec v českém formátu dd.mm.yyyy (nebo dd.mm.yyyy HH:MM)."""
    if d is None:
        return ''
    if hasattr(d, 'strftime'):
        return d.strftime(fmt)
    return str(d)


# Sentinel pro nevyplněný telefon u odběrů – v DB jen čísla, zobrazení jako "Nezadáno"
TELEFON_NEZADANO_SENTINEL = '+420000000000'


def _odber_telefon_display(telefon):
    """Vrátí 'Nezadáno' pokud telefon je prázdný nebo 000000000, jinak vrátí hodnotu (pro zobrazení/export)."""
    if not telefon or str(telefon).strip() in ('', '000000000', TELEFON_NEZADANO_SENTINEL):
        return 'Nezadáno'
    return telefon


@app.template_filter('odber_telefon')
def odber_telefon_filter(telefon):
    """Šablony: {{ odber.telefon | odber_telefon }} → Nezadáno nebo číslo."""
    return _odber_telefon_display(telefon)


@app.template_filter('cesky_datum')
def cesky_datum_filter(d):
    """Šablony: {{ hodnota | cesky_datum }} → dd.mm.yyyy."""
    return _cesky_datum(d, '%d.%m.%Y')


@app.template_filter('cesky_datum_cas')
def cesky_datum_cas_filter(d):
    """Šablony: {{ hodnota | cesky_datum_cas }} → dd.mm.yyyy HH:MM."""
    return _cesky_datum(d, '%d.%m.%Y %H:%M')


@app.context_processor
def inject_csrf_token():
    """Zajistí, že csrf_token() je dostupný ve všech šablonách."""
    try:
        from flask_wtf.csrf import generate_csrf
        return {'csrf_token': generate_csrf}
    except Exception:
        return {'csrf_token': lambda: ''}


# Časová zóna pro ČR (na Windows může chybět IANA data – pak použijeme lokální čas)
try:
    CZ_TZ = ZoneInfo("Europe/Prague") if ZoneInfo else None
except (ZoneInfoNotFoundError, Exception):
    CZ_TZ = None

def get_current_time():
    return datetime.now(CZ_TZ) if CZ_TZ else datetime.now()


def _is_postgresql():
    """True pokud používáme PostgreSQL (Render + Neon)."""
    return 'postgresql' in _db_url.lower()


def _db_year_eq(column, year):
    """Filtr: rok sloupce == year. SQLite: strftime, PostgreSQL: extract."""
    y = int(year) if year is not None else None
    if y is None:
        return db.literal(False)
    if _is_postgresql():
        return db.func.extract('year', column) == y
    return db.func.strftime('%Y', column) == str(y)


def _db_month_eq(column, month):
    """Filtr: měsíc sloupce == month. SQLite: strftime, PostgreSQL: extract."""
    m = int(month) if month is not None else None
    if m is None:
        return db.literal(False)
    if _is_postgresql():
        return db.func.extract('month', column) == m
    return db.func.strftime('%m', column) == f'{m:02d}'


# Modely
# Asociační tabulka pro many-to-many vztah mezi User a Pobocka
user_pobocky = db.Table('user_pobocky',
    db.Column('user_id', db.Integer, db.ForeignKey('user.id'), primary_key=True),
    db.Column('pobocka_id', db.Integer, db.ForeignKey('pobocka.id'), primary_key=True)
)

class User(UserMixin, db.Model):
    __tablename__ = 'users' if 'postgresql' in _db_url.lower() else 'user'  # "user" je rezervované v PostgreSQL
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(100), unique=True, nullable=False)
    password = db.Column(db.String(100), nullable=False)
    pin = db.Column(db.String(10), unique=True, nullable=True)  # PIN pro rychlé přihlášení
    pobocka_id = db.Column(db.Integer, db.ForeignKey('pobocka.id'), nullable=True)  # Přiřazená pobočka (zachováno pro zpětnou kompatibilitu)
    role = db.Column(db.String(20), default='user')  # 'admin' nebo 'user'
    jmeno = db.Column(db.String(100), nullable=True)  # Jméno uživatele pro zobrazení
    
    # Relationships
    pobocka = db.relationship('Pobocka', foreign_keys=[pobocka_id], backref='users_old', lazy=True)  # Zpětná kompatibilita
    pobocky = db.relationship('Pobocka', secondary=user_pobocky, lazy='subquery', backref=db.backref('users', lazy=True))  # Many-to-many

    # Pomocné metody pro práci s hesly – ulehčí případné další změny.
    def set_password(self, raw_password: str) -> None:
        self.password = generate_password_hash(raw_password)

    def check_password(self, raw_password: str) -> bool:
        # Zpětná kompatibilita: pokud je v DB staré "plain text" heslo,
        # dovolíme jedno poslední přihlášení a hned ho převedeme na hash.
        if self.password.startswith(('pbkdf2:', 'scrypt:', 'sha256:')):
            return check_password_hash(self.password, raw_password)
        return self.password == raw_password

    def is_admin(self) -> bool:
        """Vrací True pokud je uživatel admin."""
        return self.role == 'admin'

    def can_access_pobocka(self, pobocka_id: int) -> bool:
        """Vrací True pokud může uživatel přistupovat k dané pobočce."""
        if self.is_admin():
            return True
        # Zkontrolujeme many-to-many vztah
        if pobocka_id in [p.id for p in self.pobocky]:
            return True
        # Zpětná kompatibilita - starý pobocka_id
        return self.pobocka_id == pobocka_id
    
    def get_all_pobocky_ids(self):
        """Vrací seznam ID všech poboček, ke kterým má uživatel přístup."""
        if self.is_admin():
            return None  # Admin má přístup ke všem
        pobocky_ids = [p.id for p in self.pobocky]
        # Zpětná kompatibilita
        if self.pobocka_id and self.pobocka_id not in pobocky_ids:
            pobocky_ids.append(self.pobocka_id)
        return pobocky_ids

class Pobocka(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nazev = db.Column(db.String(100), nullable=False)
    adresa = db.Column(db.String(200), nullable=True)  # Adresa pobočky
    firma = db.Column(db.String(200), nullable=True)  # Název firmy

class Odber(db.Model):
    __table_args__ = (
        db.Index('ix_odber_pobocka_stav', 'pobocka_id', 'stav'),
        db.Index('ix_odber_pobocka_datum', 'pobocka_id', 'datum'),
    )
    id = db.Column(db.Integer, primary_key=True)
    pobocka_id = db.Column(db.Integer, db.ForeignKey('pobocka.id'), nullable=False)
    jmeno = db.Column(db.String(100), nullable=False)
    kdo_zadal = db.Column(db.String(100), nullable=False)
    telefon = db.Column(db.String(12), nullable=True)
    placeno_predem = db.Column(db.Boolean, default=False)
    datum = db.Column(db.Date, nullable=False)
    castka = db.Column(db.Float, nullable=True)
    poznamky = db.Column(db.Text, nullable=True)
    stav = db.Column(db.String(20), default='aktivní')

class Akce(db.Model):
    __table_args__ = (db.Index('ix_akce_pobocka_datum', 'pobocka_id', 'datum'),)
    id = db.Column(db.Integer, primary_key=True)
    odber_id = db.Column(db.Integer, db.ForeignKey('odber.id'), nullable=True)   # None = admin/systémová akce
    uzivatel = db.Column(db.String(100), nullable=False)
    akce = db.Column(db.String(100), nullable=False)
    datum = db.Column(db.DateTime, nullable=False)
    pobocka_id = db.Column(db.Integer, db.ForeignKey('pobocka.id'), nullable=True)  # None = admin/systémová akce


class Reklamace(db.Model):
    """Reklamace elektronických cigaret."""
    __table_args__ = (
        db.Index('ix_reklamace_pobocka_archived', 'pobocka_id', 'archived'),
        db.Index('ix_reklamace_pobocka_stav', 'pobocka_id', 'stav'),
        db.Index('ix_reklamace_pobocka_datum', 'pobocka_id', 'datum_prijmu'),
    )
    id = db.Column(db.Integer, primary_key=True)
    pobocka_id = db.Column(db.Integer, db.ForeignKey('pobocka.id'), nullable=False)
    zakaznik = db.Column(db.String(120), nullable=False)
    telefon = db.Column(db.String(20), nullable=True)
    znacka = db.Column(db.String(100), nullable=False)
    model = db.Column(db.String(100), nullable=True)
    barva = db.Column(db.String(50), nullable=True)  # Barva zboží
    datum_prijmu = db.Column(db.Date, nullable=False)
    datum_zakoupeni = db.Column(db.Date, nullable=True)  # Datum zakoupení zboží
    popis_zavady = db.Column(db.Text, nullable=False)
    zjistena_zavada_nas = db.Column(db.Text, nullable=True)  # Zjištěná závada ze strany prodejny
    stav = db.Column(db.String(30), default='Čeká')  # Čeká / Výměna kus za kus / Posláno do Ústí / Zamítnuto
    sleva_procent = db.Column(db.Float, nullable=True)  # Procenta slevy při Zamítnuto (reklamace na krajíčku)
    reseni = db.Column(db.Text, nullable=True)
    cena = db.Column(db.Float, nullable=True)
    poznamky = db.Column(db.Text, nullable=True)
    zavolano_zakaznikovi = db.Column(db.Boolean, default=False, nullable=False)
    prijal = db.Column(db.String(100), nullable=True)  # Kdo přijal reklamaci
    datum_vyrizeni = db.Column(db.Date, nullable=True)  # Datum vyřízení (při přechodu do Prošlo/Zamítnuto)
    archived = db.Column(db.Boolean, default=False, nullable=False)  # Archivováno – skryto z běžného přehledu
    archived_at = db.Column(db.DateTime, nullable=True)  # Kdy bylo archivováno
    created_at = db.Column(db.DateTime, default=get_current_time, nullable=False)

    pobocka = db.relationship('Pobocka', backref='reklamace', lazy=True)


class ReklamaceLog(db.Model):
    """Log změn a akcí nad reklamacemi pro admin historii."""
    __table_args__ = (db.Index('ix_reklamace_log_pobocka_datum', 'pobocka_id', 'datum'),)
    id = db.Column(db.Integer, primary_key=True)
    reklamace_id = db.Column(db.Integer, db.ForeignKey('reklamace.id'), nullable=False)
    uzivatel = db.Column(db.String(100), nullable=False)
    akce = db.Column(db.String(255), nullable=False)
    datum = db.Column(db.DateTime, nullable=False)
    pobocka_id = db.Column(db.Integer, db.ForeignKey('pobocka.id'), nullable=False)

# Formuláře
class PridatOdberForm(FlaskForm):
    jmeno = StringField('Jméno a příjmení', validators=[DataRequired(message='Zadejte jméno zákazníka'), Length(max=200, message='Jméno může mít maximálně 200 znaků')])
    telefon = StringField('Telefon', validators=[Optional(), Regexp(r'^\s*$|^\d{9}$', message='Telefon musí být prázdné nebo 9 číslic')])
    placeno_predem = BooleanField('Placeno předem')
    datum = DateField('Datum objednávky', validators=[DataRequired()])
    castka = FloatField('Částka v Kč', validators=[Optional()])
    poznamky = TextAreaField('Poznámky', validators=[Optional(), Length(max=5000, message='Poznámky mohou mít maximálně 5000 znaků')])
    kdo_zadal = SelectField('Zadal', coerce=str, validators=[DataRequired()])

class LoginForm(FlaskForm):
    """Přihlášení: buď PIN (4–10 číslic), nebo uživatelské jméno + heslo."""
    pin = PasswordField('PIN', validators=[Optional(), Regexp(r'^\d{4,10}$', message='PIN musí být 4-10 číslic')])
    username = StringField('Uživatelské jméno', validators=[Optional(), Length(max=100)])
    password = PasswordField('Heslo', validators=[Optional(), Length(min=6, max=100, message='Heslo musí mít 6-100 znaků')])

class AddPobockaForm(FlaskForm):
    nazev = StringField('Název pobočky', validators=[DataRequired(), Regexp(r'^[a-zA-ZáčďéěíňóřšťúůýžÁČĎÉĚÍŇÓŘŠŤÚŮÝŽ0-9\s\-\.]{2,100}$', message='Název musí obsahovat 2-100 znaků (pouze písmena, čísla, mezery, pomlčky a tečky)')])

class AddUserForm(FlaskForm):
    jmeno = StringField('Jméno uživatele', validators=[DataRequired(), Length(max=100, message='Jméno může mít maximálně 100 znaků')])
    pin = StringField('PIN (4-10 číslic)', validators=[DataRequired(), Regexp(r'^\d{4,10}$', message='PIN musí být 4-10 číslic')])
    pobocky = SelectField('Pobočky (může být více)', choices=[], validators=[Optional()], render_kw={'multiple': True, 'size': 5})  # Multiple select
    role = SelectField('Role', choices=[('user', 'Uživatel'), ('admin', 'Admin')], default='user', validators=[DataRequired()])

class EditUserForm(FlaskForm):
    jmeno = StringField('Jméno uživatele', validators=[DataRequired(), Length(max=100, message='Jméno může mít maximálně 100 znaků')])
    pin = StringField('PIN (4-10 číslic)', validators=[Optional(), Regexp(r'^\d{4,10}$', message='PIN musí být 4-10 číslic')])
    pobocky = SelectField('Pobočky (může být více)', choices=[], validators=[Optional()], render_kw={'multiple': True, 'size': 5})
    role = SelectField('Role', choices=[('user', 'Uživatel'), ('admin', 'Admin')], validators=[DataRequired()])

class EditPobockaForm(FlaskForm):
    nazev = StringField('Název pobočky', validators=[DataRequired(), Regexp(r'^[a-zA-ZáčďéěíňóřšťúůýžÁČĎÉĚÍŇÓŘŠŤÚŮÝŽ0-9\s\-\.]{2,100}$', message='Název musí obsahovat 2-100 znaků (pouze písmena, čísla, mezery, pomlčky a tečky)')])
    adresa = StringField('Adresa pobočky', validators=[Optional(), Length(max=200, message='Adresa může mít maximálně 200 znaků')])
    firma = StringField('Název firmy', validators=[Optional(), Length(max=200, message='Název firmy může mít maximálně 200 znaků')])


class ReklamaceForm(FlaskForm):
    zakaznik = StringField('Zákazník', validators=[DataRequired(message='Zadejte jméno zákazníka'), Length(max=200, message='Jméno zákazníka může mít maximálně 200 znaků')])
    telefon = StringField('Telefon', validators=[DataRequired(message='Zadejte telefonní číslo'), Regexp(r'^\d{9}$', message='Telefon musí být 9 číslic')])
    znacka = StringField('Značka', validators=[DataRequired(message='Zadejte značku zboží'), Length(max=100, message='Značka může mít maximálně 100 znaků')])
    model = StringField('Model', validators=[DataRequired(message='Zadejte model zboží'), Length(max=100, message='Model může mít maximálně 100 znaků')])
    barva = StringField('Barva', validators=[Optional(), Length(max=50, message='Barva může mít maximálně 50 znaků')])
    datum_prijmu = DateField('Datum přijetí', validators=[DataRequired()])
    datum_zakoupeni = DateField('Datum zakoupení', validators=[Optional()])
    popis_zavady = TextAreaField('Zjištěná závada (popis zákazníka)', validators=[DataRequired(), Length(max=2000, message='Popis závady může mít maximálně 2000 znaků')])
    zjistena_zavada_nas = TextAreaField('Zjištěná závada ze strany nás', validators=[Optional(), Length(max=2000)])
    stav = SelectField(
        'Stav',
        choices=[
            ('Čeká', '⏳ Čeká – čeká na vyřízení'),
            ('Výměna kus za kus', '✅ Prošlo – výměna kus za kus'),
            ('Posláno do Ústí', '📦 Posláno do Ústí – odesláno k vyřízení'),
            ('Zamítnuto', '❌ Zamítnuto – reklamace nebyla uznána'),
        ],
        default='Čeká',
        validators=[DataRequired()],
    )
    sleva_procent = FloatField('Sleva v %', validators=[Optional()], render_kw={'placeholder': 'např. 15', 'min': 0, 'max': 100, 'step': 0.5})
    reseni = TextAreaField('Řešení', validators=[Optional(), Length(max=2000, message='Řešení může mít maximálně 2000 znaků')])
    poznamky = TextAreaField('Poznámky', validators=[Optional(), Length(max=5000, message='Poznámky mohou mít maximálně 5000 znaků')])
    zavolano_zakaznikovi = BooleanField('Zavoláno zákazníkovi', validators=[Optional()], default=False)


class ReklamaceEditForm(ReklamaceForm):
    """Form pro editaci reklamace – stejné pole jako create."""
    pass


def get_odbery_stats_for_pobocky(pobocky, rok=None):
    """Vrátí statistiky odběrů pro seznam poboček.
    Aktivní = aktuálně aktivní (bez filtru roku). Zelené/červené = z aktivních podle data.
    Vydáno/nevyzvednuto = v daném roce (rok=None = aktuální rok).
    """
    if not pobocky:
        return []
    ids = [p.id for p in pobocky]
    dnes = date.today()
    rok = rok or dnes.year
    by_p = {p.id: {'aktivni': 0, 'zelene': 0, 'cervene': 0, 'vydano': 0, 'nevyzvednuto': 0} for p in pobocky}
    try:
        aktivni_list = Odber.query.filter(Odber.pobocka_id.in_(ids), Odber.stav == 'aktivní').all()
        for o in aktivni_list:
            if o.pobocka_id not in by_p:
                continue
            by_p[o.pobocka_id]['aktivni'] += 1
            if (dnes - o.datum).days <= 7:
                by_p[o.pobocka_id]['zelene'] += 1
            else:
                by_p[o.pobocka_id]['cervene'] += 1
        rok_str = str(rok)
        q_vydano = (db.session.query(Odber.pobocka_id, db.func.count(Odber.id))
            .filter(Odber.pobocka_id.in_(ids), Odber.stav == 'vydáno',
                    _db_year_eq(Odber.datum, rok_str))
            .group_by(Odber.pobocka_id))
        for pid, cnt in q_vydano.all():
            if pid in by_p:
                by_p[pid]['vydano'] = cnt
        q_nevyzvednuto = (db.session.query(Odber.pobocka_id, db.func.count(Odber.id))
            .filter(Odber.pobocka_id.in_(ids), Odber.stav == 'nevyzvednuto',
                    _db_year_eq(Odber.datum, rok_str))
            .group_by(Odber.pobocka_id))
        for pid, cnt in q_nevyzvednuto.all():
            if pid in by_p:
                by_p[pid]['nevyzvednuto'] = cnt
    except Exception:
        for p in pobocky:
            aktivni = Odber.query.filter_by(pobocka_id=p.id, stav='aktivní').all()
            by_p[p.id]['aktivni'] = len(aktivni)
            for o in aktivni:
                if (dnes - o.datum).days <= 7:
                    by_p[p.id]['zelene'] += 1
                else:
                    by_p[p.id]['cervene'] += 1
            vydano_q = Odber.query.filter_by(pobocka_id=p.id, stav='vydáno').filter(
                _db_year_eq(Odber.datum, rok))
            nevyzvednuto_q = Odber.query.filter_by(pobocka_id=p.id, stav='nevyzvednuto').filter(
                _db_year_eq(Odber.datum, rok))
            by_p[p.id]['vydano'] = vydano_q.count()
            by_p[p.id]['nevyzvednuto'] = nevyzvednuto_q.count()
    return [
        {'pobocka_id': p.id, 'nazev': p.nazev, 'aktivni': by_p[p.id]['aktivni'],
         'zelene': by_p[p.id]['zelene'], 'cervene': by_p[p.id]['cervene'],
         'vydano': by_p[p.id]['vydano'], 'nevyzvednuto': by_p[p.id]['nevyzvednuto']}
        for p in pobocky
    ]


def get_reklamace_stats_for_pobocky(pobocky, rok=None):
    """Vrátí statistiky reklamací pro seznam poboček. rok=None = všechny roky (celkem).
    sleva = Zamítnuto se sleva_procent (reklamace na krajíčku).
    Archivované reklamace se do statistik nezapočítávají."""
    if not pobocky:
        return []
    ids = [p.id for p in pobocky]
    by_p = {p.id: {'celkem': 0, 'ceka': 0, 'vymena': 0, 'poslano': 0, 'zamitnuto': 0, 'sleva': 0, 'vyrizene': 0} for p in pobocky}
    try:
        q = (db.session.query(Reklamace.pobocka_id, Reklamace.stav, db.func.count(Reklamace.id))
             .filter(Reklamace.pobocka_id.in_(ids), Reklamace.archived == False))
        if rok is not None:
            q = q.filter(_db_year_eq(Reklamace.datum_prijmu, rok))
        rows = q.group_by(Reklamace.pobocka_id, Reklamace.stav).all()
        for pid, stav, cnt in rows:
            if pid not in by_p:
                continue
            by_p[pid]['celkem'] += cnt
            if stav == 'Čeká':
                by_p[pid]['ceka'] = cnt
            elif stav == 'Výměna kus za kus':
                by_p[pid]['vymena'] = cnt
                by_p[pid]['vyrizene'] += cnt
            elif stav == 'Posláno do Ústí':
                by_p[pid]['poslano'] = cnt
                by_p[pid]['vyrizene'] += cnt
            elif stav == 'Zamítnuto':
                by_p[pid]['zamitnuto'] = cnt
        # sleva = Zamítnuto se sleva_procent
        q_sleva = (db.session.query(Reklamace.pobocka_id, db.func.count(Reklamace.id))
                   .filter(Reklamace.pobocka_id.in_(ids), Reklamace.archived == False, Reklamace.stav == 'Zamítnuto', Reklamace.sleva_procent.isnot(None)))
        if rok is not None:
            q_sleva = q_sleva.filter(_db_year_eq(Reklamace.datum_prijmu, rok))
        for pid, cnt in q_sleva.group_by(Reklamace.pobocka_id).all():
            if pid in by_p:
                by_p[pid]['sleva'] = cnt
    except Exception:
        for p in pobocky:
            base = Reklamace.query.filter_by(pobocka_id=p.id, archived=False)
            if rok is not None:
                base = base.filter(_db_year_eq(Reklamace.datum_prijmu, rok))
            by_p[p.id]['celkem'] = base.count()
            by_p[p.id]['ceka'] = base.filter_by(stav='Čeká').count()
            by_p[p.id]['vymena'] = base.filter_by(stav='Výměna kus za kus').count()
            by_p[p.id]['poslano'] = base.filter_by(stav='Posláno do Ústí').count()
            by_p[p.id]['zamitnuto'] = base.filter_by(stav='Zamítnuto').count()
            by_p[p.id]['sleva'] = base.filter_by(stav='Zamítnuto').filter(Reklamace.sleva_procent.isnot(None)).count()
            by_p[p.id]['vyrizene'] = by_p[p.id]['vymena'] + by_p[p.id]['poslano']
    return [
        {'pobocka_id': p.id, 'nazev': p.nazev, 'celkem': by_p[p.id]['celkem'],
         'ceka': by_p[p.id]['ceka'], 'vymena': by_p[p.id]['vymena'],
         'poslano': by_p[p.id]['poslano'], 'zamitnuto': by_p[p.id]['zamitnuto'],
         'sleva': by_p[p.id]['sleva'], 'vyrizene': by_p[p.id]['vyrizene']}
        for p in pobocky
    ]


def _system_pobocka_id():
    """Vrací ID první pobočky pro audit záznamy bez konkrétní pobočky (admin akce)."""
    p = Pobocka.query.order_by(Pobocka.id).first()
    return p.id if p else None


def log_reklamace_action(reklamace: Reklamace, text: str) -> None:
    """Pomocná funkce pro uložení záznamu o akci nad reklamacemi."""
    uzivatel = current_user.username if current_user.is_authenticated else 'system'
    log = ReklamaceLog(
        reklamace_id=reklamace.id,
        uzivatel=uzivatel,
        akce=text,
        datum=get_current_time(),
        pobocka_id=reklamace.pobocka_id,
    )
    db.session.add(log)

# Migrace databáze - přidání nových sloupců do existující tabulky user
def migrate_db():
    """Přidá chybějící sloupce do tabulky user, pokud neexistují.
    U PostgreSQL se migrace přeskočí – create_all() vytvoří kompletní schéma."""
    if _is_postgresql():
        return
    # NEPOUŽÍVÁME app.app_context() zde, protože to už je v init_db()
    try:
        # Zkontrolujeme, jestli sloupce existují pomocí PRAGMA (SQLite)
        result = db.session.execute(db.text("PRAGMA table_info(user)"))
        columns = [row[1] for row in result.fetchall()]
        
        # Přidáme chybějící sloupce
        if 'pin' not in columns:
            db.session.execute(db.text('ALTER TABLE user ADD COLUMN pin VARCHAR(10)'))
            db.session.commit()
        if 'pobocka_id' not in columns:
            db.session.execute(db.text('ALTER TABLE user ADD COLUMN pobocka_id INTEGER'))
            db.session.commit()
        if 'role' not in columns:
            db.session.execute(db.text('ALTER TABLE user ADD COLUMN role VARCHAR(20) DEFAULT "user"'))
            db.session.commit()
        if 'jmeno' not in columns:
            db.session.execute(db.text('ALTER TABLE user ADD COLUMN jmeno VARCHAR(100)'))
            db.session.commit()
    except Exception as e:
        # Pokud kontrola selže, zkusíme přidat sloupce přímo (pokud už neexistují)
        # SQLite vrátí chybu, pokud sloupec už existuje, což je OK
        try:
            db.session.execute(db.text('ALTER TABLE user ADD COLUMN pin VARCHAR(10)'))
            db.session.commit()
        except Exception:
            db.session.rollback()
        try:
            db.session.execute(db.text('ALTER TABLE user ADD COLUMN pobocka_id INTEGER'))
            db.session.commit()
        except Exception:
            db.session.rollback()
        try:
            db.session.execute(db.text('ALTER TABLE user ADD COLUMN role VARCHAR(20) DEFAULT "user"'))
            db.session.commit()
        except Exception:
            db.session.rollback()
        try:
            db.session.execute(db.text('ALTER TABLE user ADD COLUMN jmeno VARCHAR(100)'))
            db.session.commit()
        except Exception:
            db.session.rollback()
    
    # Migrace pro Reklamace tabulku
    try:
        result = db.session.execute(db.text("PRAGMA table_info(reklamace)"))
        reklamace_columns = [row[1] for row in result.fetchall()]
        
        if 'zavolano_zakaznikovi' not in reklamace_columns:
            db.session.execute(db.text('ALTER TABLE reklamace ADD COLUMN zavolano_zakaznikovi BOOLEAN DEFAULT 0'))
            db.session.commit()
            app.logger.info('Přidán sloupec zavolano_zakaznikovi do Reklamace')
    except Exception:
        db.session.rollback()
        try:
            db.session.execute(db.text('ALTER TABLE reklamace ADD COLUMN zavolano_zakaznikovi BOOLEAN DEFAULT 0'))
            db.session.commit()
        except Exception:
            db.session.rollback()
    
    try:
        result = db.session.execute(db.text("PRAGMA table_info(reklamace)"))
        reklamace_columns = [row[1] for row in result.fetchall()]
        
        if 'prijal' not in reklamace_columns:
            db.session.execute(db.text('ALTER TABLE reklamace ADD COLUMN prijal VARCHAR(100)'))
            db.session.commit()
            app.logger.info('Přidán sloupec prijal do Reklamace')
    except Exception:
        db.session.rollback()
        try:
            db.session.execute(db.text('ALTER TABLE reklamace ADD COLUMN prijal VARCHAR(100)'))
            db.session.commit()
        except Exception:
            db.session.rollback()
    
    # Migrace pro nová pole v Reklamace: barva a datum_zakoupeni
    try:
        result = db.session.execute(db.text("PRAGMA table_info(reklamace)"))
        reklamace_columns = [row[1] for row in result.fetchall()]
        
        if 'barva' not in reklamace_columns:
            db.session.execute(db.text('ALTER TABLE reklamace ADD COLUMN barva VARCHAR(50)'))
            db.session.commit()
            app.logger.info('Přidán sloupec barva do Reklamace')
    except Exception:
        db.session.rollback()
        try:
            db.session.execute(db.text('ALTER TABLE reklamace ADD COLUMN barva VARCHAR(50)'))
            db.session.commit()
        except Exception:
            db.session.rollback()
    
    try:
        result = db.session.execute(db.text("PRAGMA table_info(reklamace)"))
        reklamace_columns = [row[1] for row in result.fetchall()]
        
        if 'datum_zakoupeni' not in reklamace_columns:
            db.session.execute(db.text('ALTER TABLE reklamace ADD COLUMN datum_zakoupeni DATE'))
            db.session.commit()
            app.logger.info('Přidán sloupec datum_zakoupeni do Reklamace')
    except Exception:
        db.session.rollback()
        try:
            db.session.execute(db.text('ALTER TABLE reklamace ADD COLUMN datum_zakoupeni DATE'))
            db.session.commit()
        except Exception:
            db.session.rollback()
    
    # Migrace: sleva_procent v Reklamace
    try:
        result = db.session.execute(db.text("PRAGMA table_info(reklamace)"))
        reklamace_columns = [row[1] for row in result.fetchall()]
        if 'sleva_procent' not in reklamace_columns:
            db.session.execute(db.text('ALTER TABLE reklamace ADD COLUMN sleva_procent REAL'))
            db.session.commit()
            app.logger.info('Přidán sloupec sleva_procent do Reklamace')
    except Exception:
        db.session.rollback()
        try:
            db.session.execute(db.text('ALTER TABLE reklamace ADD COLUMN sleva_procent REAL'))
            db.session.commit()
        except Exception:
            db.session.rollback()
    
    # Migrace: stav Sleva -> Zamítnuto (sleva je nyní jen sloupec při Zamítnuto)
    try:
        db.session.execute(db.text("UPDATE reklamace SET stav = 'Zamítnuto' WHERE stav = 'Sleva'"))
        updated = db.session.execute(db.text("SELECT changes()")).scalar()
        if updated and updated > 0:
            db.session.commit()
            app.logger.info(f'Migrace: {updated} reklamací převedeno ze stavu Sleva na Zamítnuto')
    except Exception:
        db.session.rollback()
    
    # Migrace: archived a archived_at pro archivaci starších reklamací
    try:
        result = db.session.execute(db.text("PRAGMA table_info(reklamace)"))
        reklamace_columns = [row[1] for row in result.fetchall()]
        if 'archived' not in reklamace_columns:
            db.session.execute(db.text('ALTER TABLE reklamace ADD COLUMN archived BOOLEAN DEFAULT 0'))
            db.session.commit()
            app.logger.info('Přidán sloupec archived do Reklamace')
    except Exception:
        db.session.rollback()
        try:
            db.session.execute(db.text('ALTER TABLE reklamace ADD COLUMN archived BOOLEAN DEFAULT 0'))
            db.session.commit()
        except Exception:
            db.session.rollback()
    try:
        result = db.session.execute(db.text("PRAGMA table_info(reklamace)"))
        reklamace_columns = [row[1] for row in result.fetchall()]
        if 'archived_at' not in reklamace_columns:
            db.session.execute(db.text('ALTER TABLE reklamace ADD COLUMN archived_at DATETIME'))
            db.session.commit()
            app.logger.info('Přidán sloupec archived_at do Reklamace')
    except Exception:
        db.session.rollback()
        try:
            db.session.execute(db.text('ALTER TABLE reklamace ADD COLUMN archived_at DATETIME'))
            db.session.commit()
        except Exception:
            db.session.rollback()
    try:
        result = db.session.execute(db.text("PRAGMA table_info(reklamace)"))
        reklamace_columns = [row[1] for row in result.fetchall()]
        if 'zjistena_zavada_nas' not in reklamace_columns:
            db.session.execute(db.text('ALTER TABLE reklamace ADD COLUMN zjistena_zavada_nas TEXT'))
            db.session.commit()
            app.logger.info('Přidán sloupec zjistena_zavada_nas do Reklamace')
    except Exception:
        db.session.rollback()
        try:
            db.session.execute(db.text('ALTER TABLE reklamace ADD COLUMN zjistena_zavada_nas TEXT'))
            db.session.commit()
        except Exception:
            db.session.rollback()
    try:
        result = db.session.execute(db.text("PRAGMA table_info(reklamace)"))
        reklamace_columns = [row[1] for row in result.fetchall()]
        if 'datum_vyrizeni' not in reklamace_columns:
            db.session.execute(db.text('ALTER TABLE reklamace ADD COLUMN datum_vyrizeni DATE'))
            db.session.commit()
            app.logger.info('Přidán sloupec datum_vyrizeni do Reklamace')
    except Exception:
        db.session.rollback()
        try:
            db.session.execute(db.text('ALTER TABLE reklamace ADD COLUMN datum_vyrizeni DATE'))
            db.session.commit()
        except Exception:
            db.session.rollback()
    
    # Migrace pro nová pole v Pobocka: adresa a firma
    try:
        result = db.session.execute(db.text("PRAGMA table_info(pobocka)"))
        pobocka_columns = [row[1] for row in result.fetchall()]
        
        if 'adresa' not in pobocka_columns:
            db.session.execute(db.text('ALTER TABLE pobocka ADD COLUMN adresa VARCHAR(200)'))
            db.session.commit()
            app.logger.info('Přidán sloupec adresa do Pobocka')
    except Exception:
        db.session.rollback()
        try:
            db.session.execute(db.text('ALTER TABLE pobocka ADD COLUMN adresa VARCHAR(200)'))
            db.session.commit()
        except Exception:
            db.session.rollback()
    
    try:
        result = db.session.execute(db.text("PRAGMA table_info(pobocka)"))
        pobocka_columns = [row[1] for row in result.fetchall()]
        
        if 'firma' not in pobocka_columns:
            db.session.execute(db.text('ALTER TABLE pobocka ADD COLUMN firma VARCHAR(200)'))
            db.session.commit()
            app.logger.info('Přidán sloupec firma do Pobocka')
    except Exception:
        db.session.rollback()
        try:
            db.session.execute(db.text('ALTER TABLE pobocka ADD COLUMN firma VARCHAR(200)'))
            db.session.commit()
        except Exception:
            db.session.rollback()
    
    # Migrace tabulky akce: odber_id a pobocka_id nullable (pro admin/systémové záznamy)
    try:
        result = db.session.execute(db.text("PRAGMA table_info(akce)"))
        rows = result.fetchall()
        # SQLite: (cid, name, type, notnull, dflt_value, pk) – notnull je index 3
        need_migrate = False
        for r in rows:
            if r[1] == 'odber_id' and r[3]:
                need_migrate = True
                break
        if need_migrate:
            db.session.execute(db.text("""
                CREATE TABLE akce_new (
                    id INTEGER PRIMARY KEY,
                    odber_id INTEGER,
                    uzivatel VARCHAR(100) NOT NULL,
                    akce VARCHAR(100) NOT NULL,
                    datum DATETIME NOT NULL,
                    pobocka_id INTEGER
                )
            """))
            db.session.execute(db.text("INSERT INTO akce_new (id, odber_id, uzivatel, akce, datum, pobocka_id) SELECT id, odber_id, uzivatel, akce, datum, pobocka_id FROM akce"))
            db.session.execute(db.text("DROP TABLE akce"))
            db.session.execute(db.text("ALTER TABLE akce_new RENAME TO akce"))
            db.session.commit()
            app.logger.info('Migrace: tabulka akce – odber_id a pobocka_id nyní nullable')
    except Exception as e:
        db.session.rollback()
        app.logger.debug(f'Migrace akce (nullable): %s', e)
    
    # Indexy pro rychlejší dotazy (CREATE INDEX IF NOT EXISTS – bezpečné pro existující DB)
    for name, sql in [
        ('ix_odber_pobocka_stav', 'CREATE INDEX IF NOT EXISTS ix_odber_pobocka_stav ON odber(pobocka_id, stav)'),
        ('ix_odber_pobocka_datum', 'CREATE INDEX IF NOT EXISTS ix_odber_pobocka_datum ON odber(pobocka_id, datum)'),
        ('ix_reklamace_pobocka_archived', 'CREATE INDEX IF NOT EXISTS ix_reklamace_pobocka_archived ON reklamace(pobocka_id, archived)'),
        ('ix_reklamace_pobocka_stav', 'CREATE INDEX IF NOT EXISTS ix_reklamace_pobocka_stav ON reklamace(pobocka_id, stav)'),
        ('ix_reklamace_pobocka_datum', 'CREATE INDEX IF NOT EXISTS ix_reklamace_pobocka_datum ON reklamace(pobocka_id, datum_prijmu)'),
        ('ix_akce_pobocka_datum', 'CREATE INDEX IF NOT EXISTS ix_akce_pobocka_datum ON akce(pobocka_id, datum)'),
        ('ix_reklamace_log_pobocka_datum', 'CREATE INDEX IF NOT EXISTS ix_reklamace_log_pobocka_datum ON reklamace_log(pobocka_id, datum)'),
    ]:
        try:
            db.session.execute(db.text(sql))
            db.session.commit()
        except Exception:
            db.session.rollback()

# Inicializace databáze
def init_db():
    with app.app_context():
        db.create_all()
        # SQLite: zapnutí WAL režimu pro plynulejší a rychlejší zápisy
        try:
            if 'sqlite' in (os.environ.get('DATABASE_URL') or 'sqlite:///').lower():
                db.session.execute(db.text('PRAGMA journal_mode=WAL'))
                db.session.execute(db.text('PRAGMA synchronous=NORMAL'))
                db.session.execute(db.text('PRAGMA cache_size=-64000'))  # 64 MB cache
                db.session.commit()
        except Exception:
            db.session.rollback()
        # Spustíme migraci před vytvářením dat - MUSÍ být před jakýmkoliv dotazem na modely
        migrate_db()
        
        # Po migraci musíme znovu načíst metadata, aby SQLAlchemy věděl o nových sloupcích
        # Použijeme raw SQL dotaz pro kontrolu existence poboček
        try:
            result = db.session.execute(db.text("SELECT COUNT(*) FROM pobocka"))
            pobocka_count = result.scalar()
            if pobocka_count == 0:
                pobocky = [Pobocka(nazev='Teplice'), Pobocka(nazev='Děčín')]
                db.session.bulk_save_objects(pobocky)
                db.session.commit()
        except Exception as e:
            app.logger.error(f'Chyba při kontrole poboček: {str(e)}')
            # Pokud tabulka neexistuje, vytvoříme ji pomocí create_all
            db.create_all()
            pobocky = [Pobocka(nazev='Teplice'), Pobocka(nazev='Děčín')]
            db.session.bulk_save_objects(pobocky)
            db.session.commit()
        
        try:
            user_table = User.__tablename__
            result = db.session.execute(db.text(f"SELECT COUNT(*) FROM {user_table}"))
            user_count = result.scalar()
            if user_count == 0:
                # Vytvoření defaultního admina
                admin = User(
                    username='admin',
                    pin='0000',  # Default PIN pro admina
                    role='admin',
                    jmeno='Administrátor'
                )
                admin.set_password('admin123')
                db.session.add(admin)
                db.session.commit()
            else:
                # Aktualizujeme existující admina, pokud nemá PIN - použijeme raw SQL
                try:
                    result = db.session.execute(db.text(f"SELECT id, pin, role, jmeno FROM {user_table} WHERE username = 'admin'"))
                    admin_row = result.fetchone()
                    if admin_row:
                        admin_id, pin, role, jmeno = admin_row
                        if not pin:
                            db.session.execute(db.text(f"UPDATE {user_table} SET pin = '0000', role = 'admin', jmeno = COALESCE(jmeno, 'Administrátor') WHERE id = :id"), {'id': admin_id})
                            db.session.commit()
                except Exception as e:
                    app.logger.error(f'Chyba při aktualizaci admina: {str(e)}')
                    db.session.rollback()
        except Exception as e:
            app.logger.error(f'Chyba při kontrole uživatelů: {str(e)}')
            # Pokud tabulka neexistuje, vytvoříme ji
            db.create_all()
            admin = User(
                username='admin',
                pin='0000',
                role='admin',
                jmeno='Administrátor'
            )
            admin.set_password('admin123')
            db.session.add(admin)
            db.session.commit()
        # No default zadavatel creation

init_db()

@login_manager.user_loader
def load_user(user_id):
    """Načte uživatele – chrání před neplatným user_id a chybami."""
    try:
        return db.session.get(User, int(user_id))
    except (ValueError, TypeError):
        return None

# Routy
@app.route('/')
def index():
    # Pokud není přihlášený, přesměruj na login
    if not current_user.is_authenticated:
        return redirect(url_for('admin_login'))
    try:
        pobocky = get_user_pobocky()
        aktualni_rok = date.today().year
        reklamace_stat = get_reklamace_stats_for_pobocky(pobocky, rok=aktualni_rok)
        odbery_stat = get_odbery_stats_for_pobocky(pobocky, rok=aktualni_rok)
        return render_template('index.html', pobocky=pobocky, reklamace_stat=reklamace_stat, odbery_stat=odbery_stat, stat_rok=aktualni_rok)
    except Exception as e:
        app.logger.exception('Chyba na index (GET /): %s', e)
        raise


@app.route('/reklamace')
@login_required
def reklamace_index():
    """Hlavní stránka reklamací – přehled napříč pobočkami."""
    pobocky = get_user_pobocky()
    prehled = get_reklamace_stats_for_pobocky(pobocky)
    return render_template('reklamace_index.html', prehled=prehled)

@app.route('/branch/<int:pobocka_id>', methods=['GET', 'POST'])
@login_required
def branch(pobocka_id):
    pobocka = Pobocka.query.get_or_404(pobocka_id)
    # Kontrola přístupu k pobočce
    if not current_user.can_access_pobocka(pobocka_id):
        flash('Nemáte přístup k této pobočce!', 'danger')
        return redirect(url_for('index'))

    try:
        form_data = request.form if request.method == 'POST' and request.form else None
    except Exception:
        form_data = None
    form = PridatOdberForm(formdata=form_data)
    # Automaticky vyplníme zadavatele podle přihlášeného uživatele
    zadavatel_jmeno = current_user.jmeno or current_user.username

    # Nastavíme automaticky vybraného zadavatele
    form.kdo_zadal.choices = [(zadavatel_jmeno, zadavatel_jmeno)]
    form.kdo_zadal.data = zadavatel_jmeno

    validacni_chyba = None
    jmeno_ok = telefon_ok = True
    if request.method == 'POST':
        try:
            jmeno_ok = bool((form.jmeno.data or '').strip())
            telefon_raw = re.sub(r'\s', '', (request.form.get('telefon') or '').strip())
            telefon_ok = not telefon_raw or bool(re.match(r'^\d{9}$', telefon_raw))
            if not jmeno_ok:
                validacni_chyba = 'Zadejte jméno zákazníka.'
            elif not telefon_ok:
                validacni_chyba = 'Telefon musí být prázdný nebo 9 číslic.'
        except Exception as e:
            app.logger.warning('Branch POST validace: %s', e)
            validacni_chyba = 'Neplatná data formuláře.'

    if form.validate_on_submit() and not validacni_chyba and jmeno_ok and telefon_ok:
        # Zadavatel je automaticky nastaven, takže kontrola není potřeba
        if not form.kdo_zadal.data:
            form.kdo_zadal.data = zadavatel_jmeno
        
        # Telefon je volitelný – pokud prázdný, ukládáme +420000000000 (jen čísla v DB), zobrazení "Nezadáno"
        telefon_input = re.sub(r'\s', '', (request.form.get('telefon', '') or '').strip())
        if telefon_input and re.match(r'^\d{9}$', telefon_input):
            telefon = f'+420{telefon_input}'
        else:
            telefon = TELEFON_NEZADANO_SENTINEL
        
        try:
            odber = Odber(
                pobocka_id=pobocka_id,
                jmeno=(form.jmeno.data or '').strip(),
                kdo_zadal=form.kdo_zadal.data or zadavatel_jmeno,
                telefon=telefon,
                placeno_predem=bool(form.placeno_predem.data),
                datum=form.datum.data or date.today(),
                castka=None if form.placeno_predem.data else form.castka.data,
                poznamky=form.poznamky.data if form.poznamky.data is not None else ''
            )
            db.session.add(odber)
            db.session.commit()
            akce = Akce(
                odber_id=odber.id,
                uzivatel=current_user.username if current_user.is_authenticated else form.kdo_zadal.data,
                akce=f'Přidán odběr: {form.jmeno.data}',
                datum=get_current_time(),
                pobocka_id=pobocka_id
            )
            db.session.add(akce)
            db.session.commit()
            flash('Odběr přidán!', 'success')
            return redirect(url_for('branch', pobocka_id=pobocka_id))
        except Exception as e:
            db.session.rollback()
            flash(f'Chyba při přidávání odběru: {str(e)}', 'danger')
            _q = request.args.get('q', '').strip()
            _odbery = Odber.query.filter_by(pobocka_id=pobocka_id, stav='aktivní').order_by(Odber.datum.desc()).all()
            return render_template('branch.html', form=form, odbery=_odbery, pobocka=pobocka, prehled={}, validacni_chyba=None, filter_q=_q)

    # Filtry (GET parametry) – server-side vyhledávání
    filter_q = request.args.get('q', '').strip()
    odbery_query = Odber.query.filter_by(pobocka_id=pobocka_id, stav='aktivní')
    if filter_q:
        like = f"%{filter_q}%"
        odbery_query = odbery_query.filter(
            db.or_(
                Odber.jmeno.ilike(like),
                Odber.telefon.ilike(like),
                Odber.poznamky.ilike(like),
            )
        )
    odbery = odbery_query.order_by(Odber.datum.desc()).all()
    dnes = date.today()
    zelene = 0
    cervene = 0
    for odber in odbery:
        odber.dni = (dnes - odber.datum).days
        odber.barva = 'table-success' if odber.dni <= 7 else 'table-danger'
        if odber.dni <= 7:
            zelene += 1
        else:
            cervene += 1
    prehled = {
        'celkem': len(odbery),
        'zelene': zelene,
        'cervene': cervene
    }
    return render_template('branch.html', form=form, odbery=odbery, pobocka=pobocka, prehled=prehled, validacni_chyba=validacni_chyba if request.method == 'POST' else None, filter_q=filter_q)


@app.route('/branch/<int:pobocka_id>/export.csv')
@login_required
def branch_export_csv(pobocka_id):
    """Export aktivních odběrů pobočky do CSV."""
    pobocka = Pobocka.query.get_or_404(pobocka_id)
    if not current_user.can_access_pobocka(pobocka_id):
        flash('Nemáte přístup k této pobočce.', 'danger')
        return redirect(url_for('index'))
    odbery = Odber.query.filter_by(pobocka_id=pobocka_id, stav='aktivní').order_by(Odber.datum.desc()).all()
    output = io.StringIO()
    output.write('\ufeff')
    writer = csv.writer(output, lineterminator='\r\n')
    writer.writerow(['ID', 'Jméno', 'Telefon', 'Datum', 'Částka', 'Placeno předem', 'Poznámky', 'Zadal', 'Stav'])
    for o in odbery:
        writer.writerow([
            o.id, o.jmeno or '', _odber_telefon_display(o.telefon), _cesky_datum(o.datum),
            o.castka or '', 'Ano' if o.placeno_predem else 'Ne', o.poznamky or '', o.kdo_zadal or '', o.stav or 'aktivní'
        ])
    safe_name = re.sub(r'[^\w\s\-]', '', (pobocka.nazev or 'pobocka').replace(' ', '_'))[:50]
    filename = f"odbery_{safe_name}_{date.today().isoformat()}_{datetime.now().strftime('%H%M%S')}.csv"
    return Response(
        output.getvalue().encode('utf-8'),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


@app.route('/reklamace/branch/<int:pobocka_id>', methods=['GET', 'POST'])
@login_required
def reklamace_branch(pobocka_id):
    pobocka = Pobocka.query.get_or_404(pobocka_id)
    # Kontrola přístupu k pobočce
    if not current_user.can_access_pobocka(pobocka_id):
        flash('Nemáte přístup k této pobočce!', 'danger')
        return redirect(url_for('reklamace_index'))
    form = ReklamaceForm(formdata=request.form if request.method == 'POST' else None)

    # Auto datum (default dnes), pokud uživatel nic nevybral
    if request.method == 'GET' and not form.datum_prijmu.data:
        form.datum_prijmu.data = date.today()

    # Explicitní kontrola povinných polí (custom input pro telefon neprochází vždy WTForms)
    validacni_chyba = None
    req_ok = True
    if request.method == 'POST':
        zakaznik_ok = bool((form.zakaznik.data or '').strip())
        telefon_ok = bool(re.match(r'^\d{9}$', (request.form.get('telefon') or '').strip()))
        znacka_ok = bool((form.znacka.data or '').strip())
        model_ok = bool((form.model.data or '').strip())
        popis_ok = bool((form.popis_zavady.data or '').strip())
        datum_zak_ok = form.datum_zakoupeni.data is not None
        zaruka_ok = True
        if datum_zak_ok and form.datum_zakoupeni.data:
            try:
                d = form.datum_zakoupeni.data
                zaruka_do = date(d.year + 2, d.month, d.day)
            except ValueError:
                zaruka_do = date(d.year + 2, 2, 28)  # 29.2. v přestupném roce
            zaruka_ok = zaruka_do >= date.today()
        req_ok = zakaznik_ok and telefon_ok and znacka_ok and model_ok and popis_ok and datum_zak_ok and zaruka_ok
        if not zakaznik_ok:
            validacni_chyba = 'Zadejte jméno zákazníka.'
        elif not telefon_ok:
            validacni_chyba = 'Zadejte platné telefonní číslo (9 číslic).'
        elif not znacka_ok:
            validacni_chyba = 'Zadejte značku zboží.'
        elif not model_ok:
            validacni_chyba = 'Zadejte model zboží.'
        elif not popis_ok:
            validacni_chyba = 'Zadejte popis závady.'
        elif not datum_zak_ok:
            validacni_chyba = 'Zadejte datum zakoupení (záruka 2 roky).'
        elif not zaruka_ok:
            validacni_chyba = 'Záruka 2 roky již vypršela – nelze přijmout reklamaci.'

    if form.validate_on_submit() and req_ok and not validacni_chyba:
        telefon_input = (request.form.get('telefon') or '').strip()
        telefon = f'+420{telefon_input}' if re.match(r'^\d{9}$', telefon_input) else None
        try:
            reklamace = Reklamace(
                pobocka_id=pobocka_id,
                zakaznik=(form.zakaznik.data or '').strip(),
                telefon=telefon,
                znacka=(form.znacka.data or '').strip(),
                model=(form.model.data or '').strip() or None,
                barva=(form.barva.data or '').strip() or None,
                datum_prijmu=form.datum_prijmu.data,
                datum_zakoupeni=form.datum_zakoupeni.data,
                popis_zavady=(form.popis_zavady.data or '').strip(),
                zjistena_zavada_nas=(form.zjistena_zavada_nas.data or '').strip() or None,
                stav=form.stav.data,
                datum_vyrizeni=date.today() if form.stav.data in ('Výměna kus za kus', 'Zamítnuto') else None,
                sleva_procent=form.sleva_procent.data if form.stav.data == 'Zamítnuto' else None,
                reseni=(form.reseni.data or '').strip() or None,
                poznamky=form.poznamky.data if form.poznamky.data is not None else '',
                zavolano_zakaznikovi=bool(form.zavolano_zakaznikovi.data),
                prijal=current_user.jmeno or current_user.username or 'unknown',
            )
            db.session.add(reklamace)
            db.session.flush()  # aby mělo reklamace.id hodnotu
            log_reklamace_action(reklamace, f'Vytvořena reklamace (stav: {reklamace.stav})')
            db.session.commit()
            flash('Reklamace byla uložena. Otevírám PDF…', 'success')
            return redirect(url_for('reklamace_print', reklamace_id=reklamace.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Chyba při ukládání reklamace: {str(e)}', 'danger')

    # Filtry (GET parametry)
    stav = request.args.get('stav', '').strip()
    q = request.args.get('q', '').strip()
    date_from = request.args.get('from', '').strip()
    date_to = request.args.get('to', '').strip()
    show_archived = request.args.get('archived', '').strip().lower() in ('1', 'true', 'ano', 'yes')

    reklamace_query = Reklamace.query.filter_by(pobocka_id=pobocka_id)
    if not show_archived:
        reklamace_query = reklamace_query.filter(Reklamace.archived == False)
    if stav:
        reklamace_query = reklamace_query.filter(Reklamace.stav == stav)
    if q:
        like = f"%{q}%"
        reklamace_query = reklamace_query.filter(
            db.or_(
                Reklamace.zakaznik.ilike(like),
                Reklamace.telefon.ilike(like),
                Reklamace.znacka.ilike(like),
                Reklamace.model.ilike(like),
                Reklamace.barva.ilike(like),
            )
        )
    if date_from:
        try:
            reklamace_query = reklamace_query.filter(Reklamace.datum_prijmu >= datetime.strptime(date_from, "%Y-%m-%d").date())
        except ValueError:
            pass
    if date_to:
        try:
            reklamace_query = reklamace_query.filter(Reklamace.datum_prijmu <= datetime.strptime(date_to, "%Y-%m-%d").date())
        except ValueError:
            pass

    reklamace_qs = reklamace_query.order_by(Reklamace.datum_prijmu.desc(), Reklamace.id.desc()).all()
    return render_template(
        'reklamace_branch.html',
        pobocka=pobocka,
        form=form,
        reklamace=reklamace_qs,
        filter_stav=stav,
        filter_q=q,
        filter_from=date_from,
        filter_to=date_to,
        filter_archived=show_archived,
        validacni_chyba=validacni_chyba,
    )


@app.route('/reklamace/<int:reklamace_id>/edit', methods=['GET', 'POST'])
@login_required
def reklamace_edit(reklamace_id):
    """Úprava reklamace - vyžaduje přihlášení a oprávnění k pobočce."""
    reklamace = Reklamace.query.get_or_404(reklamace_id)
    pobocka = Pobocka.query.get_or_404(reklamace.pobocka_id)
    
    # Ověření oprávnění k pobočce
    if not current_user.can_access_pobocka(reklamace.pobocka_id):
        flash('Nemáte přístup k této pobočce!', 'danger')
        return redirect(url_for('index'))
    
    form = ReklamaceEditForm(
        formdata=request.form if request.method == 'POST' else None,
        obj=reklamace if request.method == 'GET' else None
    )
    # Načteme checkbox hodnotu
    if request.method == 'GET' and reklamace.zavolano_zakaznikovi:
        form.zavolano_zakaznikovi.data = True

    validacni_chyba = None
    req_ok = True
    if request.method == 'POST':
        zakaznik_ok = bool((form.zakaznik.data or '').strip())
        telefon_ok = bool(re.match(r'^\d{9}$', (request.form.get('telefon') or '').strip()))
        znacka_ok = bool((form.znacka.data or '').strip())
        model_ok = bool((form.model.data or '').strip())
        popis_ok = bool((form.popis_zavady.data or '').strip())
        datum_zak_ok = form.datum_zakoupeni.data is not None
        zaruka_ok = True
        if datum_zak_ok and form.datum_zakoupeni.data:
            try:
                d = form.datum_zakoupeni.data
                zaruka_do = date(d.year + 2, d.month, d.day)
            except ValueError:
                zaruka_do = date(d.year + 2, 2, 28)  # 29.2. v přestupném roce
            zaruka_ok = zaruka_do >= date.today()
        req_ok = zakaznik_ok and telefon_ok and znacka_ok and model_ok and popis_ok and datum_zak_ok and zaruka_ok
        if not zakaznik_ok:
            validacni_chyba = 'Zadejte jméno zákazníka.'
        elif not telefon_ok:
            validacni_chyba = 'Zadejte platné telefonní číslo (9 číslic).'
        elif not znacka_ok:
            validacni_chyba = 'Zadejte značku zboží.'
        elif not model_ok:
            validacni_chyba = 'Zadejte model zboží.'
        elif not popis_ok:
            validacni_chyba = 'Zadejte popis závady.'
        elif not datum_zak_ok:
            validacni_chyba = 'Zadejte datum zakoupení (záruka 2 roky).'
        elif not zaruka_ok:
            validacni_chyba = 'Záruka 2 roky již vypršela – nelze přijmout reklamaci.'

    if form.validate_on_submit() and req_ok and not validacni_chyba:
        telefon_input = (request.form.get('telefon') or '').strip()
        telefon = f'+420{telefon_input}' if re.match(r'^\d{9}$', telefon_input) else None
        try:
            reklamace.zakaznik = (form.zakaznik.data or '').strip()
            reklamace.telefon = telefon
            reklamace.znacka = (form.znacka.data or '').strip()
            reklamace.model = (form.model.data or '').strip() or None
            reklamace.barva = (form.barva.data or '').strip() or None
            reklamace.datum_prijmu = form.datum_prijmu.data
            reklamace.datum_zakoupeni = form.datum_zakoupeni.data
            reklamace.popis_zavady = (form.popis_zavady.data or '').strip()
            reklamace.zjistena_zavada_nas = (form.zjistena_zavada_nas.data or '').strip() or None
            reklamace.stav = form.stav.data
            if form.stav.data in ('Výměna kus za kus', 'Zamítnuto') and reklamace.datum_vyrizeni is None:
                reklamace.datum_vyrizeni = date.today()
            reklamace.sleva_procent = form.sleva_procent.data if form.stav.data == 'Zamítnuto' else None
            reklamace.reseni = (form.reseni.data or '').strip() or None
            reklamace.poznamky = form.poznamky.data if form.poznamky.data is not None else ''
            reklamace.zavolano_zakaznikovi = bool(form.zavolano_zakaznikovi.data)
            # Pokud ještě není nastaveno kdo přijal, nastavíme to
            if not reklamace.prijal:
                reklamace.prijal = current_user.jmeno or current_user.username
            log_reklamace_action(reklamace, f'Upravena reklamace (stav: {reklamace.stav})')
            db.session.commit()
            flash('Reklamace byla upravena.', 'success')
            return redirect(url_for('reklamace_branch', pobocka_id=pobocka.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Chyba při ukládání úprav: {str(e)}', 'danger')

    # Naplnění inputu telefonu (bez +420)
    telefon_plain = ''
    if reklamace.telefon and reklamace.telefon.startswith('+420'):
        telefon_plain = reklamace.telefon.replace('+420', '').strip()
    return render_template('reklamace_edit.html', pobocka=pobocka, reklamace=reklamace, form=form, telefon_plain=telefon_plain, validacni_chyba=validacni_chyba if request.method == 'POST' else None)


@app.route('/reklamace/<int:reklamace_id>/status', methods=['POST'])
@login_required
def reklamace_change_status(reklamace_id):
    """Rychlá změna stavu reklamace z tabulky - vyžaduje přihlášení a oprávnění k pobočce."""
    try:
        reklamace = Reklamace.query.get_or_404(reklamace_id)
        pobocka_id = reklamace.pobocka_id

        if not current_user.can_access_pobocka(pobocka_id):
            flash('Nemáte přístup k této pobočce!', 'danger')
            return redirect(url_for('index'))

        action = (request.form.get('action') or '').strip()
        mapping = {
            'ceka': 'Čeká',
            'vymena': 'Výměna kus za kus',
            'poslano_usti': 'Posláno do Ústí',
            'zamitnuto': 'Zamítnuto',
        }
        new_status = mapping.get(action)

        if not new_status:
            flash('Neplatná akce pro změnu stavu.', 'danger')
            return redirect(url_for('reklamace_branch', pobocka_id=pobocka_id))

        reklamace.stav = new_status
        if new_status in ('Výměna kus za kus', 'Zamítnuto') and reklamace.datum_vyrizeni is None:
            reklamace.datum_vyrizeni = date.today()
        log_reklamace_action(reklamace, f'Změněn stav reklamace na {new_status}')
        db.session.commit()
        flash(f'Stav reklamace změněn na "{new_status}".', 'success')
    except HTTPException:
        raise
    except Exception:
        db.session.rollback()
        app.logger.exception(f'Chyba při změně stavu reklamace {reklamace_id}')
        flash('Došlo k chybě při změně stavu. Zkuste to znovu nebo obnovte stránku.', 'danger')
        try:
            r = Reklamace.query.get(reklamace_id)
            pobocka_id = r.pobocka_id if r else None
        except Exception:
            pobocka_id = None
        if pobocka_id is not None:
            return redirect(url_for('reklamace_branch', pobocka_id=pobocka_id))
        return redirect(url_for('index'))
    return redirect(url_for('reklamace_branch', pobocka_id=pobocka_id))


@app.route('/reklamace/<int:reklamace_id>/archive', methods=['POST'])
@login_required
def reklamace_archive(reklamace_id):
    """Archivace vyřízené reklamace – přesune ji z běžného přehledu do archivu."""
    try:
        reklamace = Reklamace.query.get_or_404(reklamace_id)
        pobocka_id = reklamace.pobocka_id

        if not current_user.can_access_pobocka(pobocka_id):
            flash('Nemáte přístup k této pobočce!', 'danger')
            return redirect(url_for('index'))

        vyrizene_stavy = ('Výměna kus za kus', 'Zamítnuto')
        if reklamace.stav not in vyrizene_stavy:
            flash('Archivovat lze pouze vyřízené reklamace (Prošlo – výměna nebo Zamítnuto).', 'warning')
            return redirect(url_for('reklamace_branch', pobocka_id=pobocka_id))

        if reklamace.archived:
            flash('Reklamace je již archivována.', 'info')
            return redirect(url_for('reklamace_branch', pobocka_id=pobocka_id))

        reklamace.archived = True
        reklamace.archived_at = get_current_time()
        log_reklamace_action(reklamace, f'Archivována reklamace (stav: {reklamace.stav})')
        db.session.commit()
        flash('Reklamace byla archivována. Prohlížení v Admin → Archiv reklamací.', 'success')
    except HTTPException:
        raise
    except Exception:
        db.session.rollback()
        app.logger.exception(f'Chyba při archivaci reklamace {reklamace_id}')
        flash('Došlo k chybě při archivaci. Zkuste to znovu nebo obnovte stránku.', 'danger')
        try:
            r = Reklamace.query.get(reklamace_id)
            pobocka_id = r.pobocka_id if r else None
        except Exception:
            pobocka_id = None
        if pobocka_id is not None:
            return redirect(url_for('reklamace_branch', pobocka_id=pobocka_id))
        return redirect(url_for('index'))
    return redirect(url_for('reklamace_branch', pobocka_id=pobocka_id))


@app.route('/reklamace/<int:reklamace_id>/print')
@login_required
def reklamace_print(reklamace_id):
    """Tisk reklamace - vyžaduje přihlášení a oprávnění k pobočce."""
    reklamace = Reklamace.query.get_or_404(reklamace_id)
    pobocka = Pobocka.query.get_or_404(reklamace.pobocka_id)
    
    # Ověření oprávnění k pobočce
    if not current_user.can_access_pobocka(reklamace.pobocka_id):
        flash('Nemáte přístup k této pobočce!', 'danger')
        return redirect(url_for('index'))
    
    try:
        log_reklamace_action(reklamace, 'Vytištěn PDF reklamačního formuláře')
        db.session.commit()
    except Exception:
        db.session.rollback()
    
    return render_template('reklamace_print.html', reklamace=reklamace, pobocka=pobocka)


@app.route('/reklamace/<int:reklamace_id>/protokol')
@login_required
def reklamace_protokol(reklamace_id):
    """Protokol o vyřízení reklamace – závěrečný doklad pro zákazníka (tisk pouze u vyřízených)."""
    reklamace = Reklamace.query.get_or_404(reklamace_id)
    pobocka = Pobocka.query.get_or_404(reklamace.pobocka_id)
    if not current_user.can_access_pobocka(reklamace.pobocka_id):
        flash('Nemáte přístup k této pobočce!', 'danger')
        return redirect(url_for('index'))
    vyrizene_stavy = ('Výměna kus za kus', 'Zamítnuto')
    if reklamace.stav not in vyrizene_stavy:
        flash('Protokol lze vytisknout až u reklamace ve závěrečném stavu (Prošlo – výměna nebo Zamítnuto).', 'warning')
        return redirect(url_for('reklamace_branch', pobocka_id=pobocka.id))
    datum_vyrizeni_cz = _cesky_datum(reklamace.datum_vyrizeni) if reklamace.datum_vyrizeni else _cesky_datum(date.today())
    return render_template('reklamace_protokol.html', reklamace=reklamace, pobocka=pobocka, datum_vyrizeni_cz=datum_vyrizeni_cz)


@app.route('/admin/export/all.xlsx')
@login_required
def admin_export_excel():
    """Export všech dat do Excel souboru."""
    if not (current_user.is_authenticated and current_user.is_admin()):
        flash('Nemáte oprávnění!', 'danger')
        return redirect(url_for('index'))
    
    if not HAS_EXCEL:
        flash('Excel export není dostupný. Nainstalujte openpyxl: pip install openpyxl', 'warning')
        return redirect(url_for('admin_dashboard'))
    
    try:
        wb = Workbook()
        
        # List pro odběry
        ws_odbery = wb.active
        ws_odbery.title = "Odběry"
        ws_odbery.append(['ID', 'Pobočka', 'Jméno', 'Telefon', 'Datum', 'Stav', 'Částka', 'Kdo zadal', 'Poznámky'])
        
        header_fill = PatternFill(start_color="4a90e2", end_color="4a90e2", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws_odbery[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        pobocky_dict = {p.id: p.nazev for p in Pobocka.query.all()}
        odbery = Odber.query.order_by(Odber.datum.desc()).all()
        for odber in odbery:
            pobocka_nazev = pobocky_dict.get(odber.pobocka_id, "N/A")
            row = ws_odbery.append([
                odber.id,
                pobocka_nazev,
                odber.jmeno or '',
                _odber_telefon_display(odber.telefon),
                _cesky_datum(odber.datum),
                odber.stav or '',
                odber.castka or 0,
                odber.kdo_zadal or '',
                odber.poznamky or ''
            ])
            # Nastavíme formátování pro textové buňky
            for col_idx in [2, 3, 5, 8, 9]:  # Pobočka, Jméno, Stav, Kdo zadal, Poznámky
                cell = ws_odbery.cell(row=ws_odbery.max_row, column=col_idx)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        # List pro reklamace
        ws_reklamace = wb.create_sheet("Reklamace")
        ws_reklamace.append(['ID', 'Pobočka', 'Zákazník', 'Telefon', 'Značka', 'Model', 'Barva', 'Datum přijmu', 'Datum zakoupení', 'Stav', 'Cena', 'Zavoláno', 'Přijal', 'Archivováno', 'Poznámky'])
        
        for cell in ws_reklamace[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        reklamace = Reklamace.query.order_by(Reklamace.datum_prijmu.desc()).all()
        for rekl in reklamace:
            pobocka_nazev = pobocky_dict.get(rekl.pobocka_id, "N/A")
            row = ws_reklamace.append([
                rekl.id,
                pobocka_nazev,
                rekl.zakaznik or '',
                rekl.telefon or '',
                rekl.znacka or '',
                rekl.model or '',
                rekl.barva or '',
                _cesky_datum(rekl.datum_prijmu),
                _cesky_datum(rekl.datum_zakoupeni),
                rekl.stav or '',
                rekl.cena or 0,
                'Ano' if rekl.zavolano_zakaznikovi else 'Ne',
                rekl.prijal or '',
                'Ano' if rekl.archived else 'Ne',
                rekl.poznamky or ''
            ])
            # Nastavíme formátování pro textové buňky
            for col_idx in [2, 3, 4, 5, 6, 10, 13, 15]:  # Pobočka, Zákazník, Značka, Model, Barva, Stav, Přijal, Poznámky
                cell = ws_reklamace.cell(row=ws_reklamace.max_row, column=col_idx)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        # Auto-width sloupců
        for ws in [ws_odbery, ws_reklamace]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f'export_vse_{date.today().strftime("%Y%m%d")}_{datetime.now().strftime("%H%M%S")}.xlsx'
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}
        )
    except Exception as e:
        app.logger.error(f'Chyba při exportu do Excel: {str(e)}')
        flash(f'Chyba při exportu: {str(e)}', 'danger')
        return redirect(url_for('admin_dashboard'))


@app.route('/admin/export/all.csv')
@login_required
def admin_export_all():
    """Export všech dat pro admina."""
    if not (current_user.is_authenticated and current_user.is_admin()):
        flash('Nemáte oprávnění!', 'danger')
        return redirect(url_for('index'))
    
    output = io.StringIO()
    output.write('\ufeff')
    writer = csv.writer(output, lineterminator='\r\n')
    
    # Export všech reklamací
    writer.writerow(['=== REKLAMACE ==='])
    writer.writerow([
        'ID', 'Pobočka', 'Zákazník', 'Telefon', 'Značka', 'Model', 'Barva', 'Datum přijmu', 'Datum zakoupení',
        'Popis závady', 'Zjištěná závada ze strany nás', 'Stav', 'Datum vyřízení', 'Řešení', 'Poznámky', 'Vytvořeno'
    ])
    pobocky_dict = {p.id: p.nazev for p in Pobocka.query.all()}
    for r in Reklamace.query.order_by(Reklamace.datum_prijmu.desc()).all():
        writer.writerow([
            r.id, pobocky_dict.get(r.pobocka_id, 'Neznámá'), r.zakaznik, r.telefon or '',
            r.znacka, r.model or '', r.barva or '', 
            _cesky_datum(r.datum_prijmu),
            _cesky_datum(r.datum_zakoupeni),
            (r.popis_zavady or '').replace('\n', ' ').strip(),
            (r.zjistena_zavada_nas or '').replace('\n', ' ').strip(),
            r.stav,
            _cesky_datum(r.datum_vyrizeni) if r.datum_vyrizeni else '',
            r.reseni or '', r.poznamky or '',
            _cesky_datum(r.created_at, '%d.%m.%Y %H:%M')
        ])
    
    writer.writerow([])
    writer.writerow(['=== ODBĚRY ==='])
    writer.writerow([
        'ID', 'Pobočka', 'Zadavatel', 'Datum', 'Stav', 'Poznámky'
    ])
    for o in Odber.query.order_by(Odber.datum.desc()).all():
        writer.writerow([
            o.id, pobocky_dict.get(o.pobocka_id, 'Neznámá'), o.kdo_zadal or '',
            _cesky_datum(o.datum), o.stav, o.poznamky or ''
        ])
    
    return Response(
        output.getvalue().encode('utf-8'),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename="admin_export_{date.today().isoformat()}.csv"'}
    )


@app.route('/admin/reklamace-archiv')
@login_required
def admin_reklamace_archiv():
    """Admin: prohlížení všech reklamací včetně archivovaných – vyhledávání, filtry."""
    if not (current_user.is_authenticated and current_user.is_admin()):
        flash('Nemáte oprávnění!', 'danger')
        return redirect(url_for('index'))
    
    q = request.args.get('q', '').strip()
    pobocka_id = request.args.get('pobocka', '').strip()
    stav = request.args.get('stav', '').strip()
    archived_only = request.args.get('archived', '').strip().lower() in ('1', 'true', 'ano', 'yes')
    
    reklamace_query = Reklamace.query
    if pobocka_id:
        try:
            reklamace_query = reklamace_query.filter(Reklamace.pobocka_id == int(pobocka_id))
        except ValueError:
            pass
    if stav:
        reklamace_query = reklamace_query.filter(Reklamace.stav == stav)
    if archived_only:
        reklamace_query = reklamace_query.filter(Reklamace.archived == True)
    if q:
        like = f"%{q}%"
        reklamace_query = reklamace_query.filter(
            db.or_(
                Reklamace.zakaznik.ilike(like),
                Reklamace.telefon.ilike(like),
                Reklamace.znacka.ilike(like),
                Reklamace.model.ilike(like),
                Reklamace.barva.ilike(like),
            )
        )
    
    reklamace_list = reklamace_query.order_by(Reklamace.datum_prijmu.desc(), Reklamace.id.desc()).limit(500).all()
    pobocky = Pobocka.query.order_by(Pobocka.nazev).all()
    
    return render_template(
        'admin_reklamace_archiv.html',
        reklamace=reklamace_list,
        pobocky=pobocky,
        filter_q=q,
        filter_pobocka=pobocka_id,
        filter_stav=stav,
        filter_archived=archived_only,
    )


@app.route('/reklamace/branch/<int:pobocka_id>/export.csv')
@login_required
def reklamace_export_csv(pobocka_id):
    """Export reklamací pobočky do CSV s jasným názvem souboru."""
    pobocka = Pobocka.query.get_or_404(pobocka_id)
    if not current_user.can_access_pobocka(pobocka_id):
        flash('Nemáte přístup k této pobočce.', 'danger')
        return redirect(url_for('reklamace_index'))
    reklamace_qs = Reklamace.query.filter_by(pobocka_id=pobocka_id).order_by(Reklamace.datum_prijmu.desc(), Reklamace.id.desc()).all()

    output = io.StringIO()
    output.write('\ufeff')
    writer = csv.writer(output, lineterminator='\r\n')
    writer.writerow([
        'ID', 'Pobočka', 'Datum přijmu', 'Datum zakoupení', 'Zákazník', 'Telefon',
        'Značka', 'Model', 'Barva', 'Stav', 'Datum vyřízení', 'Popis závady', 'Zjištěná závada ze strany nás', 'Řešení', 'Poznámky', 'Vytvořeno'
    ])
    for r in reklamace_qs:
        writer.writerow([
            r.id, pobocka.nazev,
            _cesky_datum(r.datum_prijmu),
            _cesky_datum(r.datum_zakoupeni),
            r.zakaznik, r.telefon or '', r.znacka, r.model or '', r.barva or '',
            r.stav,
            _cesky_datum(r.datum_vyrizeni) if r.datum_vyrizeni else '',
            (r.popis_zavady or '').replace('\n', ' ').strip(),
            (r.zjistena_zavada_nas or '').replace('\n', ' ').strip(),
            (r.reseni or '').replace('\n', ' ').strip(),
            (r.poznamky or '').replace('\n', ' ').strip(),
            _cesky_datum(r.created_at, '%d.%m.%Y %H:%M'),
        ])

    safe_name = re.sub(r'[^\w\s\-]', '', (pobocka.nazev or 'pobocka').replace(' ', '_'))[:50]
    filename = f"reklamace_{safe_name}_{date.today().isoformat()}_{datetime.now().strftime('%H%M%S')}.csv"
    return Response(
        output.getvalue().encode('utf-8'),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )

@app.route('/update/<int:id>', methods=['POST'])
@login_required
def update(id):
    """Aktualizace stavu odběru - vyžaduje přihlášení a oprávnění k pobočce."""
    try:
        odber = Odber.query.get_or_404(id)
        
        # Ověření oprávnění k pobočce
        if not current_user.can_access_pobocka(odber.pobocka_id):
            flash('Nemáte přístup k této pobočce!', 'danger')
            return redirect(url_for('index'))
        
        akce = request.form.get('action')
        if akce not in ['vydano', 'nevyzvednuto', 'smazat']:
            flash('Neplatná akce!', 'danger')
            return redirect(url_for('branch', pobocka_id=odber.pobocka_id))
        
        if akce == 'vydano':
            odber.stav = 'vydáno'
        elif akce == 'nevyzvednuto':
            odber.stav = 'nevyzvednuto'
        elif akce == 'smazat':
            odber.stav = 'smazano'
        
        akce_log = Akce(
            odber_id=id,
            uzivatel=current_user.username or current_user.jmeno or 'unknown',
            akce=f'Stav změněn na {odber.stav}',
            datum=get_current_time(),
            pobocka_id=odber.pobocka_id
        )
        db.session.add(akce_log)
        db.session.commit()
        flash('Stav aktualizován!', 'success')
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Chyba při aktualizaci stavu odběru {id}: {str(e)}')
        flash(f'Chyba při aktualizaci: {str(e)}', 'danger')
    
    return redirect(url_for('branch', pobocka_id=odber.pobocka_id))


@app.route('/branch/<int:pobocka_id>/bulk-update', methods=['POST'])
@login_required
def branch_bulk_update(pobocka_id):
    """Hromadná změna stavu odběrů – vydáno nebo nevyzvednuto."""
    pobocka = Pobocka.query.get_or_404(pobocka_id)
    if not current_user.can_access_pobocka(pobocka_id):
        flash('Nemáte přístup k této pobočce!', 'danger')
        return redirect(url_for('index'))
    action = request.form.get('action')
    if action not in ('vydano', 'nevyzvednuto'):
        flash('Neplatná akce.', 'danger')
        return redirect(url_for('branch', pobocka_id=pobocka_id, q=request.args.get('q')))
    ids = request.form.getlist('ids')
    if not ids:
        flash('Nejsou vybrány žádné odběry.', 'warning')
        return redirect(url_for('branch', pobocka_id=pobocka_id, q=request.args.get('q')))
    new_stav = 'vydáno' if action == 'vydano' else 'nevyzvednuto'
    count = 0
    try:
        for _id in ids:
            try:
                oid = int(_id)
            except (ValueError, TypeError):
                continue
            odber = Odber.query.filter_by(id=oid, pobocka_id=pobocka_id, stav='aktivní').first()
            if odber:
                odber.stav = new_stav
                db.session.add(Akce(
                    odber_id=odber.id,
                    uzivatel=current_user.username or current_user.jmeno or 'unknown',
                    akce=f'Hromadně: stav změněn na {new_stav}',
                    datum=get_current_time(),
                    pobocka_id=pobocka_id
                ))
                count += 1
        db.session.commit()
        flash(f'{count} odběrů označeno jako „{new_stav}“.', 'success')
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Chyba při hromadné aktualizaci: {str(e)}')
        flash(f'Chyba při aktualizaci: {str(e)}', 'danger')
    q = request.args.get('q')
    return redirect(url_for('branch', pobocka_id=pobocka_id, q=q if q else None))


@app.route('/reklamace/branch/<int:pobocka_id>/bulk-archive', methods=['POST'])
@login_required
def reklamace_bulk_archive(pobocka_id):
    """Hromadná archivace vyřízených reklamací – jen Prošlo nebo Zamítnuto."""
    pobocka = Pobocka.query.get_or_404(pobocka_id)
    if not current_user.can_access_pobocka(pobocka_id):
        flash('Nemáte přístup k této pobočce!', 'danger')
        return redirect(url_for('reklamace_index'))
    vyrizene_stavy = ('Výměna kus za kus', 'Zamítnuto')
    ids = request.form.getlist('ids')
    if not ids:
        flash('Nejsou vybrány žádné reklamace.', 'warning')
        return redirect(url_for('reklamace_branch', pobocka_id=pobocka_id))
    count = 0
    try:
        for _id in ids:
            try:
                rid = int(_id)
            except (ValueError, TypeError):
                continue
            r = Reklamace.query.filter_by(id=rid, pobocka_id=pobocka_id).first()
            if r and r.stav in vyrizene_stavy and not r.archived:
                r.archived = True
                r.archived_at = get_current_time()
                log_reklamace_action(r, f'Archivována reklamace (stav: {r.stav})')
                count += 1
        db.session.commit()
        flash(f'{count} reklamací archivováno.', 'success')
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Chyba při hromadné archivaci: {str(e)}')
        flash(f'Chyba při archivaci: {str(e)}', 'danger')
    return redirect(url_for('reklamace_branch', pobocka_id=pobocka_id, **{k: v for k, v in request.args.items()}))


@app.route('/update_notes/<int:id>', methods=['POST'])
@login_required
def update_notes(id):
    """Aktualizace poznámek odběru - vyžaduje přihlášení a oprávnění k pobočce."""
    try:
        odber = Odber.query.get_or_404(id)
        
        # Ověření oprávnění k pobočce
        if not current_user.can_access_pobocka(odber.pobocka_id):
            return jsonify({'status': 'error', 'message': 'Nemáte přístup k této pobočce!'}), 403
        
        data = request.get_json(silent=True) or {}
        if not isinstance(data, dict):
            return jsonify({'status': 'error', 'message': 'Neplatný formát dat'}), 400
        # Bezpečné načtení poznámek (vždy řetězec, max 5000 znaků)
        new_poznamky = str(data.get('poznamky') or '').strip()[:5000]
        if len(new_poznamky) > 5000:
            return jsonify({'status': 'error', 'message': 'Poznámky jsou příliš dlouhé (max 5000 znaků)'}), 400

        app.logger.info(f'Zpracovávám poznámky pro odber_id: {id}, nové poznámky: {new_poznamky[:50]}...')

        odber.poznamky = new_poznamky
        akce_text = 'Upraveny poznámky: ' + (new_poznamky[:50] + '...' if len(new_poznamky) > 50 else new_poznamky)
        if not new_poznamky:
            akce_text = 'Upraveny poznámky: (prázdné)'

        # Validate pobocka_id
        if not odber.pobocka_id:
            app.logger.error(f'Neplatné pobocka_id pro odber_id: {id}')
            return jsonify({'status': 'error', 'message': 'Neplatná pobočka pro tento odběr'}), 400

        # Ensure uzivatel is not None
        uzivatel = current_user.username or current_user.jmeno or (odber.kdo_zadal or 'unknown')
        if not uzivatel:
            app.logger.warning(f'Uzivatel je None pro odber_id: {id}, nastavuji na "unknown"')
            uzivatel = 'unknown'

        akce = Akce(
            odber_id=id,
            uzivatel=uzivatel,
            akce=akce_text,
            datum=get_current_time(),
            pobocka_id=odber.pobocka_id
        )

        db.session.add(akce)
        db.session.commit()
        app.logger.info(f'Poznámky úspěšně uloženy pro odber_id: {id}, poznamky: {new_poznamky[:50]}...')
        return jsonify({'status': 'success', 'poznamky': new_poznamky})
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Chyba při ukládání poznámek pro odber_id: {id}, chyba: {str(e)}')
        return jsonify({'status': 'error', 'message': f'Chyba při ukládání: {str(e)}'}), 500

def _is_safe_redirect_url(target):
    """Ověří, že redirect URL je relativní (ochrana před open redirect)."""
    if not target or not target.strip():
        return False
    return target.startswith('/') and not target.startswith('//')


@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if current_user.is_authenticated:
        next_url = request.args.get('next')
        if next_url and _is_safe_redirect_url(next_url):
            return redirect(next_url)
        return redirect(url_for('admin_dashboard') if current_user.is_admin() else url_for('index'))

    form = LoginForm()
    if form.validate_on_submit():
        user = None
        pin_val = (form.pin.data or '').strip()
        username_val = (form.username.data or '').strip()
        password_val = (form.password.data or '').strip()

        if pin_val:
            user = User.query.filter_by(pin=pin_val).first()
            if not user:
                flash('Neplatný PIN.', 'danger')
        elif username_val and password_val:
            user = User.query.filter_by(username=username_val).first()
            if not user or not user.check_password(password_val):
                flash('Neplatné uživatelské jméno nebo heslo.', 'danger')
                user = None
        else:
            flash('Zadejte PIN.', 'danger')

        if user:
            login_user(user)
            flash(f'Vítejte, {user.jmeno or user.username}!', 'success')
            next_url = request.args.get('next')
            if next_url and _is_safe_redirect_url(next_url):
                return redirect(next_url)
            return redirect(url_for('admin_dashboard') if user.is_admin() else url_for('index'))

    return render_template('admin_login.html', form=form)

def get_user_pobocky():
    """Vrací seznam poboček, ke kterým má uživatel přístup."""
    if not current_user.is_authenticated:
        return Pobocka.query.all()  # Nepřihlášení vidí vše
    if getattr(current_user, 'is_admin', None) and current_user.is_admin():
        return Pobocka.query.all()
    # Many-to-many přiřazené pobočky
    pobocky = getattr(current_user, 'pobocky', None)
    if pobocky:
        return list(pobocky)
    # Zpětná kompatibilita: jeden pobocka_id
    pid = getattr(current_user, 'pobocka_id', None)
    if pid:
        pobocka = db.session.get(Pobocka, pid)
        return [pobocka] if pobocka else []
    return []

@app.route('/admin/dashboard', methods=['GET', 'POST'])
@login_required
def admin_dashboard():
    # Pouze admin má přístup k dashboardu
    if not (current_user.is_authenticated and current_user.is_admin()):
        flash('Nemáte oprávnění k přístupu k admin dashboardu!', 'danger')
        return redirect(url_for('index'))
    # Filtry podle roku
    selected_year = request.args.get('rok', str(date.today().year))
    try:
        selected_year = int(selected_year)
    except (ValueError, TypeError):
        selected_year = date.today().year
    
    # Filtrování podle pobočky uživatele
    try:
        user_pobocky = get_user_pobocky()
        pobocky_ids = [p.id for p in user_pobocky] if user_pobocky else []
        
        if current_user.is_authenticated and current_user.is_admin():
            pobocky = Pobocka.query.all()
        else:
            pobocky = user_pobocky if user_pobocky else []
        users = User.query.all() if (current_user.is_authenticated and current_user.is_admin()) else []
        
        # Přehled odběrů podle roku – agregované dotazy (místo N dotazů na pobočku)
        odber_agg = db.session.query(
            Odber.pobocka_id, Odber.stav,
            db.func.count(Odber.id).label('cnt'),
            db.func.sum(Odber.castka).label('castka_sum')
        ).filter(_db_year_eq(Odber.datum, selected_year)).group_by(Odber.pobocka_id, Odber.stav).all()
        aktivni_roky = Odber.query.filter(
            _db_year_eq(Odber.datum, selected_year), Odber.stav == 'aktivní'
        ).with_entities(Odber.pobocka_id, Odber.datum).all()
        by_p_odber = defaultdict(lambda: {'aktivni': 0, 'vydano': 0, 'nevyzvednuto': 0, 'smazano': 0, 'castka_vydano': 0, 'zelene': 0, 'cervene': 0, 'celkem_rok': 0})
        for pid, stav, cnt, csum in odber_agg:
            by_p_odber[pid]['celkem_rok'] += cnt
            if stav == 'aktivní':
                by_p_odber[pid]['aktivni'] = cnt
            elif stav == 'vydáno':
                by_p_odber[pid]['vydano'] = cnt
                by_p_odber[pid]['castka_vydano'] = (csum or 0)
            elif stav == 'nevyzvednuto':
                by_p_odber[pid]['nevyzvednuto'] = cnt
            elif stav == 'smazano':
                by_p_odber[pid]['smazano'] = cnt
        dnes = date.today()
        for pid, d in aktivni_roky:
            if d is None:
                continue
            if (dnes - d).days <= 7:
                by_p_odber[pid]['zelene'] += 1
            else:
                by_p_odber[pid]['cervene'] += 1
        prehled = [
            {
                'nazev': p.nazev,
                'aktivni': by_p_odber[p.id]['aktivni'],
                'vydano': by_p_odber[p.id]['vydano'],
                'nevyzvednuto': by_p_odber[p.id]['nevyzvednuto'],
                'smazano': by_p_odber[p.id]['smazano'],
                'castka_vydano': by_p_odber[p.id]['castka_vydano'],
                'zelene': by_p_odber[p.id]['zelene'],
                'cervene': by_p_odber[p.id]['cervene'],
                'celkem_rok': by_p_odber[p.id]['celkem_rok'],
            }
            for p in pobocky
        ]
    except Exception as e:
        app.logger.error(f'Chyba v admin dashboard při načítání dat: {str(e)}')
        pobocky = []
        pobocky_ids = []
        users = []
        prehled = []

    pobocky_dict = {p.id: p.nazev for p in pobocky} if pobocky else {}

    # Historie odběrů a admin akcí (filtrování podle pobočky)
    akce_query = Akce.query
    if current_user.is_authenticated and not current_user.is_admin() and pobocky_ids:
        akce_query = akce_query.filter(Akce.pobocka_id.in_(pobocky_ids))
    akce_logs = []
    for a in akce_query.order_by(Akce.datum.desc()).limit(200).all():
        akce_logs.append({
            'datum': a.datum,
            'pobocka': pobocky_dict.get(a.pobocka_id, 'Není známo'),
            'uzivatel': a.uzivatel,
            'akce': a.akce,
            'typ': 'Odběr / admin'
        })

    # Historie reklamací (filtrování podle pobočky)
    reklamace_query = ReklamaceLog.query
    if current_user.is_authenticated and not current_user.is_admin() and pobocky_ids:
        reklamace_query = reklamace_query.filter(ReklamaceLog.pobocka_id.in_(pobocky_ids))
    reklamace_logs = []
    for rlog in reklamace_query.order_by(ReklamaceLog.datum.desc()).limit(200).all():
        reklamace_logs.append({
            'datum': rlog.datum,
            'pobocka': pobocky_dict.get(rlog.pobocka_id, 'Není známo'),
            'uzivatel': rlog.uzivatel,
            'akce': rlog.akce,
            'typ': 'Reklamace'
        })

    # Společná časová osa
    try:
        historie = sorted(akce_logs + reklamace_logs, key=lambda x: x['datum'], reverse=True)
    except Exception as e:
        app.logger.error(f'Chyba při řazení historie: {str(e)}')
        historie = []
    
    # Přehled reklamací pro admin dashboard (filtrování podle pobočky a roku)
    reklamace_prehled = []
    celkove_statistiky = {
        'celkem_reklamaci': 0,
        'ceka_reklamace': 0,
        'vymena_reklamace': 0,
        'poslano_reklamace': 0,
        'sleva_reklamace': 0,
        'zamitnuto_reklamace': 0,
        'vyrizene_reklamace': 0,
        'celkem_odberu': 0,
        'celkem_pobocek': len(pobocky),
        'celkem_uzivatelu': len(users) if (current_user.is_authenticated and current_user.is_admin()) else 0,
    }
    
    try:
        # Reklamace podle roku – jeden agregovaný dotaz místo N×6 dotazů
        rekl_agg = db.session.query(
            Reklamace.pobocka_id, Reklamace.stav, db.func.count(Reklamace.id).label('cnt')
        ).filter(
            _db_year_eq(Reklamace.datum_prijmu, selected_year), Reklamace.archived == False
        ).group_by(Reklamace.pobocka_id, Reklamace.stav).all()
        rekl_sleva = db.session.query(Reklamace.pobocka_id, db.func.count(Reklamace.id).label('cnt')).filter(
            _db_year_eq(Reklamace.datum_prijmu, selected_year), Reklamace.archived == False,
            Reklamace.stav == 'Zamítnuto', Reklamace.sleva_procent.isnot(None)
        ).group_by(Reklamace.pobocka_id).all()
        by_p_rekl = defaultdict(lambda: {'celkem': 0, 'ceka': 0, 'vymena': 0, 'poslano': 0, 'zamitnuto': 0, 'sleva': 0})
        for pid, stav, cnt in rekl_agg:
            by_p_rekl[pid]['celkem'] += cnt
            if stav == 'Čeká':
                by_p_rekl[pid]['ceka'] = cnt
            elif stav == 'Výměna kus za kus':
                by_p_rekl[pid]['vymena'] = cnt
            elif stav == 'Posláno do Ústí':
                by_p_rekl[pid]['poslano'] = cnt
            elif stav == 'Zamítnuto':
                by_p_rekl[pid]['zamitnuto'] = cnt
        for pid, cnt in rekl_sleva:
            by_p_rekl[pid]['sleva'] = cnt
        reklamace_prehled = [
            {
                'pobocka_id': p.id, 'nazev': p.nazev,
                'celkem': by_p_rekl[p.id]['celkem'], 'ceka': by_p_rekl[p.id]['ceka'],
                'vymena': by_p_rekl[p.id]['vymena'], 'poslano': by_p_rekl[p.id]['poslano'],
                'sleva': by_p_rekl[p.id]['sleva'], 'zamitnuto': by_p_rekl[p.id]['zamitnuto'],
                'vyrizene': by_p_rekl[p.id]['vymena'] + by_p_rekl[p.id]['poslano'],
            }
            for p in pobocky
        ]
        for p in pobocky:
            r = by_p_rekl[p.id]
            celkove_statistiky['celkem_reklamaci'] += r['celkem']
            celkove_statistiky['ceka_reklamace'] = celkove_statistiky.get('ceka_reklamace', 0) + r['ceka']
            celkove_statistiky['vymena_reklamace'] = celkove_statistiky.get('vymena_reklamace', 0) + r['vymena']
            celkove_statistiky['poslano_reklamace'] = celkove_statistiky.get('poslano_reklamace', 0) + r['poslano']
            celkove_statistiky['sleva_reklamace'] = celkove_statistiky.get('sleva_reklamace', 0) + r['sleva']
            celkove_statistiky['zamitnuto_reklamace'] = celkove_statistiky.get('zamitnuto_reklamace', 0) + r['zamitnuto']
            celkove_statistiky['vyrizene_reklamace'] += r['vymena'] + r['poslano']
            celkove_statistiky['celkem_odberu'] += by_p_odber[p.id]['celkem_rok']
    except Exception as e:
        app.logger.error(f'Chyba při načítání reklamací: {str(e)}')
        reklamace_prehled = []

    # Měsíční data pro grafy (Chart.js)
    chart_odbery = [0] * 12
    chart_reklamace = [0] * 12
    trend_odber = None  # % změna vs minulý rok
    trend_reklamace = None
    try:
        for mesic in range(1, 13):
            chart_odbery[mesic - 1] = Odber.query.filter(
                _db_year_eq(Odber.datum, selected_year), _db_month_eq(Odber.datum, mesic)
            ).count()
            chart_reklamace[mesic - 1] = Reklamace.query.filter(
                _db_year_eq(Reklamace.datum_prijmu, selected_year), _db_month_eq(Reklamace.datum_prijmu, mesic),
                Reklamace.archived == False
            ).count()
        # Trend vs minulý rok (pro summary karty)
        prev_year = selected_year - 1
        prev_odber = Odber.query.filter(_db_year_eq(Odber.datum, prev_year)).count()
        prev_rekl = Reklamace.query.filter(_db_year_eq(Reklamace.datum_prijmu, prev_year), Reklamace.archived == False).count()
        celkem_o = celkove_statistiky.get('celkem_odberu', 0) or 0
        celkem_r = celkove_statistiky.get('celkem_reklamaci', 0) or 0
        if prev_odber and celkem_o is not None:
            trend_odber = round((celkem_o - prev_odber) / prev_odber * 100, 2)
        if prev_rekl and celkem_r is not None:
            trend_reklamace = round((celkem_r - prev_rekl) / prev_rekl * 100, 2)
    except Exception as e:
        app.logger.error(f'Chyba při načítání měsíčních dat pro grafy: {str(e)}')

    # Fallback: pokud na serveru chybí admin_base_sneat.html, použij TailAdmin
    _sneat_path = os.path.join(app.root_path, 'templates', 'admin_base_sneat.html')
    admin_base = 'admin_base_sneat.html' if os.path.isfile(_sneat_path) else 'admin_base_tailadmin.html'

    try:
        return render_template(
            'admin_dashboard.html',
            admin_base=admin_base,
            prehled=prehled,
            akce=historie,
            reklamace_prehled=reklamace_prehled,
            statistiky=celkove_statistiky,
            selected_year=selected_year,
            is_admin=current_user.is_authenticated and current_user.is_admin(),
            chart_odbery=chart_odbery,
            chart_reklamace=chart_reklamace,
            trend_odber=trend_odber,
            trend_reklamace=trend_reklamace,
        )
    except Exception as e:
        app.logger.error(f'Chyba při renderování admin dashboard: {str(e)}')
        import traceback
        app.logger.error(traceback.format_exc())
        flash(f'Chyba při načítání dashboardu: {str(e)}', 'danger')
        return render_template('admin_dashboard.html',
            admin_base=admin_base,
            prehled=[],
            akce=[],
            reklamace_prehled=[],
            statistiky={},
            selected_year=date.today().year,
            is_admin=False,
            chart_odbery=[0]*12,
            chart_reklamace=[0]*12,
            trend_odber=None,
            trend_reklamace=None,
        )


@app.route('/admin/statistiky')
@login_required
def admin_statistiky():
    """Detailní statistiky a přehledy pro admina."""
    if not (current_user.is_authenticated and current_user.is_admin()):
        flash('Nemáte oprávnění!', 'danger')
        return redirect(url_for('index'))
    
    try:
        # Filtry podle roku a měsíce
        selected_year = request.args.get('rok', str(date.today().year))
        selected_month = request.args.get('mesic', '')
        selected_pobocka = request.args.get('pobocka', '')
        
        try:
            selected_year = int(selected_year)
        except (ValueError, TypeError):
            selected_year = date.today().year
        
        try:
            selected_month = int(selected_month) if selected_month else None
        except (ValueError, TypeError):
            selected_month = None
        
        try:
            selected_pobocka = int(selected_pobocka) if selected_pobocka else None
        except (ValueError, TypeError):
            selected_pobocka = None
        
        # Načtení poboček
        pobocky = Pobocka.query.all()
        pobocky_dict = {p.id: p.nazev for p in pobocky}
        app.logger.debug(f'Načteno {len(pobocky)} poboček, rok: {selected_year}, měsíc: {selected_month}, pobočka: {selected_pobocka}')
        
        # Základní query pro odběry a reklamace (SQLite kompatibilní)
        # Filtrujeme pouze záznamy s platným datem
        # Použijeme strftime pro porovnání roku (vrací string, takže porovnáváme se stringem)
        # Debug: zkontrolujeme, kolik záznamů máme celkem
        total_odbery = Odber.query.count()
        total_reklamace = Reklamace.query.count()
        app.logger.debug(f'Celkem v DB: {total_odbery} odběrů, {total_reklamace} reklamací')
        
        odbery_query = Odber.query.filter(
            Odber.datum.isnot(None),
            _db_year_eq(Odber.datum, selected_year)
        )
        reklamace_query = Reklamace.query.filter(
            Reklamace.datum_prijmu.isnot(None),
            _db_year_eq(Reklamace.datum_prijmu, selected_year)
        )
        
        # Debug: zkontrolujeme, kolik záznamů máme po filtrování
        odbery_count_after_filter = odbery_query.count()
        reklamace_count_after_filter = reklamace_query.count()
        app.logger.debug(f'Po filtrování roku {selected_year}: {odbery_count_after_filter} odběrů, {reklamace_count_after_filter} reklamací')
        
        if selected_month:
            odbery_query = odbery_query.filter(_db_month_eq(Odber.datum, selected_month))
            reklamace_query = reklamace_query.filter(_db_month_eq(Reklamace.datum_prijmu, selected_month))
        
        if selected_pobocka:
            odbery_query = odbery_query.filter_by(pobocka_id=selected_pobocka)
            reklamace_query = reklamace_query.filter_by(pobocka_id=selected_pobocka)
        
        # Měsíční statistiky - inicializujeme všechny měsíce
        mesicni_odbery = {i: {'celkem': 0, 'aktivni': 0, 'vydano': 0, 'castka': 0} for i in range(1, 13)}
        mesicni_reklamace = {i: {'celkem': 0, 'ceka': 0, 'vymena': 0, 'poslano': 0, 'zamitnuto': 0, 'cena': 0} for i in range(1, 13)}
        for mesic in range(1, 13):
            odbery_mesic = Odber.query.filter(
                Odber.datum.isnot(None),
                _db_year_eq(Odber.datum, selected_year),
                _db_month_eq(Odber.datum, mesic)
            )
            if selected_pobocka:
                odbery_mesic = odbery_mesic.filter_by(pobocka_id=selected_pobocka)
            
            reklamace_mesic = Reklamace.query.filter(
                Reklamace.datum_prijmu.isnot(None),
                _db_year_eq(Reklamace.datum_prijmu, selected_year),
                _db_month_eq(Reklamace.datum_prijmu, mesic)
            )
            if selected_pobocka:
                reklamace_mesic = reklamace_mesic.filter_by(pobocka_id=selected_pobocka)
            
            try:
                odbery_count = odbery_mesic.count()
                odbery_vydano_list = odbery_mesic.filter_by(stav='vydáno').all()
                mesicni_odbery[mesic] = {
                    'celkem': odbery_count or 0,
                    'aktivni': odbery_mesic.filter_by(stav='aktivní').count() or 0,
                    'vydano': len(odbery_vydano_list) or 0,
                    'castka': sum(o.castka or 0 for o in odbery_vydano_list) or 0
                }
            except Exception as e:
                app.logger.error(f'Chyba při načítání měsíčních odběrů pro měsíc {mesic}: {str(e)}')
                import traceback
                app.logger.error(traceback.format_exc())
                mesicni_odbery[mesic] = {'celkem': 0, 'aktivni': 0, 'vydano': 0, 'castka': 0}
            
            try:
                reklamace_count = reklamace_mesic.count()
                reklamace_all_list = reklamace_mesic.all()
                mesicni_reklamace[mesic] = {
                    'celkem': reklamace_count or 0,
                    'ceka': reklamace_mesic.filter_by(stav='Čeká').count() or 0,
                    'vymena': reklamace_mesic.filter_by(stav='Výměna kus za kus').count() or 0,
                    'poslano': reklamace_mesic.filter_by(stav='Posláno do Ústí').count() or 0,
                    'zamitnuto': reklamace_mesic.filter_by(stav='Zamítnuto').count() or 0,
                    'cena': sum(r.cena or 0 for r in reklamace_all_list) or 0
                }
            except Exception as e:
                app.logger.error(f'Chyba při načítání měsíčních reklamací pro měsíc {mesic}: {str(e)}')
                import traceback
                app.logger.error(traceback.format_exc())
                mesicni_reklamace[mesic] = {'celkem': 0, 'ceka': 0, 'vymena': 0, 'poslano': 0, 'zamitnuto': 0, 'cena': 0}
        
        # Statistiky podle poboček
        pobocky_stats = []
        for pobocka in pobocky:
            if selected_pobocka and pobocka.id != selected_pobocka:
                continue
            
            try:
                odbery_pob = odbery_query.filter_by(pobocka_id=pobocka.id)
                reklamace_pob = reklamace_query.filter_by(pobocka_id=pobocka.id)
                
                odbery_pob_vydano = odbery_pob.filter_by(stav='vydáno').all()
                reklamace_pob_all = reklamace_pob.all()
                
                pobocky_stats.append({
                    'id': pobocka.id,
                    'nazev': pobocka.nazev,
                    'odbery': {
                        'celkem': odbery_pob.count() or 0,
                        'aktivni': odbery_pob.filter_by(stav='aktivní').count() or 0,
                        'vydano': len(odbery_pob_vydano) or 0,
                        'castka': sum(o.castka or 0 for o in odbery_pob_vydano) or 0
                    },
                    'reklamace': {
                        'celkem': reklamace_pob.count() or 0,
                        'ceka': reklamace_pob.filter_by(stav='Čeká').count() or 0,
                        'vymena': reklamace_pob.filter_by(stav='Výměna kus za kus').count() or 0,
                        'poslano': reklamace_pob.filter_by(stav='Posláno do Ústí').count() or 0,
                        'zamitnuto': reklamace_pob.filter_by(stav='Zamítnuto').count() or 0,
                        'cena': sum(r.cena or 0 for r in reklamace_pob_all) or 0
                    }
                })
            except Exception as e:
                app.logger.error(f'Chyba při načítání statistik pro pobočku {pobocka.nazev}: {str(e)}')
                pobocky_stats.append({
                    'id': pobocka.id,
                    'nazev': pobocka.nazev,
                    'odbery': {'celkem': 0, 'aktivni': 0, 'vydano': 0, 'castka': 0},
                    'reklamace': {'celkem': 0, 'ceka': 0, 'vymena': 0, 'poslano': 0, 'zamitnuto': 0, 'cena': 0}
                })
        
        # Celkové statistiky
        try:
            odbery_count_total = odbery_query.count()
            odbery_vydano = odbery_query.filter_by(stav='vydáno').all()
            odbery_castka = sum(o.castka or 0 for o in odbery_vydano)
            app.logger.debug(f'Celkem odběrů: {odbery_count_total}, vydáno: {len(odbery_vydano)}, částka: {odbery_castka}')
        except Exception as e:
            app.logger.error(f'Chyba při načítání celkových statistik odběrů: {str(e)}')
            odbery_castka = 0
        
        try:
            reklamace_count_total = reklamace_query.count()
            reklamace_all = reklamace_query.all()
            reklamace_cena = sum(r.cena or 0 for r in reklamace_all)
            app.logger.debug(f'Celkem reklamací: {reklamace_count_total}, cena: {reklamace_cena}')
        except Exception as e:
            app.logger.error(f'Chyba při načítání celkových statistik reklamací: {str(e)}')
            reklamace_cena = 0
        
        celkove_stats = {
            'odbery': {
                'celkem': odbery_query.count() or 0,
                'aktivni': odbery_query.filter_by(stav='aktivní').count() or 0,
                'vydano': odbery_query.filter_by(stav='vydáno').count() or 0,
                'nevyzvednuto': odbery_query.filter_by(stav='nevyzvednuto').count() or 0,
                'smazano': odbery_query.filter_by(stav='smazano').count() or 0,
                'castka': odbery_castka
            },
            'reklamace': {
                'celkem': reklamace_query.count() or 0,
                'ceka': reklamace_query.filter_by(stav='Čeká').count() or 0,
                'vymena': reklamace_query.filter_by(stav='Výměna kus za kus').count() or 0,
                'poslano': reklamace_query.filter_by(stav='Posláno do Ústí').count() or 0,
                'zamitnuto': reklamace_query.filter_by(stav='Zamítnuto').count() or 0,
                'cena': reklamace_cena,
                'zavolano': reklamace_query.filter_by(zavolano_zakaznikovi=True).count() or 0
            }
        }
        
        # Top zákazníci (podle počtu odběrů)
        try:
            top_zakaznici = db.session.query(
                Odber.jmeno,
                db.func.count(Odber.id).label('pocet'),
                db.func.sum(Odber.castka).label('celkem')
            ).filter(
                Odber.datum.isnot(None),
                _db_year_eq(Odber.datum, selected_year)
            )
            if selected_pobocka:
                top_zakaznici = top_zakaznici.filter_by(pobocka_id=selected_pobocka)
            top_zakaznici = top_zakaznici.group_by(Odber.jmeno).order_by(db.func.count(Odber.id).desc()).limit(10).all()
        except Exception as e:
            app.logger.error(f'Chyba při načítání top zákazníků: {str(e)}')
            top_zakaznici = []
        
        # Top značky reklamací
        try:
            top_znacky = db.session.query(
                Reklamace.znacka,
                db.func.count(Reklamace.id).label('pocet')
            ).filter(
                Reklamace.datum_prijmu.isnot(None),
                _db_year_eq(Reklamace.datum_prijmu, selected_year)
            )
            if selected_pobocka:
                top_znacky = top_znacky.filter_by(pobocka_id=selected_pobocka)
            top_znacky = top_znacky.group_by(Reklamace.znacka).order_by(db.func.count(Reklamace.id).desc()).limit(10).all()
        except Exception as e:
            app.logger.error(f'Chyba při načítání top značek: {str(e)}')
            top_znacky = []
        
        return render_template(
            'admin_statistiky.html',
            selected_year=selected_year,
            selected_month=selected_month,
            selected_pobocka=selected_pobocka,
            pobocky=pobocky,
            mesicni_odbery=mesicni_odbery,
            mesicni_reklamace=mesicni_reklamace,
            pobocky_stats=pobocky_stats,
            celkove_stats=celkove_stats,
            top_zakaznici=top_zakaznici,
            top_znacky=top_znacky,
            current_year=date.today().year,
            is_admin=True
        )
    except Exception as e:
        app.logger.error(f'Chyba v admin statistiky: {str(e)}')
        import traceback
        app.logger.error(traceback.format_exc())
        flash(f'Chyba při načítání statistik: {str(e)}', 'danger')
        # Vytvoříme prázdné struktury pro šablonu
        empty_mesicni_odbery = {i: {'celkem': 0, 'aktivni': 0, 'vydano': 0, 'castka': 0} for i in range(1, 13)}
        empty_mesicni_reklamace = {i: {'celkem': 0, 'ceka': 0, 'vymena': 0, 'poslano': 0, 'zamitnuto': 0, 'cena': 0} for i in range(1, 13)}
        return render_template('admin_statistiky.html',
            selected_year=date.today().year,
            selected_month=None,
            selected_pobocka=None,
            pobocky=[],
            mesicni_odbery=empty_mesicni_odbery,
            mesicni_reklamace=empty_mesicni_reklamace,
            pobocky_stats=[],
            celkove_stats={
                'odbery': {'celkem': 0, 'aktivni': 0, 'vydano': 0, 'nevyzvednuto': 0, 'castka': 0},
                'reklamace': {'celkem': 0, 'ceka': 0, 'vymena': 0, 'poslano': 0, 'zamitnuto': 0, 'cena': 0}
            },
            top_zakaznici=[],
            top_znacky=[],
            current_year=date.today().year,
            is_admin=True
        )


@app.route('/admin/users', methods=['GET', 'POST'])
@login_required
def admin_users():
    """Samostatná stránka – uživatelé (seznam + přidat)."""
    if not (current_user.is_authenticated and current_user.is_admin()):
        flash('Nemáte oprávnění!', 'danger')
        return redirect(url_for('index'))
    user_form = AddUserForm()
    all_pobocky = Pobocka.query.all()
    user_form.pobocky.choices = [(str(p.id), p.nazev) for p in all_pobocky]
    users = User.query.all()

    if user_form.validate_on_submit() and 'jmeno' in request.form:
        jmeno_clean = (user_form.jmeno.data or '').strip()[:100]
        pin_clean = (user_form.pin.data or '').strip()
        username_clean = jmeno_clean.lower().replace(' ', '_')[:100]
        existing_pin = User.query.filter_by(pin=pin_clean).first()
        if existing_pin:
            flash('PIN již existuje!', 'danger')
        elif User.query.filter_by(username=username_clean).first():
            flash('Uživatel s tímto jménem již existuje!', 'danger')
        else:
            user = User(username=username_clean, pin=pin_clean, jmeno=jmeno_clean, role=user_form.role.data)
            user.set_password(pin_clean)
            pobocky_data = request.form.getlist('pobocky') or user_form.pobocky.data or []
            if pobocky_data:
                pobocky_ids = []
                for p in pobocky_data:
                    if not p:
                        continue
                    try:
                        pid = int(p)
                        if db.session.get(Pobocka, pid):
                            pobocky_ids.append(pid)
                    except (ValueError, TypeError):
                        continue
                if pobocky_ids:
                    user.pobocky = Pobocka.query.filter(Pobocka.id.in_(pobocky_ids)).all()
                    user.pobocka_id = user.pobocky[0].id
            try:
                db.session.add(user)
                db.session.commit()
                akce = Akce(odber_id=None, uzivatel=current_user.username, akce=f'Přidán uživatel: {user.jmeno}', datum=get_current_time(), pobocka_id=_system_pobocka_id())
                db.session.add(akce)
                db.session.commit()
                flash('Uživatel přidán!', 'success')
                return redirect(url_for('admin_users'))
            except Exception as e:
                db.session.rollback()
                flash(f'Chyba: {str(e)}', 'danger')

    _sneat_path = os.path.join(app.root_path, 'templates', 'admin_base_sneat.html')
    admin_base = 'admin_base_sneat.html' if os.path.isfile(_sneat_path) else 'admin_base_tailadmin.html'
    return render_template('admin_users.html', admin_base=admin_base, user_form=user_form, users=users, all_pobocky=all_pobocky)


@app.route('/admin/pobocky', methods=['GET', 'POST'])
@login_required
def admin_pobocky():
    """Samostatná stránka – pobočky (seznam + přidat)."""
    if not (current_user.is_authenticated and current_user.is_admin()):
        flash('Nemáte oprávnění!', 'danger')
        return redirect(url_for('index'))
    pobocka_form = AddPobockaForm()
    pobocky = Pobocka.query.all()

    if pobocka_form.validate_on_submit() and 'nazev' in request.form:
        pobocka = Pobocka(nazev=pobocka_form.nazev.data)
        db.session.add(pobocka)
        db.session.commit()
        akce = Akce(odber_id=None, uzivatel=current_user.username, akce=f'Přidána pobočka: {pobocka.nazev}', datum=get_current_time(), pobocka_id=pobocka.id)
        db.session.add(akce)
        db.session.commit()
        flash('Pobočka přidána!', 'success')
        return redirect(url_for('admin_pobocky'))

    _sneat_path = os.path.join(app.root_path, 'templates', 'admin_base_sneat.html')
    admin_base = 'admin_base_sneat.html' if os.path.isfile(_sneat_path) else 'admin_base_tailadmin.html'
    return render_template('admin_pobocky.html', admin_base=admin_base, pobocka_form=pobocka_form, pobocky=pobocky)


@app.route('/admin/user/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_user(id):
    """Editace uživatele."""
    if not (current_user.is_authenticated and current_user.is_admin()):
        flash('Nemáte oprávnění!', 'danger')
        return redirect(url_for('admin_users'))
    
    user = User.query.get_or_404(id)
    form = EditUserForm()
    
    # Naplníme choices pro pobočky (pro zpětnou kompatibilitu)
    all_pobocky = Pobocka.query.all()
    form.pobocky.choices = [(str(p.id), p.nazev) for p in all_pobocky]
    
    # Získáme ID poboček uživatele pro checkboxy
    user_pobocky_ids = [p.id for p in user.pobocky]
    
    if request.method == 'GET':
        form.jmeno.data = user.jmeno
        form.pin.data = user.pin
        form.role.data = user.role
        form.pobocky.data = [str(p.id) for p in user.pobocky]
    
    if form.validate_on_submit():
        # Sanitizace vstupů
        jmeno_clean = form.jmeno.data.strip()[:100] if form.jmeno.data else user.jmeno
        if not jmeno_clean or len(jmeno_clean) < 2:
            flash('Jméno musí mít minimálně 2 znaky!', 'danger')
            return render_template('admin_edit_user_checkboxes.html', form=form, user=user, all_pobocky=all_pobocky, user_pobocky_ids=user_pobocky_ids)
        
        user.jmeno = jmeno_clean
        
        if form.pin.data:
            pin_clean = form.pin.data.strip()
            if pin_clean:
                # Kontrola unikátnosti PINu (kromě aktuálního uživatele)
                existing_pin = User.query.filter(User.pin == pin_clean, User.id != id).first()
                if existing_pin:
                    flash('PIN již existuje u jiného uživatele!', 'danger')
                    return render_template('admin_edit_user_checkboxes.html', form=form, user=user, all_pobocky=all_pobocky, user_pobocky_ids=user_pobocky_ids)
                user.pin = pin_clean
                user.set_password(pin_clean)  # přihlášení jen přes PIN
        
        user.role = form.role.data
        
        # Aktualizace poboček - použijeme checkboxy z request.form.getlist
        pobocky_data = request.form.getlist('pobocky')
        
        if pobocky_data:
            pobocky_ids = []
            for p_id in pobocky_data:
                if p_id:
                    try:
                        pob_id = int(p_id)
                        # Ověření, že pobočka existuje
                        if db.session.get(Pobocka, pob_id):
                            pobocky_ids.append(pob_id)
                    except (ValueError, TypeError):
                        continue
            
            if pobocky_ids:
                pobocky_objects = Pobocka.query.filter(Pobocka.id.in_(pobocky_ids)).all()
                user.pobocky = pobocky_objects
                app.logger.info(f'Uživatel {user.username} má nyní {len(pobocky_objects)} poboček: {[p.nazev for p in pobocky_objects]}')
                # Zpětná kompatibilita
                if pobocky_objects:
                    user.pobocka_id = pobocky_objects[0].id
            else:
                user.pobocky = []
                user.pobocka_id = None
        else:
            user.pobocky = []
            user.pobocka_id = None

        try:
            db.session.commit()
            akce = Akce(
                odber_id=None,
                uzivatel=current_user.username,
                akce=f'Upraven uživatel: {user.jmeno}',
                datum=get_current_time(),
                pobocka_id=_system_pobocka_id()
            )
            db.session.add(akce)
            db.session.commit()
            flash('Uživatel upraven!', 'success')
            return redirect(url_for('admin_users'))
        except Exception as e:
            db.session.rollback()
            app.logger.error(f'Chyba při ukládání uživatele: {str(e)}')
            flash(f'Chyba při ukládání: {str(e)}', 'danger')
    
    return render_template('admin_edit_user_checkboxes.html', form=form, user=user, all_pobocky=all_pobocky, user_pobocky_ids=user_pobocky_ids)


@app.route('/admin/pobocka/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_pobocka(id):
    """Editace pobočky."""
    if not (current_user.is_authenticated and current_user.is_admin()):
        flash('Nemáte oprávnění!', 'danger')
        return redirect(url_for('admin_pobocky'))
    
    pobocka = Pobocka.query.get_or_404(id)
    form = EditPobockaForm()
    
    if request.method == 'GET':
        form.nazev.data = pobocka.nazev
        form.adresa.data = pobocka.adresa
        form.firma.data = pobocka.firma
    
    if form.validate_on_submit():
        # Sanitizace vstupu
        new_nazev = form.nazev.data.strip()[:100] if form.nazev.data else pobocka.nazev
        
        # Kontrola, zda název už neexistuje (kromě aktuální pobočky)
        existing = Pobocka.query.filter(Pobocka.nazev == new_nazev, Pobocka.id != id).first()
        if existing:
            flash('Pobočka s tímto názvem již existuje!', 'danger')
            return render_template('admin_edit_pobocka.html', form=form, pobocka=pobocka)
        
        old_nazev = pobocka.nazev
        pobocka.nazev = new_nazev
        pobocka.adresa = (form.adresa.data or '').strip()[:200] if form.adresa.data else None
        pobocka.firma = (form.firma.data or '').strip()[:200] if form.firma.data else None
        
        try:
            db.session.commit()
            akce = Akce(
                odber_id=None,
                uzivatel=current_user.username,
                akce=f'Upravena pobočka: {old_nazev} → {pobocka.nazev}',
                datum=get_current_time(),
                pobocka_id=pobocka.id
            )
            db.session.add(akce)
            db.session.commit()
            flash('Pobočka upravena!', 'success')
            return redirect(url_for('admin_pobocky'))
        except Exception as e:
            db.session.rollback()
            app.logger.error(f'Chyba při ukládání pobočky: {str(e)}')
            flash(f'Chyba při ukládání: {str(e)}', 'danger')
    
    return render_template('admin_edit_pobocka.html', form=form, pobocka=pobocka)


@app.route('/delete_user/<int:id>', methods=['POST'])
@login_required
def delete_user(id):
    if not (current_user.is_authenticated and current_user.is_admin()):
        flash('Nemáte oprávnění!', 'danger')
        return redirect(url_for('admin_dashboard'))
    user = User.query.get_or_404(id)
    if user.id == current_user.id:
        flash('Nemůžete smazat sami sebe!', 'danger')
        return redirect(url_for('admin_dashboard'))
    akce = Akce(
        odber_id=None,
        uzivatel=current_user.username,
        akce=f'Smazán uživatel: {user.jmeno or user.username}',
        datum=get_current_time(),
        pobocka_id=_system_pobocka_id()
    )
    db.session.add(akce)
    db.session.delete(user)
    db.session.commit()
    flash('Uživatel smazán!', 'success')
    return redirect(url_for('admin_users'))


@app.route('/delete_pobocka/<int:id>', methods=['POST'])
@login_required
def delete_pobocka(id):
    """Smazání pobočky."""
    if not (current_user.is_authenticated and current_user.is_admin()):
        flash('Nemáte oprávnění!', 'danger')
        return redirect(url_for('admin_pobocky'))
    
    pobocka = Pobocka.query.get_or_404(id)
    
    # Kontrola, jestli pobočka nemá žádné odběry nebo reklamace
    odbery_count = Odber.query.filter_by(pobocka_id=id).count()
    reklamace_count = Reklamace.query.filter_by(pobocka_id=id).count()
    
    if odbery_count > 0 or reklamace_count > 0:
        flash(f'Nelze smazat pobočku! Má {odbery_count} odběrů a {reklamace_count} reklamací.', 'danger')
        return redirect(url_for('admin_pobocky'))
    
    nazev = pobocka.nazev
    db.session.delete(pobocka)
    db.session.commit()
    
    akce = Akce(
        odber_id=None,
        uzivatel=current_user.username,
        akce=f'Smazána pobočka: {nazev}',
        datum=get_current_time(),
        pobocka_id=_system_pobocka_id()
    )
    db.session.add(akce)
    db.session.commit()
    flash('Pobočka smazána!', 'success')
    return redirect(url_for('admin_pobocky'))

@app.route('/logout')
def logout():
    if current_user.is_authenticated:
        logout_user()
        flash('Byli jste odhlášeni.', 'info')
    return redirect(url_for('index'))


# ---------- PPL Sklad (zásilky) ----------
import sqlite3 as _sqlite3
from datetime import timedelta

# Nenapipané zásilky při inventuře se označí ke smazání; fyzicky se smažou po tomto počtu dní
PPL_INVENTURA_DAYS_UNTIL_DELETE = 7

# Složka pro PPL DB – pokud je složka s app.py read-only, nastavte PPL_DATA_DIR na zapisovatelnou cestu
_ppl_dir = os.environ.get('PPL_DATA_DIR', '').strip()
if _ppl_dir:
    if not os.path.isdir(_ppl_dir):
        try:
            os.makedirs(_ppl_dir, exist_ok=True)
        except OSError:
            _ppl_dir = ''
    if not _ppl_dir:
        _ppl_dir = os.path.dirname(os.path.abspath(__file__))
else:
    _ppl_dir = os.path.dirname(os.path.abspath(__file__))
PPL_DATABASE = os.path.join(_ppl_dir, 'ppl_warehouse.db')
PPL_HISTORY_DB = os.path.join(_ppl_dir, 'ppl_history.db')
# Čekání na zámek (sekundy) – při "database is locked" SQLite počká místo okamžité chyby
PPL_DB_TIMEOUT = 15

def _ppl_connect(db_path, timeout=PPL_DB_TIMEOUT):
    conn = _sqlite3.connect(db_path, timeout=float(timeout))
    try:
        conn.execute('PRAGMA journal_mode=WAL')
    except Exception:
        pass
    return conn

def _ppl_conn(db_path=PPL_DATABASE):
    """Připojí se k PPL DB. Pokud soubor nebo tabulky chybí, vytvoří nové (_ppl_init)."""
    need_init = (
        (db_path == PPL_DATABASE and not os.path.isfile(PPL_DATABASE)) or
        (db_path == PPL_HISTORY_DB and not os.path.isfile(PPL_HISTORY_DB))
    )
    if need_init:
        _ppl_init()
    try:
        conn = _ppl_connect(db_path)
    except _sqlite3.OperationalError:
        _ppl_init()
        conn = _ppl_connect(db_path)
    if db_path == PPL_DATABASE:
        try:
            conn.execute('SELECT 1 FROM parcels LIMIT 1')
        except _sqlite3.OperationalError:
            conn.close()
            _ppl_init()
            conn = _ppl_connect(db_path)
    return conn

def _ppl_init():
    """Vytvoří PPL DB soubory a tabulky, pokud chybí. Bezpečné volat opakovaně."""
    dir_ = os.path.dirname(PPL_DATABASE)
    if dir_:
        try:
            os.makedirs(dir_, exist_ok=True)
        except OSError:
            pass
    conn = _ppl_connect(PPL_DATABASE)
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS parcels (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            pobocka_id INTEGER NOT NULL DEFAULT 1,
            parcel_code TEXT NOT NULL,
            last_four_digits TEXT NOT NULL,
            shelf TEXT NOT NULL,
            notes TEXT,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    for col, typ in [
        ('pobocka_id', 'INTEGER NOT NULL DEFAULT 1'),
        ('released_at', 'DATETIME'),
        ('k_smazani_po', 'DATE'),
    ]:
        try:
            conn.execute('ALTER TABLE parcels ADD COLUMN ' + col + ' ' + typ)
            conn.commit()
        except _sqlite3.OperationalError:
            conn.rollback()
    conn.close()
    conn = _ppl_connect(PPL_DATABASE)
    conn.execute('CREATE TABLE IF NOT EXISTS inventura_dates (pobocka_id INTEGER PRIMARY KEY, last_date DATE)')
    conn.execute('''CREATE TABLE IF NOT EXISTS inventura_session (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        pobocka_id INTEGER NOT NULL,
        parcel_code TEXT NOT NULL,
        shelf TEXT NOT NULL,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')
    conn.commit()
    conn.close()
    dir_h = os.path.dirname(PPL_HISTORY_DB)
    if dir_h:
        try:
            os.makedirs(dir_h, exist_ok=True)
        except OSError:
            pass
    conn = _ppl_connect(PPL_HISTORY_DB)
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            pobocka_id INTEGER,
            action TEXT NOT NULL,
            parcel_code TEXT,
            shelf TEXT,
            notes TEXT,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    try:
        conn.execute('ALTER TABLE history ADD COLUMN pobocka_id INTEGER')
        conn.commit()
    except _sqlite3.OperationalError:
        pass
    conn.close()

def _ppl_log(action, parcel_code=None, shelf=None, notes=None, pobocka_id=None):
    if not os.path.isfile(PPL_HISTORY_DB):
        _ppl_init()
    conn = _ppl_connect(PPL_HISTORY_DB)
    try:
        conn.execute('INSERT INTO history (action, parcel_code, shelf, notes, pobocka_id) VALUES (?,?,?,?,?)',
                     (action, parcel_code, shelf, notes, pobocka_id))
    except _sqlite3.OperationalError:
        conn.execute('INSERT INTO history (action, parcel_code, shelf, notes) VALUES (?,?,?,?)',
                     (action, parcel_code, shelf, notes))
    conn.commit()
    conn.close()

try:
    _ppl_init()
except Exception:
    pass

# Validace PPL: 41498792168-54302 nebo 41498792168 – vždy poslední 4 číslice z první části (2168)
# Některé skenery přidají ke 11místnému číslu kontrolní cifru → 12 číslic; tu odstraníme.
def _ppl_validate_code(raw):
    raw = (raw or '').strip().replace(' ', '')
    if not raw:
        return None, 'Zadejte číslo zásilky.'
    if '-' in raw:
        parts = raw.split('-', 1)
        if len(parts) != 2 or not parts[0].isdigit() or not parts[1].isdigit():
            return None, 'Formát: číslice-číslice (např. 41498792168-54302).'
        code = f'{parts[0]}-{parts[1]}'
        last_four = parts[0][-4:] if len(parts[0]) >= 4 else parts[0]
    else:
        if not raw.isdigit() or len(raw) < 4:
            return None, 'Číslo zásilky musí mít alespoň 4 číslice (nebo formát 41498792168-54302).'
        # 12 číslic bez pomlčky = 11místné číslo + kontrolní cifra ze skeneru → bereme prvních 11
        if len(raw) == 12:
            raw = raw[:11]
        code = raw
        last_four = raw[-4:]
    return (code, last_four), None


@app.route('/ppl')
@login_required
def ppl_index():
    """PPL – výběr pobočky nebo přesměrování na sklad pobočky."""
    pobocky = get_user_pobocky()
    if not pobocky:
        flash('Nemáte přiřazenu žádnou pobočku.', 'warning')
        return redirect(url_for('index'))
    if len(pobocky) == 1:
        return redirect(url_for('ppl_branch', pobocka_id=pobocky[0].id))
    return render_template('ppl/vyber_pobocky.html', pobocky=pobocky)


def _ppl_inventura_reminder(pobocka_id):
    """Vrátí (last_date, next_due) pro připomínku inventury. next_due = last + 1 měsíc. Když inventura nikdy nebyla, next_due = dnes."""
    conn = _ppl_conn()
    try:
        row = conn.execute('SELECT last_date FROM inventura_dates WHERE pobocka_id = ?', (pobocka_id,)).fetchone()
    except _sqlite3.OperationalError:
        row = None
    conn.close()
    from datetime import timedelta
    today = date.today()
    if not row or not row[0]:
        return None, today.isoformat()
    last = row[0]
    try:
        last_d = date.fromisoformat(last) if isinstance(last, str) else last
        next_due = (last_d + timedelta(days=30)).isoformat()
        return last, next_due
    except Exception:
        return last, today.isoformat()


@app.route('/ppl/<int:pobocka_id>')
@login_required
def ppl_branch(pobocka_id):
    """Sklad PPL pro danou pobočku."""
    pobocka = Pobocka.query.get_or_404(pobocka_id)
    if not current_user.can_access_pobocka(pobocka_id):
        flash('Nemáte přístup k této pobočce.', 'danger')
        return redirect(url_for('ppl_index'))
    today_iso = date.today().isoformat()
    inventura_last_cz = None
    inventura_next_due_cz = None
    inventura_next_due = today_iso
    try:
        inventura_last, inventura_next_due = _ppl_inventura_reminder(pobocka_id)
        inventura_last_cz = _cesky_datum(date.fromisoformat(inventura_last)) if inventura_last else None
        inventura_next_due_cz = _cesky_datum(date.fromisoformat(inventura_next_due)) if inventura_next_due else None
    except Exception as e:
        app.logger.warning('PPL inventura_reminder: %s', e)
    return render_template('ppl/index.html', pobocka=pobocka, inventura_last=inventura_last_cz, inventura_next_due=inventura_next_due_cz, inventura_next_due_iso=inventura_next_due, today_iso=today_iso)


@app.route('/ppl/<int:pobocka_id>/sklad')
@login_required
def ppl_sklad(pobocka_id):
    pobocka = Pobocka.query.get_or_404(pobocka_id)
    if not current_user.can_access_pobocka(pobocka_id):
        flash('Nemáte přístup.', 'danger')
        return redirect(url_for('ppl_index'))
    return render_template('ppl/sklad.html', pobocka=pobocka)


@app.route('/ppl/<int:pobocka_id>/historie')
@login_required
def ppl_historie(pobocka_id):
    pobocka = Pobocka.query.get_or_404(pobocka_id)
    if not current_user.can_access_pobocka(pobocka_id):
        flash('Nemáte přístup.', 'danger')
        return redirect(url_for('ppl_index'))
    return render_template('ppl/historie.html', pobocka=pobocka)


@app.route('/ppl/<int:pobocka_id>/inventura')
@login_required
def ppl_inventura(pobocka_id):
    """Inventura – pipání balíků a polic; nenapipané se označí ke smazání za 1 měsíc."""
    pobocka = Pobocka.query.get_or_404(pobocka_id)
    if not current_user.can_access_pobocka(pobocka_id):
        flash('Nemáte přístup.', 'danger')
        return redirect(url_for('ppl_index'))
    return render_template('ppl/inventura.html', pobocka=pobocka)


def _ppl_normalize_shelf(shelf):
    """Jednotná normalizace umístění pro porovnávání: velká písmena, jedna mezera."""
    return re.sub(r'\s+', ' ', (shelf or '').strip()).upper()


def _ppl_canonical_parcel(raw_code):
    """Z libovolného tvaru kódu (70182912688 nebo 70182912688-54302) vrátí (base_11, last4).
    Vždy pracujeme s 11místným základem – oba tvary = jedna zásilka."""
    raw = (raw_code or '').strip().replace(' ', '')
    if not raw or not raw.replace('-', '').isdigit():
        return None, None
    if '-' in raw:
        base = raw.split('-', 1)[0].strip()
    else:
        base = raw
    if not base.isdigit():
        return None, None
    if len(base) == 12:
        base = base[:11]
    elif len(base) > 11:
        base = base[:11]
    if len(base) < 4:
        return base, base
    return base, base[-4:]


def _ppl_inventura_verified_set(scanned_pairs):
    """Z napipaných dvojic (zásilka + umístění) vrátí set pro rychlé hledání.
    Pouze zásilky V TOMTO SETU zůstanou. Používáme kanonický tvar: vždy base_11 (70182912688),
    takže 70182912688 i 70182912688-54302 = stejná zásilka."""
    out = set()
    for p in (scanned_pairs or []):
        raw = (p.get('parcel_code') or p.get('parcel') or '').strip().replace(' ', '')
        shelf = _ppl_normalize_shelf(p.get('shelf') or p.get('police'))
        if not raw or not shelf:
            continue
        base_11, last4 = _ppl_canonical_parcel(raw)
        if not base_11:
            continue
        out.add((base_11, shelf))
        out.add((last4, shelf))
    return out


def _ppl_ensure_inventura_session(conn):
    """Vytvoří tabulku inventura_session, pokud neexistuje (např. stará DB)."""
    try:
        conn.execute('''CREATE TABLE IF NOT EXISTS inventura_session (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            pobocka_id INTEGER NOT NULL,
            parcel_code TEXT NOT NULL,
            shelf TEXT NOT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )''')
        conn.commit()
    except _sqlite3.OperationalError:
        conn.rollback()


@app.route('/ppl/<int:pobocka_id>/api/inventura/pairs', methods=['GET'])
@login_required
def ppl_api_inventura_pairs(pobocka_id):
    """Vrátí napipané dvojice (zásilka + police) z aktuální inventury (uložené po jednom)."""
    if not current_user.can_access_pobocka(pobocka_id):
        return jsonify({'error': 'Přístup odepřen.'}), 403
    conn = _ppl_conn()
    _ppl_ensure_inventura_session(conn)
    try:
        cur = conn.execute(
            'SELECT parcel_code, shelf FROM inventura_session WHERE pobocka_id = ? ORDER BY id',
            (pobocka_id,)
        )
        rows = cur.fetchall()
    except _sqlite3.OperationalError:
        rows = []
    conn.close()
    return jsonify([{'parcel_code': r[0], 'shelf': r[1]} for r in rows])


@app.route('/ppl/<int:pobocka_id>/api/inventura/pair', methods=['POST'])
@login_required
def ppl_api_inventura_pair(pobocka_id):
    """Uloží jednu dvojici (zásilka + police) – napipáno v inventuře. Po uložení lze pipat další."""
    if not current_user.can_access_pobocka(pobocka_id):
        return jsonify({'error': 'Přístup odepřen.'}), 403
    data = request.get_json(silent=True) or {}
    parcel = (data.get('parcel_code') or data.get('parcel') or '').strip().replace(' ', '')
    shelf = (data.get('shelf') or data.get('police') or '').strip()
    if not parcel or not shelf:
        return jsonify({'error': 'Zadejte číslo zásilky a umístění.'}), 400
    if len(parcel) == 12 and parcel.isdigit():
        parcel = parcel[:11]
    conn = _ppl_conn()
    _ppl_ensure_inventura_session(conn)
    try:
        conn.execute(
            'INSERT INTO inventura_session (pobocka_id, parcel_code, shelf) VALUES (?, ?, ?)',
            (pobocka_id, parcel, shelf.upper())
        )
        conn.commit()
    except _sqlite3.OperationalError as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': 'Chyba ukládání.'}), 500
    conn.close()
    return jsonify({'message': 'Ověřeno.', 'parcel_code': parcel, 'shelf': shelf.upper()}), 201


@app.route('/ppl/<int:pobocka_id>/api/inventura/finish', methods=['POST'])
@login_required
def ppl_api_inventura_finish(pobocka_id):
    """Ukončit inventuru: CO JSTE NAPIPAL, ZŮSTANE (zásilka+umístění v verified). CO JSTE NENAPIPAL, označí se ke smazání za 7 dní."""
    if not current_user.can_access_pobocka(pobocka_id):
        return jsonify({'error': 'Přístup odepřen.'}), 403
    data = request.get_json(silent=True) or {}
    scanned_pairs = data.get('scanned_pairs') or []
    if not scanned_pairs:
        conn = _ppl_conn()
        _ppl_ensure_inventura_session(conn)
        try:
            cur = conn.execute('SELECT parcel_code, shelf FROM inventura_session WHERE pobocka_id = ?', (pobocka_id,))
            scanned_pairs = [{'parcel_code': r[0], 'shelf': r[1]} for r in cur.fetchall()]
        except _sqlite3.OperationalError:
            pass
        conn.close()
    verified = _ppl_inventura_verified_set(scanned_pairs)
    conn = _ppl_conn()
    _ppl_ensure_inventura_session(conn)

    # Propis napipaných dvojic (zásilka + umístění) do tabulky parcels – přehled po umístěních zobrazí aktuální police bez refresh
    try:
        cur_all = conn.execute('SELECT id, parcel_code, last_four_digits FROM parcels WHERE pobocka_id = ?', (pobocka_id,))
        all_parcels = cur_all.fetchall()
        for row in all_parcels:
            pid, code, last_four = row[0], (row[1] or '').strip(), (row[2] or '').strip()
            code_base_11, code_last4 = _ppl_canonical_parcel(code)
            if not code_base_11:
                continue
            for pair in (scanned_pairs or []):
                raw = (pair.get('parcel_code') or pair.get('parcel') or '').strip().replace(' ', '')
                shelf_val = _ppl_normalize_shelf((pair.get('shelf') or pair.get('police') or '').strip())
                if not raw or not shelf_val:
                    continue
                p_base, p_last4 = _ppl_canonical_parcel(raw)
                if code_base_11 == p_base or (code_last4 and code_last4 == p_last4):
                    conn.execute('UPDATE parcels SET shelf = ? WHERE id = ? AND pobocka_id = ?', (shelf_val, pid, pobocka_id))
                    break
        conn.commit()
    except Exception as e:
        app.logger.warning('PPL inventura propis polic: %s', e)
        try:
            conn.rollback()
        except Exception:
            pass

    try:
        cur = conn.execute('''
            SELECT id, parcel_code, last_four_digits, shelf FROM parcels
            WHERE pobocka_id = ? AND released_at IS NULL
              AND (k_smazani_po IS NULL OR date(k_smazani_po) > date('now'))
        ''', (pobocka_id,))
    except _sqlite3.OperationalError:
        cur = conn.execute('SELECT id, parcel_code, last_four_digits, shelf FROM parcels WHERE pobocka_id = ?', (pobocka_id,))
    rows = cur.fetchall()
    today = date.today()
    delete_on = today + timedelta(days=PPL_INVENTURA_DAYS_UNTIL_DELETE)
    marked = 0
    for r in rows:
        pid, code, last_four, shelf_raw = r[0], (r[1] or '').strip(), (r[2] or '').strip(), (r[3] or '').strip()
        shelf = _ppl_normalize_shelf(shelf_raw)
        code_base_11, code_last4 = _ppl_canonical_parcel(code)
        if not code_base_11:
            continue
        # Napipané = v verified (stejná kanonická norma: 70182912688 = 70182912688-54302). Nenapipané → označíme ke smazání.
        if (code_base_11, shelf) in verified or (code_last4, shelf) in verified:
            continue
        conn.execute(
            'UPDATE parcels SET k_smazani_po = ? WHERE id = ? AND pobocka_id = ?',
            (delete_on.isoformat(), pid, pobocka_id)
        )
        marked += 1
    conn.commit()
    try:
        conn.execute('INSERT OR REPLACE INTO inventura_dates (pobocka_id, last_date) VALUES (?, ?)', (pobocka_id, today.isoformat()))
        conn.commit()
    except _sqlite3.OperationalError:
        pass
    try:
        conn.execute('DELETE FROM inventura_session WHERE pobocka_id = ?', (pobocka_id,))
        conn.commit()
    except _sqlite3.OperationalError:
        pass
    conn.close()
    kept = len(rows) - marked
    _ppl_log('INVENTURA_FINISH', notes=f'Označeno ke smazání za {PPL_INVENTURA_DAYS_UNTIL_DELETE} dní: {marked} zásilek.', pobocka_id=pobocka_id)
    return jsonify({
        'message': f'Inventura ukončena. {kept} zásilek zůstalo, {marked} označeno ke smazání za {PPL_INVENTURA_DAYS_UNTIL_DELETE} dní.',
        'kept': kept,
        'marked': marked,
        'days': PPL_INVENTURA_DAYS_UNTIL_DELETE
    }), 200


@app.route('/ppl/<int:pobocka_id>/api/parcels/to-delete', methods=['GET'])
@login_required
def ppl_api_parcels_to_delete(pobocka_id):
    """Zásilky označené ke smazání (k_smazani_po v budoucnu)."""
    if not current_user.can_access_pobocka(pobocka_id):
        return jsonify({'error': 'Přístup odepřen.'}), 403
    try:
        conn = _ppl_conn()
        try:
            cur = conn.execute('''
                SELECT id, parcel_code, last_four_digits, shelf, notes, k_smazani_po
                FROM parcels
                WHERE pobocka_id = ? AND released_at IS NULL
                  AND k_smazani_po IS NOT NULL AND date(k_smazani_po) > date('now')
                ORDER BY k_smazani_po, parcel_code
            ''', (pobocka_id,))
            rows = cur.fetchall()
        except _sqlite3.OperationalError:
            conn.close()
            try:
                _ppl_init()
            except Exception:
                pass
            return jsonify([])
        conn.close()
        return jsonify([{'id': r[0], 'parcel_code': r[1], 'last_four': r[2], 'shelf': r[3], 'notes': r[4] or '', 'k_smazani_po': r[5]} for r in rows])
    except Exception as e:
        app.logger.warning('PPL GET to-delete: %s', e)
        return jsonify([])


@app.route('/ppl/<int:pobocka_id>/api/parcels/<int:parcel_id>/cancel-delete', methods=['POST'])
@login_required
def ppl_api_parcel_cancel_delete(pobocka_id, parcel_id):
    """Zruší označení ke smazání (vymaže k_smazani_po)."""
    if not current_user.can_access_pobocka(pobocka_id):
        return jsonify({'error': 'Přístup odepřen.'}), 403
    try:
        conn = _ppl_conn()
        cur = conn.execute('UPDATE parcels SET k_smazani_po = NULL WHERE id = ? AND pobocka_id = ?', (parcel_id, pobocka_id))
        conn.commit()
        conn.close()
        if cur.rowcount:
            return jsonify({'message': 'Označení ke smazání zrušeno.'}), 200
        return jsonify({'error': 'Zásilka nenalezena nebo není označena ke smazání.'}), 404
    except _sqlite3.OperationalError:
        return jsonify({'error': 'Zásilka nenalezena nebo není označena ke smazání.'}), 404


@app.route('/ppl/<int:pobocka_id>/api/parcels/<int:parcel_id>/mark-delete', methods=['POST'])
@login_required
def ppl_api_parcel_mark_delete(pobocka_id, parcel_id):
    """Označí zásilku ke smazání (nastaví k_smazani_po)."""
    if not current_user.can_access_pobocka(pobocka_id):
        return jsonify({'error': 'Přístup odepřen.'}), 403
    try:
        conn = _ppl_conn()
        row = conn.execute(
            'SELECT id FROM parcels WHERE id = ? AND pobocka_id = ? AND released_at IS NULL',
            (parcel_id, pobocka_id),
        ).fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Zásilka nenalezena nebo již vydána.'}), 404
        delete_on = date.today() + timedelta(days=PPL_INVENTURA_DAYS_UNTIL_DELETE)
        conn.execute(
            'UPDATE parcels SET k_smazani_po = ? WHERE id = ? AND pobocka_id = ?',
            (delete_on.isoformat(), parcel_id, pobocka_id),
        )
        conn.commit()
        conn.close()
        return jsonify({'message': 'Zásilka označena ke smazání.'}), 200
    except Exception as e:
        app.logger.warning('PPL mark-delete: %s', e)
        return jsonify({'error': 'Chyba při označení.'}), 500


@app.route('/ppl/<int:pobocka_id>/api/parcels', methods=['GET'])
@login_required
def ppl_api_get_parcels(pobocka_id):
    if not current_user.can_access_pobocka(pobocka_id):
        return jsonify({'error': 'Přístup odepřen.'}), 403
    try:
        conn = _ppl_conn()
        rows = _ppl_fetch_parcels_rows(conn, pobocka_id)
        conn.close()
        out = []
        for r in rows:
            if len(r) >= 6:
                out.append({'id': r[0], 'parcel_code': r[1], 'last_four': r[2], 'shelf': r[3], 'notes': r[4] or '', 'timestamp': r[5]})
            elif len(r) >= 4:
                out.append({'id': r[0], 'parcel_code': r[1], 'last_four': r[2], 'shelf': r[3], 'notes': '', 'timestamp': None})
            else:
                out.append({'id': r[0], 'parcel_code': r[1], 'last_four': '', 'shelf': r[2], 'notes': '', 'timestamp': None})
        return jsonify(out)
    except Exception as e:
        app.logger.warning('PPL GET parcels: %s', e)
        return jsonify([])


@app.route('/ppl/<int:pobocka_id>/api/parcels', methods=['POST'])
@login_required
def ppl_api_add_parcel(pobocka_id):
    if not current_user.can_access_pobocka(pobocka_id):
        return jsonify({'error': 'Přístup odepřen.'}), 403
    data = request.get_json(silent=True) or {}
    raw = (data.get('parcelCode') or data.get('parcel_code') or '').strip()
    shelf = (data.get('shelf') or '').strip()
    notes = (data.get('notes') or '').strip()
    result, err = _ppl_validate_code(raw)
    if err:
        return jsonify({'error': err}), 400
    parcel_code, last_four = result
    try:
        _ppl_init()
        conn = _ppl_conn()
        conn.execute(
            'INSERT INTO parcels (pobocka_id, parcel_code, last_four_digits, shelf, notes) VALUES (?,?,?,?,?)',
            (pobocka_id, parcel_code, last_four, shelf, notes)
        )
        conn.commit()
        conn.close()
        _ppl_log('ADD', parcel_code=parcel_code, shelf=shelf, notes=notes, pobocka_id=pobocka_id)
        return jsonify({'message': 'Zásilka přidána.', 'parcel_code': parcel_code, 'last_four': last_four}), 201
    except Exception as e:
        import traceback
        app.logger.error(f'PPL add parcel: {e}\n{traceback.format_exc()}')
        err_msg = str(e) if e else 'Chyba při ukládání.'
        return jsonify({'error': 'Chyba při ukládání.', 'detail': err_msg}), 500


@app.route('/ppl/<int:pobocka_id>/api/parcels/<int:parcel_id>', methods=['PUT'])
@login_required
def ppl_api_update_parcel(pobocka_id, parcel_id):
    if not current_user.can_access_pobocka(pobocka_id):
        return jsonify({'error': 'Přístup odepřen.'}), 403
    data = request.get_json(silent=True) or {}
    notes = (data.get('notes') or '').strip()
    conn = _ppl_conn()
    cur = conn.execute('UPDATE parcels SET notes = ? WHERE id = ? AND pobocka_id = ?', (notes, parcel_id, pobocka_id))
    conn.commit()
    conn.close()
    if cur.rowcount:
        _ppl_log('UPDATE_NOTES', parcel_code=str(parcel_id), notes=notes, pobocka_id=pobocka_id)
        return jsonify({'message': 'Poznámky upraveny.'}), 200
    return jsonify({'error': 'Zásilka nenalezena.'}), 404


@app.route('/ppl/<int:pobocka_id>/api/parcels/<int:parcel_id>/vydat', methods=['POST'])
@login_required
def ppl_api_vydat_parcel(pobocka_id, parcel_id):
    """Označit zásilku jako vydanou zákazníkovi – zmizí z aktivního skladu."""
    if not current_user.can_access_pobocka(pobocka_id):
        return jsonify({'error': 'Přístup odepřen.'}), 403
    conn = _ppl_conn()
    row = conn.execute('SELECT parcel_code, shelf FROM parcels WHERE id = ? AND pobocka_id = ? AND (released_at IS NULL OR released_at = "")', (parcel_id, pobocka_id)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Zásilka nenalezena nebo již vydána.'}), 404
    now = get_current_time().strftime('%Y-%m-%d %H:%M:%S')
    conn.execute('UPDATE parcels SET released_at = ? WHERE id = ? AND pobocka_id = ?', (now, parcel_id, pobocka_id))
    conn.commit()
    conn.close()
    _ppl_log('VYDANO', parcel_code=row[0], shelf=row[1], pobocka_id=pobocka_id)
    return jsonify({'message': 'Zásilka vydána.'}), 200


@app.route('/ppl/<int:pobocka_id>/api/parcels/<int:parcel_id>', methods=['DELETE'])
@login_required
def ppl_api_delete_parcel(pobocka_id, parcel_id):
    if not current_user.can_access_pobocka(pobocka_id):
        return jsonify({'error': 'Přístup odepřen.'}), 403
    conn = _ppl_conn()
    row = conn.execute('SELECT parcel_code, shelf, notes FROM parcels WHERE id = ? AND pobocka_id = ?', (parcel_id, pobocka_id)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Zásilka nenalezena.'}), 404
    conn.execute('DELETE FROM parcels WHERE id = ? AND pobocka_id = ?', (parcel_id, pobocka_id))
    conn.commit()
    conn.close()
    _ppl_log('DELETE', parcel_code=row[0], shelf=row[1], notes=row[2], pobocka_id=pobocka_id)
    return jsonify({'message': 'Zásilka smazána.'}), 200


def _ppl_fetch_parcels_rows(conn, pobocka_id, order_by_shelf=False):
    """Stejná data jako hlavní seznam zásilek: vyčistit prošlé ke smazání, pak SELECT. order_by_shelf=True pro přehled po umístění."""
    try:
        conn.execute('DELETE FROM parcels WHERE pobocka_id = ? AND k_smazani_po IS NOT NULL AND date(k_smazani_po) <= date("now")', (pobocka_id,))
        conn.commit()
    except _sqlite3.OperationalError:
        try:
            conn.rollback()
        except Exception:
            pass
    order = 'ORDER BY shelf, parcel_code' if order_by_shelf else 'ORDER BY timestamp DESC'
    order_fallback = 'ORDER BY shelf, parcel_code' if order_by_shelf else 'ORDER BY parcel_code'
    try:
        cur = conn.execute(
            'SELECT id, parcel_code, last_four_digits, shelf, notes, timestamp FROM parcels '
            'WHERE pobocka_id = ? AND released_at IS NULL AND (k_smazani_po IS NULL OR date(k_smazani_po) > date(\'now\')) ' + order,
            (pobocka_id,),
        )
        return cur.fetchall()
    except _sqlite3.OperationalError:
        try:
            cur = conn.execute(
                'SELECT id, parcel_code, last_four_digits, shelf, notes, timestamp FROM parcels WHERE pobocka_id = ? ' + order,
                (pobocka_id,),
            )
            return cur.fetchall()
        except _sqlite3.OperationalError:
            cur = conn.execute(
                'SELECT id, parcel_code, shelf FROM parcels WHERE pobocka_id = ? ' + order_fallback,
                (pobocka_id,),
            )
            return cur.fetchall()


@app.route('/ppl/<int:pobocka_id>/api/stock', methods=['GET'])
@login_required
def ppl_api_stock(pobocka_id):
    """Přehled po umístěních: stejná data jako hlavní seznam, seskupená podle shelf."""
    if not current_user.can_access_pobocka(pobocka_id):
        return jsonify({'error': 'Přístup odepřen.'}), 403
    try:
        conn = _ppl_conn()
        rows = _ppl_fetch_parcels_rows(conn, pobocka_id, order_by_shelf=True)
        conn.close()
        by_shelf = {}
        for r in rows:
            if len(r) >= 4:
                sid, code, last_four, shelf = r[0], (r[1] or '').strip(), (r[2] or '').strip(), (r[3] or '').strip()
            else:
                sid, code, shelf = r[0], (r[1] or '').strip(), (r[2] or '').strip()
                last_four = ''
            if shelf not in by_shelf:
                by_shelf[shelf] = []
            by_shelf[shelf].append({'id': sid, 'parcel_code': code, 'last_four': last_four})
        def shelf_sort_key(s):
            if not s:
                return (0, '')
            m = re.match(r'^([A-Za-z]*)(\d*)$', s.strip())
            a, n = (m.group(1) or '').upper(), m.group(2) or '0'
            return (a, int(n) if n.isdigit() else 0)
        shelves_sorted = sorted(by_shelf.keys(), key=shelf_sort_key)
        total = sum(len(by_shelf[s]) for s in shelves_sorted)
        result = {
            'shelves': [{'shelf': s, 'count': len(by_shelf[s]), 'parcels': by_shelf[s]} for s in shelves_sorted],
            'total': total,
        }
        return jsonify(result)
    except Exception as e:
        app.logger.warning('PPL api/stock: %s', e)
        return jsonify({'shelves': [], 'total': 0})


@app.route('/ppl/<int:pobocka_id>/api/history', methods=['GET'])
@login_required
def ppl_api_history(pobocka_id):
    if not current_user.can_access_pobocka(pobocka_id):
        return jsonify({'error': 'Přístup odepřen.'}), 403
    conn = _ppl_conn(PPL_HISTORY_DB)
    try:
        cur = conn.execute(
            'SELECT id, action, parcel_code, shelf, notes, timestamp FROM history WHERE pobocka_id = ? ORDER BY timestamp DESC LIMIT 500',
            (pobocka_id,)
        )
    except _sqlite3.OperationalError:
        cur = conn.execute('SELECT id, action, parcel_code, shelf, notes, timestamp FROM history ORDER BY timestamp DESC LIMIT 500')
    rows = cur.fetchall()
    conn.close()
    return jsonify([{'id': r[0], 'action': r[1], 'parcel_code': r[2], 'shelf': r[3], 'notes': r[4], 'timestamp': r[5]} for r in rows])


# Error handlers
@app.errorhandler(404)
def not_found_error(error):
    """Handler pro 404 chyby – nikdy nesmí znovu vyhodit."""
    try:
        app.logger.warning('404: %s', request.url)
        return render_template('error.html', error_code=404, error_message='Stránka nenalezena'), 404
    except Exception:
        return (
            '<!DOCTYPE html><html><head><meta charset="utf-8"><title>404</title></head>'
            '<body><h1>404</h1><p>Stránka nenalezena.</p><p><a href="/">Domů</a></p></body></html>',
            404,
            {'Content-Type': 'text/html; charset=utf-8'}
        )


@app.errorhandler(500)
def internal_error(error):
    """Handler pro 500 chyby – nikdy nesmí znovu vyhodit výjimku."""
    try:
        db.session.rollback()
    except Exception:
        pass
    try:
        app.logger.error('500 chyba: %s', str(error), exc_info=True)
    except Exception:
        pass
    try:
        return render_template('error.html',
                               error_code=500,
                               error_message='Vnitřní chyba serveru'), 500
    except Exception as e:
        return (
            '<!DOCTYPE html><html><head><meta charset="utf-8"><title>Chyba 500</title></head>'
            '<body><h1>Chyba 500</h1><p>Došlo k vnitřní chybě serveru. Zkuste to později.</p>'
            '<p><a href="/">Zpět na hlavní stránku</a></p></body></html>',
            500,
            {'Content-Type': 'text/html; charset=utf-8'}
        )


@app.errorhandler(400)
def bad_request_error(error):
    """Handler pro 400 (např. neplatný CSRF nebo špatná data) – nikdy nesmí znovu vyhodit."""
    try:
        app.logger.warning('400: %s', request.url)
        return render_template('error.html', error_code=400, error_message='Neplatný požadavek (např. vypršela session – obnovte stránku a zkuste znovu)'), 400
    except Exception:
        return (
            '<!DOCTYPE html><html><head><meta charset="utf-8"><title>400</title></head>'
            '<body><h1>400</h1><p>Neplatný požadavek. Obnovte stránku (F5) a zkuste znovu.</p><p><a href="/">Domů</a></p></body></html>',
            400,
            {'Content-Type': 'text/html; charset=utf-8'}
        )


@app.errorhandler(403)
def forbidden_error(error):
    """Handler pro 403 chyby – nikdy nesmí znovu vyhodit."""
    try:
        app.logger.warning('403: %s', request.url)
        return render_template('error.html', error_code=403, error_message='Přístup zamítnut'), 403
    except Exception:
        return (
            '<!DOCTYPE html><html><head><meta charset="utf-8"><title>403</title></head>'
            '<body><h1>403</h1><p>Přístup zamítnut.</p><p><a href="/">Domů</a></p></body></html>',
            403,
            {'Content-Type': 'text/html; charset=utf-8'}
        )


# Health check endpoint pro monitoring
@app.route('/health')
def health_check():
    """Health check endpoint pro monitoring aplikace."""
    try:
        # Test databázového připojení
        db.session.execute(db.text('SELECT 1'))
        return jsonify({
            'status': 'healthy',
            'database': 'connected',
            'timestamp': datetime.now().isoformat()
        }), 200
    except Exception as e:
        app.logger.error(f'Health check failed: {str(e)}')
        return jsonify({
            'status': 'unhealthy',
            'database': 'disconnected',
            'error': str(e),
            'timestamp': datetime.now().isoformat()
        }), 503


if __name__ == '__main__':
    # Stabilní spuštění: host 0.0.0.0 (přístup z sítě), port z env, bez debugu
    host = os.environ.get('FLASK_RUN_HOST', '0.0.0.0')
    port = int(os.environ.get('PORT', os.environ.get('FLASK_RUN_PORT', '8080')))
    app.run(host=host, port=port, debug=False, threaded=True)